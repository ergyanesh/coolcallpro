"""
State Hub Page Generator — Programmatic SEO
Reads states.xlsx + cities.xlsx and generates template-driven HTML pages.
Every factual claim comes from the spreadsheet data. No improvised prose.
"""

import json
import openpyxl
from collections import defaultdict
from datetime import date
from html import escape as _html_escape
from pathlib import Path
import os
import sys

# ============================================================
# IMAGE ATTRIBUTIONS (loaded once)
# ============================================================
_ATTRIBUTIONS_PATH = Path(__file__).parent / "images" / "_attributions.json"
try:
    _IMAGE_ATTRIBUTIONS = json.loads(_ATTRIBUTIONS_PATH.read_text(encoding="utf-8"))
except (FileNotFoundError, json.JSONDecodeError):
    _IMAGE_ATTRIBUTIONS = {}


# ============================================================
# STATE-HUB EXTRA SECTIONS
# Sections inserted between Cost Overview and Service Areas on every state
# hub: Common HVAC Issues (climate-seasonal cluster links) + Climate Hazards
# (FEMA NRI event-driven HVAC impacts). Each builder returns PURE content
# ('    <!-- comment -->\\n    <section>...\\n    </section>', no leading or
# trailing whitespace). The caller joins non-empty blocks with blank lines.
# Mirrors scripts/add_state_hub_common_issues.py + add_state_hub_climate_hazards.py
# so regenerated state hubs stay consistent with the live-patched files.
# ============================================================

_COMMON_ISSUES_SUMMER = [
    ("../articles/complete-ac-troubleshooting-guide",
     "Complete AC troubleshooting guide",
     "the full diagnostic framework for every AC failure mode"),
    ("../articles/ac-not-cooling-below-80",
     "AC running but not cooling below 80&deg;?",
     "low refrigerant, weak capacitor, dirty coil, or undersized system"),
    ("../articles/ac-circuit-breaker-trips",
     "AC circuit breaker keeps tripping",
     "failing compressor, weak capacitor, or dirty condenser coil"),
    ("../articles/ac-freezing-up-in-summer",
     "AC freezing up in summer",
     "airflow restriction or refrigerant problem"),
    ("../articles/water-dripping-from-vent",
     "Water dripping from ceiling vent",
     "clogged condensate drain or frozen-coil thaw"),
    ("../articles/ac-contactor-clicking",
     "AC contactor clicking but nothing happens",
     "electrical fault in the outdoor unit"),
]

_COMMON_ISSUES_WINTER = [
    ("../article-furnace",
     "Furnace not igniting?",
     "ignition failure diagnosis and repair costs"),
    ("../articles/furnace-blowing-cold-air-winter",
     "Furnace blowing cold air in winter",
     "filter, ignitor, flame sensor, or gas valve fault"),
    ("../article-carbon-monoxide",
     "Carbon monoxide: the invisible killer",
     "CO detection, warning signs, and safety steps"),
    ("../article-winter-storm",
     "Protect your HVAC during a winter storm",
     "freeze prep, power-outage safeguards, post-storm inspection"),
    ("../article-heat-pump",
     "Heat pump not working?",
     "cold-weather performance, defrost cycle, common failures"),
    ("../article-maintenance",
     "12-month HVAC maintenance checklist",
     "seasonal tune-up timing and pro-only tasks"),
]

_COMMON_ISSUES_BOTH = [
    ("../articles/complete-ac-troubleshooting-guide",
     "Complete AC troubleshooting guide",
     "diagnosis for every AC failure mode"),
    ("../article-furnace",
     "Furnace not igniting?",
     "ignition failure diagnosis and repair costs"),
    ("../article-heat-pump",
     "Heat pump not working?",
     "year-round heat-pump performance and repairs"),
    ("../articles/ac-circuit-breaker-trips",
     "AC circuit breaker keeps tripping",
     "electrical fault in the outdoor unit"),
    ("../article-maintenance",
     "12-month HVAC maintenance checklist",
     "seasonal tune-ups for both cooling and heating"),
    ("../articles/2026-hvac-cost-guide",
     "Honest 2026 HVAC cost guide",
     "diagnostic, repair, and replacement pricing"),
]

_COMMON_ISSUES_INTRO = {
    "Summer": "Cooling is the dominant HVAC demand in {state}. The most common emergency and troubleshooting topics for a cooling-driven climate:",
    "Winter": "{state}'s long heating season drives most HVAC calls. Common furnace and heating failures we see:",
    "Both": "{state} sees both cooling and heating demand year-round. Common HVAC troubleshooting topics for a mixed-demand climate:",
}


def common_hvac_issues_section_html(state_name, peak):
    """Render the Common HVAC Issues <section> for a state hub. Peak demand
    season comes from states.xlsx. Returns '' for any peak value outside
    Summer/Winter/Both (defensive; all 51 state rows should be one of the
    three)."""
    peak_key = str(peak or "").strip()
    if peak_key == "Summer":
        links = _COMMON_ISSUES_SUMMER
    elif peak_key == "Winter":
        links = _COMMON_ISSUES_WINTER
    elif peak_key == "Both":
        links = _COMMON_ISSUES_BOTH
    else:
        return ""
    intro = _COMMON_ISSUES_INTRO[peak_key].format(state=state_name)
    lis = "\n".join(
        f'            <li><a href="{href}" style="color: var(--orange-dark); font-weight: 600;">{label}</a> &mdash; {desc}</li>'
        for href, label, desc in links
    )
    return (
        "    <!-- Common HVAC Issues -->\n"
        "    <section class=\"section\" style=\"padding: 0;\">\n"
        "      <div class=\"container\">\n"
        "        <div class=\"city-services\" style=\"margin-bottom: 40px;\">\n"
        f"          <h2>Common HVAC Issues in {state_name}</h2>\n"
        f"          <p style=\"font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);\">{intro}</p>\n"
        "          <ul>\n"
        f"{lis}\n"
        "          </ul>\n"
        "        </div>\n"
        "      </div>\n"
        "    </section>"
    )


# Hazard vocabulary (name -> impact + pro-action copy). Per-state hazard
# lists live in states.xlsx "Climate Hazards" column. Sources:
#   - https://hazards.fema.gov/nri/ (FEMA National Risk Index)
#   - https://www.ncei.noaa.gov/access/billions/ (NOAA NCEI)
_HAZARDS_PATH = Path(__file__).parent / "climate_hazards.json"
try:
    _CLIMATE_HAZARD_VOCAB = json.loads(_HAZARDS_PATH.read_text(encoding="utf-8"))["hazards"]
except (FileNotFoundError, json.JSONDecodeError, KeyError):
    _CLIMATE_HAZARD_VOCAB = {}


def _format_hazard_list_prose(hazards):
    strong = [f"<strong>{h}</strong>" for h in hazards]
    if len(strong) == 1:
        return strong[0]
    if len(strong) == 2:
        return f"{strong[0]} and {strong[1]}"
    return ", ".join(strong[:-1]) + f", and {strong[-1]}"


def climate_hazards_section_html(state_name, hazards_csv):
    """Render the Climate Hazards <section> for a state hub. Returns '' if no
    hazard data is present for the state, so generated pages degrade cleanly
    when xlsx hasn't been updated yet."""
    if not hazards_csv or str(hazards_csv).strip().lower() in ("", "nan"):
        return ""
    hazards = [h.strip() for h in str(hazards_csv).split(",") if h.strip()]
    if not hazards:
        return ""
    intro_list = _format_hazard_list_prose(hazards)
    lis = "\n".join(
        f'            <li><strong>{hz}:</strong> {_CLIMATE_HAZARD_VOCAB[hz]["impact"]} {_CLIMATE_HAZARD_VOCAB[hz]["action"]}</li>'
        for hz in hazards
        if hz in _CLIMATE_HAZARD_VOCAB
    )
    unknown = [hz for hz in hazards if hz not in _CLIMATE_HAZARD_VOCAB]
    if unknown:
        raise KeyError(f"Unknown climate hazard(s) {unknown} for state '{state_name}' "
                       f"-- add them to climate_hazards.json")
    return (
        "    <!-- Climate Hazards -->\n"
        "    <section class=\"section\" style=\"padding: 0;\">\n"
        "      <div class=\"container\">\n"
        "        <div class=\"city-services\" style=\"margin-bottom: 40px;\">\n"
        f"          <h2>Climate Hazards Affecting {state_name} HVAC Systems</h2>\n"
        f"          <p style=\"font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);\">According to the FEMA National Risk Index and NOAA storm data, {state_name} faces elevated exposure to {intro_list}. Each event damages or stresses residential HVAC systems in specific ways that an HVAC professional should diagnose after the fact.</p>\n"
        "          <ul>\n"
        f"{lis}\n"
        "          </ul>\n"
        "          <p style=\"font-size: 0.9rem; color: var(--gray-700); margin-top: 20px;\">Source: <a href=\"https://www.fema.gov/flood-maps/products-tools/national-risk-index\" target=\"_blank\" rel=\"nofollow noopener\" style=\"color: var(--orange-dark);\">FEMA National Risk Index</a> and <a href=\"https://www.noaa.gov/climate\" target=\"_blank\" rel=\"nofollow noopener\" style=\"color: var(--orange-dark);\">NOAA climate data</a>.</p>\n"
        "        </div>\n"
        "      </div>\n"
        "    </section>"
    )


def why_pro_section_html(d):
    """Build the "Why Call a {State} HVAC Pro" section from xlsx row data.

    All four bullet facts come from states.xlsx (Summer High, Winter Low,
    CDD, HDD, SEER2 min+region, License Requirement, Licensing Body/Agency,
    Climate Hazards). The FEMA NRI citation and the 2023-01-01 federal
    efficiency-standard effective date are independently-verifiable public
    facts. No invented data.

    This function mirrors build_block() in
    scripts/replace_state_hub_how_it_works.py so regenerated state hubs
    stay byte-identical to the patched output."""
    name = d["State Name"]
    sh = d.get("Avg Summer High (F)")
    wl = d.get("Avg Winter Low (F)")
    cdd = d.get("Annual Cooling Degree Days")
    hdd = d.get("Annual Heating Degree Days")
    seer_min = d.get("SEER2 Minimum")
    seer_region = str(d.get("SEER2 Region", "") or "").strip()
    req = str(d.get("State HVAC License Requirement", "") or "").strip()
    agency = str(d.get("Licensing Body/Agency", "") or "").strip()
    hazards_csv = str(d.get("Climate Hazards", "") or "").strip()
    hazards = [h.strip() for h in hazards_csv.split(",") if h.strip()]

    def _n(v):
        if v is None:
            return ""
        try:
            return f"{int(float(v)):,}"
        except (TypeError, ValueError):
            return str(v)

    if req.lower().startswith("yes"):
        lic_p = (
            f"{name} requires HVAC contractors to hold a state license from the "
            f"<strong>{agency}</strong>. Unlicensed work can void manufacturer "
            f"warranties and complicate insurance claims."
        )
    else:
        lic_p = (
            f"{name} does not issue a statewide HVAC contractor license; licensing "
            f"and permits are handled by local cities and counties. Hire a contractor "
            f"who carries liability insurance, pulls permits with your local building "
            f"department, and can provide local references."
        )

    sh_s = _n(sh)
    wl_s = _n(wl)
    cdd_s = _n(cdd)
    hdd_s = _n(hdd)
    try:
        seer_min_s = f"{float(seer_min):g}"
    except (TypeError, ValueError):
        seer_min_s = str(seer_min)
    hz1, hz2, hz3 = (hazards + ["", "", ""])[:3]

    return (
        f"    <!-- Why Call a {name} Pro -->\n"
        f"    <section class=\"section section-dark\" id=\"why-pro\">\n"
        f"      <div class=\"container\">\n"
        f"        <div class=\"section-header\">\n"
        f"          <span class=\"section-tag\"\n"
        f"            style=\"background: rgba(255,165,0,0.15); color: var(--orange-light); border-color: rgba(255,165,0,0.3);\">&#128205;\n"
        f"            {name} HVAC Context</span>\n"
        f"          <h2 class=\"section-title\">Why Call a {name} HVAC Pro</h2>\n"
        f"          <p class=\"section-sub\">Four facts from climate and licensing data specific to <strong>{name}</strong>.</p>\n"
        f"        </div>\n"
        f"\n"
        f"        <div class=\"steps-grid\">\n"
        f"\n"
        f"          <div class=\"step-card\">\n"
        f"            <div class=\"step-number\">01</div>\n"
        f"            <div class=\"step-icon\">&#127777;&#65039;</div>\n"
        f"            <h3>Climate Load</h3>\n"
        f"            <p>{name} averages <strong>{sh_s}&deg;F</strong> summer highs and <strong>{wl_s}&deg;F</strong> winter lows, with <strong>{cdd_s}</strong> annual cooling degree days and <strong>{hdd_s}</strong> heating degree days. A mis-sized or poorly-charged system fails early under that load.</p>\n"
        f"          </div>\n"
        f"\n"
        f"          <div class=\"step-card\">\n"
        f"            <div class=\"step-number\">02</div>\n"
        f"            <div class=\"step-icon\">&#128221;</div>\n"
        f"            <h3>Licensing</h3>\n"
        f"            <p>{lic_p}</p>\n"
        f"          </div>\n"
        f"\n"
        f"          <div class=\"step-card\">\n"
        f"            <div class=\"step-number\">03</div>\n"
        f"            <div class=\"step-icon\">&#128161;</div>\n"
        f"            <h3>Efficiency Code</h3>\n"
        f"            <p>New central AC units installed in {name} must meet <strong>SEER2 {seer_min_s}</strong> (<strong>{seer_region}</strong> SEER2 region) under the federal efficiency standard effective January 1, 2023. A pro ensures the installed system is compliant and sized correctly.</p>\n"
        f"          </div>\n"
        f"\n"
        f"          <div class=\"step-card\">\n"
        f"            <div class=\"step-number\">04</div>\n"
        f"            <div class=\"step-icon\">&#127744;&#65039;</div>\n"
        f"            <h3>Weather Risks</h3>\n"
        f"            <p>FEMA's National Risk Index identifies <strong>{hz1}</strong>, <strong>{hz2}</strong>, and <strong>{hz3}</strong> as elevated hazards in {name}. A technician inspects outdoor equipment for damage after any major weather event before restart.</p>\n"
        f"          </div>\n"
        f"\n"
        f"        </div>\n"
        f"\n"
        f"        <div style=\"text-align: center; margin-top: 48px;\">\n"
        f"          <a href=\"tel:+18445821795\" class=\"btn btn-primary btn-lg btn-vibrate\"><span class=\"phone-icon\">&#128222;</span> Call Now &#8212; (844) 582-1795</a>\n"
        f"        </div>\n"
        f"      </div>\n"
        f"    </section>"
    )


# State-energy-office vocabulary (agency name + URL per state abbreviation).
# Sourced from NASEO State Energy Offices Directory (fetched 2026-04-21) and
# stored in state_energy_offices.json for auditability. See that file's
# _notes for the 3 URL substitutions vs NASEO's raw data.
_ENERGY_OFFICES_PATH = Path(__file__).parent / "state_energy_offices.json"
try:
    _ENERGY_OFFICES = json.loads(_ENERGY_OFFICES_PATH.read_text(encoding="utf-8"))["offices"]
except (FileNotFoundError, json.JSONDecodeError, KeyError):
    _ENERGY_OFFICES = {}


def _strip_scheme(url: str) -> str:
    """User-friendly display form of a URL: drop https:// and trailing slash."""
    import re as _re_local
    return _re_local.sub(r"^https?://", "", url).rstrip("/")


def state_energy_office_section_html(state_name, abbr):
    """Render the State Energy Office and HVAC Rebate Resources section.
    Returns '' when we have no data for the abbreviation (defensive; all 51
    rows in states.xlsx should resolve via state_energy_offices.json)."""
    entry = _ENERGY_OFFICES.get(abbr)
    if not entry:
        return ""
    office = entry.get("name") or ""
    url = entry.get("url") or ""
    if not office or not url:
        return ""
    display = _strip_scheme(url)
    # DSIRE's programs subdomain (programs.dsireusa.org/system/program?state=XX)
    # blocks direct external-referer links intermittently ("Access denied").
    # Linking to the dsireusa.org homepage is stable -- users enter their ZIP
    # code in the homepage search to reach state-specific program listings.
    dsire_url = "https://www.dsireusa.org/"
    # Detect a federal-fallback URL (only for states whose own energy office
    # is currently unreachable, e.g. Ohio -> energy.gov). Swap the description
    # so we don't call DOE "the state-level energy agency".
    is_federal_fallback = url.rstrip('/').endswith('energy.gov')
    if is_federal_fallback:
        office_desc = (f"The federal energy agency. {state_name}'s own state-level administrator "
                       "is currently unreachable online; the DOE homepage links to national "
                       "HVAC rebate programs, IRA administration resources, and weatherization "
                       "assistance information.")
    else:
        office_desc = (f"The state-level energy agency that coordinates HVAC rebates, "
                       f"weatherization assistance, and IRA program administration in {state_name}.")
    return (
        "    <!-- State Energy Office Resources -->\n"
        "    <section class=\"section\" style=\"padding: 0;\">\n"
        "      <div class=\"container\">\n"
        "        <div class=\"city-services\" style=\"margin-bottom: 40px;\">\n"
        f"          <h2>{state_name} Energy Office and HVAC Rebate Resources</h2>\n"
        f"          <p style=\"font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);\">For current HVAC rebates, energy-efficiency incentives, and federal Inflation Reduction Act (IRA) program coordination in {state_name}, consult these authoritative sources:</p>\n"
        "          <ul>\n"
        f"            <li><strong>{office}</strong> &mdash; <a href=\"{url}\" target=\"_blank\" rel=\"nofollow noopener\" style=\"color: var(--orange-dark); font-weight: 600;\">{display}</a>. {office_desc}</li>\n"
        f"            <li><strong>DSIRE {state_name} listing</strong> &mdash; <a href=\"{dsire_url}\" target=\"_blank\" rel=\"nofollow noopener\" style=\"color: var(--orange-dark); font-weight: 600;\">dsireusa.org</a>. N.C. State University's comprehensive database of every federal, state, local, and utility incentive program. Enter your ZIP code on the homepage to see programs available to {state_name} homeowners.</li>\n"
        "            <li><strong>ENERGY STAR Heating &amp; Cooling Products</strong> &mdash; <a href=\"https://www.energystar.gov/products/heating_cooling\" target=\"_blank\" rel=\"nofollow noopener\" style=\"color: var(--orange-dark); font-weight: 600;\">energystar.gov/products/heating_cooling</a>. The federal Section 25C tax credit was terminated for installations after Dec 31, 2025 by the One Big Beautiful Bill Act (Public Law 119-21). State HEAR rebates (rolling out through 2026) and utility rebates remain the primary federal-funded incentives.</li>\n"
        "          </ul>\n"
        "          <p style=\"font-size: 0.9rem; color: var(--gray-700); margin-top: 20px;\">Source: <a href=\"https://www.naseo.org/state-energy-offices\" target=\"_blank\" rel=\"nofollow noopener\" style=\"color: var(--orange-dark);\">NASEO State Energy Offices Directory</a> (National Association of State Energy Officials).</p>\n"
        "        </div>\n"
        "      </div>\n"
        "    </section>"
    )


def _compose_state_hub_extra_sections(*section_blocks):
    """Join non-empty section blocks between Cost Overview and Service Areas.

    Template uses: '    </section>{extra}    <!-- Service Areas -->'
    Return value produces a single blank line before each section plus a
    terminator so the last section's '</section>' flows directly into
    '    <!-- Service Areas -->' with no trailing blank line.

    Cases:
      - No blocks:   return '\\n\\n' (preserves the pre-patch blank line
                     between Cost Overview and Service Areas)
      - Any blocks:  return '\\n\\n<b1>\\n\\n<b2>...\\n' (blank lines between
                     each block, single newline before Service Areas anchor)
    """
    non_empty = [s for s in section_blocks if s]
    if not non_empty:
        return "\n\n"
    return "\n\n" + "\n\n".join(non_empty) + "\n"


def state_figure_html(state_slug, state_name):
    """Render attributed <figure> for a state image. Returns '' if missing."""
    rec = _IMAGE_ATTRIBUTIONS.get(f"state/{state_slug}")
    if not rec:
        return ""
    src = f"../{rec['file']}"
    width = rec.get("width") or 1600
    height = rec.get("height") or 900
    artist = _html_escape(rec.get("artist") or "Unknown photographer")
    license_short = _html_escape(rec.get("license") or "")
    license_url = _html_escape(rec.get("license_url") or "")
    commons_url = _html_escape(rec.get("commons_url") or rec.get("image_source_url") or "")
    description = (rec.get("description") or "").strip()
    if description and len(description) > 12:
        alt_text = description.split(". ")[0].strip()[:140]
        if not alt_text.endswith(('.', '!', '?')):
            alt_text = alt_text + f" — {state_name}"
    else:
        alt_text = f"{state_name} — state landmark photo"
    alt_text = _html_escape(alt_text)
    license_html_inner = (
        f'<a href="{license_url}" target="_blank" rel="nofollow noopener">{license_short}</a>'
        if license_url and license_short else license_short
    )
    return f'''
        <figure class="location-figure" style="margin: 0 auto 40px; max-width: 1100px;">
          <img src="{src}" alt="{alt_text}" width="{width}" height="{height}" loading="lazy" decoding="async" style="width: 100%; height: auto; border-radius: var(--radius); display: block;" />
          <figcaption style="display: inline-block; margin-top: 12px; padding: 8px 14px; background: var(--gray-50); border: 1px solid var(--gray-100); border-radius: 999px; font-size: 0.78rem; color: var(--gray-600); line-height: 1.5;">
            <span aria-hidden="true" style="margin-right: 6px;">&#128247;</span>{artist} &middot; {license_html_inner} via <a href="{commons_url}" target="_blank" rel="nofollow noopener" style="color: var(--gray-700); text-decoration: underline;">Wikimedia Commons</a> &middot; <a href="../image-credits" style="color: var(--gray-700); text-decoration: underline;">credits</a>
          </figcaption>
        </figure>
'''

# Cities with generated pages (xlsx may not be updated yet if file was locked)
GENERATED_CITIES = {
    # Batch 1
    "Los Angeles", "New York City", "San Jose", "Austin", "Jacksonville",
    "Fort Worth", "San Diego", "San Francisco", "Columbus", "Indianapolis",
    "Nashville", "Washington", "Seattle", "El Paso", "Louisville",
    # Batch 2
    "Portland", "Memphis", "Baltimore", "Albuquerque", "Tucson",
    "Fresno", "Sacramento", "Mesa", "Omaha", "Miami",
    "Colorado Springs", "Raleigh", "Virginia Beach", "Oakland",
    "Tulsa", "Cleveland", "Wichita",
    # Batch 3
    "Allentown", "Anchorage", "Augusta", "Baton Rouge", "Boise",
    "Bridgeport", "Brownsville", "Buffalo", "Chattanooga", "Dayton",
    "Des Moines", "Eugene", "Fayetteville", "Fort Wayne", "Glendale",
    "Grand Rapids", "Hartford", "Huntsville", "Irvine", "Knoxville",
    "Laredo", "Lexington", "Lincoln", "Little Rock", "Lubbock",
    "Madison", "Mobile", "Modesto", "Newark", "Peoria",
    "Pittsburgh", "Providence", "Reno", "Richmond", "Salt Lake City",
    "Shreveport", "Spokane", "Springfield", "Syracuse", "Tallahassee",
    "Toledo", "Winston-Salem", "Worcester",
    # Batch 4 — 12 remaining states (9 April 2026)
    "Honolulu", "Charleston", "Columbia", "Greenville", "Jackson",
    "Gulfport", "Portland", "Bangor", "Morgantown",
    "Billings", "Missoula", "Fargo", "Bismarck", "Sioux Falls",
    "Rapid City", "Wilmington", "Burlington", "Cheyenne", "Manchester",
}

# ─── Data Loading ───────────────────────────────────────────────────────

def load_states(path='states.xlsx'):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    states = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        abbr = data['State Abbreviation']
        states[abbr] = data
    return states

def load_cities(path='cities_updated.xlsx'):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    by_state = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        by_state[data['State']].append(data)
    return by_state

# ─── Helper Functions ───────────────────────────────────────────────────

def slug(name):
    """State name to URL slug: 'New York' -> 'new-york'"""
    return name.lower().replace(' ', '-')

def city_file(city_name, abbr):
    """City name + abbr to filename: 'Birmingham', 'AL' -> 'birmingham-al'"""
    return city_name.lower().replace(' ', '-').replace('.', '').replace("'", '') + '-' + abbr.lower()

def fmt_pop_short(pop):
    """5237750 -> '5.2M'"""
    if not pop: return 'N/A'
    m = int(pop) / 1_000_000
    return f"{m:.1f}M" if m >= 1 else f"{int(pop):,}"

def fmt_pop_long(pop):
    """5237750 -> '5.2 million'"""
    if not pop: return 'N/A'
    m = int(pop) / 1_000_000
    return f"{m:.1f} million"

def fmt_zones(zones_str):
    """'2A, 3A' -> 'Zones 2A and 3A'"""
    if not zones_str: return ''
    zones = [z.strip() for z in zones_str.split(',')]
    if len(zones) == 1: return f"Zone {zones[0]}"
    if len(zones) == 2: return f"Zones {zones[0]} and {zones[1]}"
    return f"Zones {', '.join(zones[:-1])}, and {zones[-1]}"

def peak_description(peak, summer_high, winter_low):
    """Generate climate demand sentence from peak season + temps."""
    p = (peak or '').lower()
    sh = int(summer_high) if summer_high else '?'
    wl = int(winter_low) if winter_low else '?'
    if 'dual' in p:
        return (f"With summer highs averaging {sh}&deg;F and winter lows dropping to "
                f"{wl}&deg;F, your HVAC system handles heavy demand in both seasons.")
    elif 'winter' in p:
        return (f"With winter lows dropping to {wl}&deg;F, heating is the primary demand "
                f"on your HVAC system. Summer highs near {sh}&deg;F still require reliable cooling.")
    else:
        return (f"With summer highs averaging {sh}&deg;F, cooling is the primary demand "
                f"on your HVAC system. Winter lows near {wl}&deg;F mean heating needs are "
                f"moderate but still important.")

def system_desc(system_type, state_name):
    """Generate system type sentence from data."""
    if not system_type: return ''
    s = system_type.lower()
    if 'mini-split' in s or 'ductless' in s:
        return (f"Most homes across {state_name} rely on mini-split and ductless systems "
                f"for year-round cooling, with heat pumps providing efficient climate control.")
    if '/' in s:
        return (f"Homes across {state_name} typically use either heat pumps or a central "
                f"AC and gas furnace combination, depending on local climate and energy costs.")
    if 'heat pump' in s and 'central ac' not in s:
        return (f"Most homes across {state_name} rely on heat pumps for year-round comfort, "
                f"providing both efficient cooling and heating.")
    if 'central ac + gas furnace' in s:
        return (f"Most homes across {state_name} use a central AC and gas furnace combination "
                f"to handle both cooling and heating demands.")
    return f"Common HVAC systems across {state_name} include {system_type.lower()}."

def article_link(system_type):
    """Pick relevant article link based on dominant system."""
    s = (system_type or '').lower()
    if 'heat pump' in s:
        return '../article-heat-pump', 'heat pump options'
    if 'furnace' in s or 'gas' in s:
        return '../article-furnace', 'furnace options'
    return '../article-heat-pump', 'heat pump options'


def hero_climate_lead(peak, system, zones):
    """Return the noun-phrase lead clause for the hero subtext (without the
    state possessive prefix). Derived from xlsx fields; no invented claims.

    Subarctic is reserved for states containing IECC Climate Zone 8 in the
    Climate Zones Present column (Alaska only in current data). Zone 7
    alone ("Very Cold" per IECC) does not qualify because states like MI,
    MN, WI have population centers in warmer zones and labeling their
    entire climate "subarctic" would overreach the xlsx data."""
    p = (peak or '').lower()
    s = (system or '').lower()
    z = str(zones or '').lower()
    z_tokens = {t.strip() for t in z.replace(',', ' ').split() if t.strip()}

    if 'mini-split' in s or ('1a' in z_tokens and '2a' not in z_tokens):
        return 'year-round cooling climate'
    if 'winter' in p and '8' in z_tokens:
        return 'subarctic climate and extended heating season'
    if 'winter' in p:
        return 'long heating season'
    if 'summer' in p:
        return 'cooling-dominated climate'
    return 'mixed cooling and heating demand'


def get_state_climate_profile(peak, system, zones):
    """Determine state climate profile from data fields for conditional text."""
    p = (peak or '').lower()
    s = (system or '').lower()
    z = str(zones or '').lower()

    # Tropical: Hawaii (mini-split dominant or Zone 1A only)
    if 'mini-split' in s or ('1a' in z and '2a' not in z):
        return {
            "meta_service": "AC repair, heat pump service, and cooling installation",
            "og_service": "Emergency AC and cooling service available 24/7.",
            "hero_service": "Emergency AC repair, heat pump service, and cooling system installation available 24/7.",
            "cost_type": "ac_tropical",
            "cost_faq_q": "How much does AC replacement cost",
            "services": [
                f"Emergency AC Repair in {{name}}",
                "Central Air Conditioning Installation &amp; Replacement",
                "Heat Pump &amp; Mini-Split Installation &amp; Service",
                "Ductwork Inspection, Cleaning &amp; Mold Prevention",
                "HVAC System Maintenance &amp; Tune-Ups",
                "Indoor Air Quality &amp; Ventilation Solutions",
            ],
        }

    # Subarctic/Heavy winter: Peak=Winter + zones 7 or 8
    if 'winter' in p and ('7' in z or '8' in z):
        return {
            "meta_service": "furnace repair, heating service, and HVAC installation",
            "og_service": "Emergency furnace repair and heating service available 24/7.",
            "hero_service": "Emergency furnace repair, heating service, and HVAC installation available 24/7.",
            "cost_type": "furnace_first",
            "cost_faq_q": "How much does HVAC replacement cost",
            "services": [
                f"Emergency Furnace Repair in {{name}}",
                "High-Efficiency Furnace Installation &amp; Replacement",
                "Boiler Service &amp; Radiant Heating Systems",
                "HVAC System Maintenance &amp; Tune-Ups",
                "Ductwork Inspection, Insulation &amp; Sealing",
                "Indoor Air Quality &amp; Ventilation Solutions",
            ],
        }

    # Winter-dominant: Peak=Winter (but not subarctic)
    if 'winter' in p:
        return {
            "meta_service": "furnace repair, AC service, and HVAC installation",
            "og_service": "Emergency furnace repair and HVAC service available 24/7.",
            "hero_service": "Emergency furnace repair, AC service, and HVAC installation available 24/7.",
            "cost_type": "furnace_first",
            "cost_faq_q": "How much does HVAC replacement cost",
            "services": [
                f"Emergency Furnace Repair in {{name}}",
                "High-Efficiency Furnace Installation &amp; Replacement",
                "Central Air Conditioning Repair &amp; Replacement",
                "Heat Pump Installation &amp; Service",
                "HVAC System Maintenance &amp; Tune-Ups",
                "Ductwork Inspection, Cleaning &amp; Sealing",
                "Indoor Air Quality &amp; Ventilation Solutions",
            ],
        }

    # Summer-dominant: Peak=Summer
    if 'summer' in p:
        return {
            "meta_service": "AC repair, HVAC service, and cooling installation",
            "og_service": "Emergency AC repair and HVAC service available 24/7.",
            "hero_service": "Emergency AC repair, heat pump service, and HVAC installation available 24/7.",
            "cost_type": "ac_first",
            "cost_faq_q": "How much does HVAC replacement cost",
            "services": [
                f"Emergency AC Repair in {{name}}",
                "Central Air Conditioning Installation &amp; Replacement",
                "Heat Pump Installation &amp; Service",
                "Furnace Repair &amp; Heating Service",
                "HVAC System Maintenance &amp; Tune-Ups",
                "Ductwork Inspection, Cleaning &amp; Sealing",
                "Indoor Air Quality &amp; Ventilation Solutions",
            ],
        }

    # Balanced (default): Peak=Both or Dual
    return {
        "meta_service": "AC repair, furnace service, and heat pump installation",
        "og_service": "Emergency AC and furnace service available 24/7.",
        "hero_service": "Emergency AC repair, furnace service, and heat pump installation available 24/7.",
        "cost_type": "both",
        "cost_faq_q": "How much does HVAC replacement cost",
        "services": [
            f"Emergency AC &amp; Furnace Repair in {{name}}",
            "Central Air Conditioning Installation &amp; Replacement",
            "Heat Pump Installation &amp; Service",
            "Furnace Installation &amp; Replacement",
            "HVAC System Maintenance &amp; Tune-Ups",
            "Ductwork Inspection, Cleaning &amp; Sealing",
            "Indoor Air Quality &amp; Ventilation Solutions",
        ],
    }


def state_cost_text(name, ac_cost, furnace_cost, cost_type, html=True):
    """Generate cost paragraph text for state hubs. html=False for schema."""
    b = lambda t: f"<strong>{t}</strong>" if html else t
    base = "Costs vary by city, system size, efficiency rating, and installation complexity."
    if cost_type == "ac_tropical":
        return (f"A standard central AC replacement in {name} typically costs "
                f"{b(ac_cost)}, and mini-split/heat pump installations run "
                f"{b(furnace_cost)}. {base}")
    elif cost_type == "furnace_first":
        return (f"A furnace replacement in {name} typically costs "
                f"{b(furnace_cost)}, and AC replacements run "
                f"{b(ac_cost)}. {base}")
    elif cost_type == "ac_first":
        return (f"A standard central AC replacement in {name} typically costs "
                f"{b(ac_cost)}, and furnace installations run "
                f"{b(furnace_cost)}. {base}")
    else:  # both
        return (f"A standard central AC replacement in {name} typically costs "
                f"{b(ac_cost)}, while furnace installations run "
                f"{b(furnace_cost)}. {base}")

def _url_is_real(url: str) -> bool:
    """True when the License Lookup URL cell holds an actual http(s) link.
    Some xlsx rows carry descriptive placeholders like
    "N/A - check local jurisdictions" -- treat those as no-URL so we don't
    emit an <a href="N/A ..."> that resolves to a 404 on coolcallpro.com."""
    return bool(url) and str(url).lower().startswith(('http://', 'https://'))


def license_html(d):
    """Generate licensing section paragraph.

    Three branches:
      - Yes   : statewide licensing -> cite agency + link if URL is real.
      - Varies: licensing handled locally -> no agency-as-regulator copy
                (the xlsx agency cell for Varies states reads as a
                description, not a name, and pasting it verbatim produces
                garbled prose like "license issued by the No state-level
                HVAC licensing...").
      - No    : (legacy catch-all) -> "does not require ... However, {detail}"
    """
    req = str(d.get('State HVAC License Requirement', '') or '').strip()
    body = d.get('Licensing Body/Agency', '') or ''
    url = d.get('License Lookup URL', '') or ''
    name = d['State Name']

    safety = (' See our <a href="../safety" style="color: var(--orange-dark); '
              'font-weight: 600;">safety tips</a> for more on what to verify before hiring.')

    req_l = req.lower()

    if req_l.startswith('yes'):
        link_tag = (f'<a href="{url}" style="color: var(--orange-dark); font-weight: 600;" '
                    f'target="_blank" rel="noopener">license lookup tool</a>') \
            if _url_is_real(url) else 'their website'
        return (f"{name} requires HVAC contractors to hold a license issued by the "
                f"<strong>{body}</strong>. Before hiring, verify your contractor's credentials "
                f"through the board's {link_tag}. Licensed contractors carry insurance, pull "
                f"permits correctly, and stand behind their work.{safety}")

    if req_l.startswith('varies'):
        # Optional state-level portal link if xlsx provides a real URL
        if _url_is_real(url):
            portal = (f' For partial state-level information, see the '
                      f'<a href="{url}" style="color: var(--orange-dark); font-weight: 600;" '
                      f'target="_blank" rel="noopener">state licensing portal</a>.')
        else:
            portal = ''
        return (f"{name} does not require a statewide HVAC contractor license. "
                f"Licensing and permit rules are set at the city or county level by "
                f"local jurisdictions.{portal} Before hiring, verify your contractor "
                f"carries liability insurance, pulls permits with your local building "
                f"department, and can provide local references.{safety}")

    # Legacy 'No state HVAC license; ...' style (currently unused in xlsx but kept
    # defensively in case future rows use that phrasing).
    detail = req.replace('No state HVAC license; ', '').replace('No state HVAC license;', '').strip()
    if detail and detail[0].isupper():
        detail = detail[0].lower() + detail[1:]
    return (f"{name} does not require a statewide HVAC license. However, {detail}. "
            f"Always verify your contractor carries proper insurance and pulls necessary "
            f"permits with your local building department.{safety}")


def license_faq(d):
    """FAQ answer for licensing question.

    Same three branches as license_html. Content mirrors it but is tighter
    (no safety-tips trailer, since the FAQ is Q+A scoped)."""
    req = str(d.get('State HVAC License Requirement', '') or '').strip()
    body = d.get('Licensing Body/Agency', '') or ''
    url = d.get('License Lookup URL', '') or ''
    name = d['State Name']

    req_l = req.lower()

    if req_l.startswith('yes'):
        link = (f'<a href="{url}" style="color: var(--orange-dark); font-weight: 600;" '
                f'target="_blank" rel="noopener">online license lookup</a>') \
            if _url_is_real(url) else 'their website'
        return (f"{name} requires all HVAC contractors to hold a license issued by the "
                f"<strong>{body}</strong>. You can verify any contractor's credentials through "
                f"the board's {link}.")

    if req_l.startswith('varies'):
        portal = ''
        if _url_is_real(url):
            portal = (f' Some state-level information is available via the '
                      f'<a href="{url}" style="color: var(--orange-dark); font-weight: 600;" '
                      f'target="_blank" rel="noopener">state licensing portal</a>.')
        return (f"{name} does not require a statewide HVAC contractor license; "
                f"licensing and permit rules are set by local cities and counties.{portal} "
                f"Before hiring, verify your contractor carries liability insurance, "
                f"pulls permits with your local building department, and can provide "
                f"local references.")

    # Legacy 'No ...' branch
    detail = req.replace('No state HVAC license; ', '').replace('No state HVAC license;', '').strip()
    if detail and detail[0].isupper():
        detail = detail[0].lower() + detail[1:]
    return (f"{name} does not require a statewide HVAC license. However, {detail}. "
            f"Always verify your contractor carries proper insurance and pulls necessary permits.")

def permits_faq(d, *, html=True):
    """FAQ #6 answer -- "What HVAC permits are required in {state}?".

    Two factual templates keyed off the "State HVAC License Requirement"
    column (36 "Yes" states, 15 "Varies by jurisdiction"). The Yes-state
    answer cites the "Licensing Body/Agency" column verbatim; the Varies
    answer notes that licensing + permits are handled locally.

    html=True returns visible-FAQ HTML (<strong>, &mdash;); html=False
    returns schema-safe plain text with a literal em-dash (U+2014)."""
    req = str(d.get('State HVAC License Requirement', '') or '').strip()
    body = str(d.get('Licensing Body/Agency', '') or '').strip()
    name = d['State Name']
    dash = "&mdash;" if html else "—"
    agency_rendered = f"<strong>{body}</strong>" if html else body

    if req.lower().startswith('yes'):
        return (
            f"HVAC work in {name} typically requires a mechanical or building permit "
            f"from your local city or county building department. Covered work includes central AC "
            f"replacement, furnace installation, refrigerant-line modifications, and duct changes. "
            f"The installer must also hold a state license issued by the {agency_rendered}, "
            f"and in most jurisdictions your technician pulls the permit on your behalf. "
            f"Permit fees, inspection requirements, and submittal rules vary by municipality {dash} "
            f"confirm with your technician before work begins."
        )
    return (
        f"{name} does not impose a statewide HVAC contractor license. Permit and "
        f"licensing rules are set by the local city or county building department where the "
        f"work occurs. Most jurisdictions require a mechanical or building permit for central "
        f"AC replacement, furnace installation, gas-line connections, and duct modifications. "
        f"Ask your HVAC technician to confirm the specific permit requirements for your address "
        f"{dash} they typically pull the permit on your behalf."
    )


def system_faq(d):
    """FAQ answer for system type question."""
    system = d.get('Dominant HVAC System Type', '') or ''
    name = d['State Name']
    art_url, art_text = article_link(system)
    art_tag = f'<a href="{art_url}" style="color: var(--orange-dark); font-weight: 600;">{art_text} guide</a>'

    s = system.lower()
    if 'mini-split' in s or 'ductless' in s:
        return (f"<strong>Mini-split and ductless systems</strong> are the most common across {name}. "
                f"The state's tropical climate means cooling is the primary need year-round, and "
                f"heat pumps provide efficient climate control. Explore our {art_tag} to learn more.")
    if '/' in s:
        return (f"Homes across {name} commonly use either <strong>heat pumps</strong> or "
                f"<strong>central AC with gas furnace</strong> systems, depending on the local "
                f"climate zone and energy costs. Explore our {art_tag} to learn more.")
    if 'heat pump' in s and 'central ac' not in s:
        return (f"<strong>Heat pumps</strong> are the most common system type across {name}. "
                f"The state's climate makes heat pumps ideal &#8212; they provide both cooling "
                f"and heating efficiently. Explore our {art_tag} to learn more.")
    if 'central ac + gas furnace' in s:
        return (f"<strong>Central AC paired with a gas furnace</strong> is the most common "
                f"system type across {name}. This combination handles both cooling and heating "
                f"demands efficiently. Explore our {art_tag} to learn more.")
    return (f"Common HVAC systems across {name} include <strong>{system}</strong>. "
            f"Explore our {art_tag} to learn more.")

def rebate_list_html(rebates_str):
    """Generate rebate <li> items from comma-separated rebate programs."""
    items = []
    if rebates_str:
        for prog in str(rebates_str).split(','):
            prog = prog.strip()
            if prog:
                items.append(f'            <li><strong>{prog}</strong></li>')
    items.append(
        '            <li><strong>Federal Tax Credits</strong> &#8212; may apply for qualifying '
        'high-efficiency installations; check <a href="https://www.energystar.gov/about/'
        'federal-tax-credits" target="_blank" rel="noopener" style="color: var(--orange-dark); '
        'font-weight: 600;">energystar.gov</a> for current amounts and eligibility</li>')
    return '\n'.join(items)

def cities_grid_html(city_list, abbr):
    """Generate city grid cards. Only show cities that have (or will have) their own page."""
    parts = []
    for city in sorted(city_list, key=lambda c: c['City']):
        name = city['City']
        cf = city_file(name, abbr)
        if city.get('City Page Published') == 'Yes' or name in GENERATED_CITIES:
            parts.append(f'''          <a href="{cf}" class="city-grid-card">
            <span class="card-icon">&#128205;</span> {name}, {abbr}
          </a>''')
        # Skip absorbed/removed cities — no Coming Soon entries
    return '\n'.join(parts)

def neighbor_links_html(neighbors_str, hub_abbrs, all_states):
    """Generate neighboring state link elements."""
    if not neighbors_str or 'none' in str(neighbors_str).lower(): return ''
    name_to_abbr = {d['State Name']: a for a, d in all_states.items()}
    parts = []
    for name in str(neighbors_str).split(','):
        name = name.strip()
        if not name: continue
        s = slug(name)
        abbr = name_to_abbr.get(name)
        cls = 'neighbor-link' if abbr in hub_abbrs else 'neighbor-link coming-soon'
        parts.append(f'          <a href="{s}" class="{cls}">&#127758; {name}</a>')
    return '\n'.join(parts)

# ─── HTML Template ──────────────────────────────────────────────────────

def _read_existing_reviewed_date(out_path):
    """Parse the Last-reviewed <time datetime="YYYY-MM-DD"> value from an
    existing state hub file. Returns a date or None if the file does not
    exist or the tag can't be located.

    Used to keep the reviewed date sticky across regenerations: a structural
    or template-only regeneration should not bump the "Last reviewed" signal
    that Google uses as a freshness marker. Pass --refresh-date to main()
    to bypass this and force today's date on output."""
    if not out_path:
        return None
    try:
        path_obj = Path(out_path)
        if not path_obj.exists():
            return None
        text = path_obj.read_text(encoding="utf-8")
    except OSError:
        return None
    import re as _re_local
    match = _re_local.search(
        r'class="meta-line meta-reviewed">Last reviewed\s*<time datetime="(\d{4}-\d{2}-\d{2})">',
        text,
    )
    if not match:
        return None
    try:
        return date.fromisoformat(match.group(1))
    except ValueError:
        return None


def generate_html(d, city_list, hub_abbrs, all_states, out_path=None, refresh_date=False):
    """Generate complete state hub HTML from state data dict.

    out_path / refresh_date control the "Last reviewed" date:
      - refresh_date=True: always use today's date (use when content changed).
      - refresh_date=False (default) + out_path exists: reuse the date
        already in the file. Regeneration does not bump the freshness signal.
      - refresh_date=False + no existing file: fall back to today's date."""

    name = d['State Name']
    abbr = d['State Abbreviation']
    s = slug(name)
    pop_short = fmt_pop_short(d.get('State Population'))
    pop_long = fmt_pop_long(d.get('State Population'))
    homeown = d.get('Homeownership Rate %', '')
    zones = d.get('Climate Zones Present', '')
    zones_fmt = fmt_zones(zones)
    sh = int(d['Avg Summer High (F)']) if d.get('Avg Summer High (F)') else ''
    wl = int(d['Avg Winter Low (F)']) if d.get('Avg Winter Low (F)') else ''
    elec = d.get('Avg Electricity Rate (cents/kWh)', '')
    gas = d.get('Avg Natural Gas Rate (dollars/therm)', '')
    seer_region = d.get('SEER2 Region', '')
    seer_min = d.get('SEER2 Minimum', '')
    ac_cost = d.get('Avg AC Install Cost', '')
    furnace_cost = d.get('Avg Furnace Install Cost', '')
    utilities = d.get('Top 3 Utility Companies', '')
    rebates = d.get('State/Utility Rebate Programs', '')
    system = d.get('Dominant HVAC System Type', '')
    peak = d.get('Peak HVAC Demand Season', '')
    neighbors = d.get('Neighboring States', '')
    hazards_csv = d.get('Climate Hazards', '')

    state_profile = get_state_climate_profile(peak, system, zones)
    hero_lead = hero_climate_lead(peak, system, zones)
    state_figure = state_figure_html(s, name)
    peak_desc = peak_description(peak, sh, wl)
    sys_desc = system_desc(system, name)
    art_url, art_text = article_link(system)
    lic_html = license_html(d)
    lic_faq = license_faq(d)
    sys_faq = system_faq(d)
    permits_faq_html_str = permits_faq(d, html=True)
    permits_faq_schema = permits_faq(d, html=False).replace('"', '\\"')
    why_pro_section = why_pro_section_html(d)
    # Strip HTML for schema (JSON-LD can't have unescaped quotes from HTML tags)
    import re as _re
    lic_faq_schema = _re.sub(r'<[^>]+>', '', lic_faq).replace('"', '\\"')
    sys_faq_schema = _re.sub(r'<[^>]+>', '', sys_faq).replace('"', '\\"').replace('&#8212;', '—')
    rebate_items = rebate_list_html(rebates)
    city_grid = cities_grid_html(city_list, abbr)
    neighbor_html = neighbor_links_html(neighbors, hub_abbrs, all_states)
    state_hub_extra_sections = _compose_state_hub_extra_sections(
        common_hvac_issues_section_html(name, peak),
        climate_hazards_section_html(name, hazards_csv),
        state_energy_office_section_html(name, abbr),
    )
    neighbor_section = f'''
    <!-- Neighboring States -->
    <section class="section" style="padding: 32px 0 48px;">
      <div class="container" style="max-width: 800px;">
        <div class="city-services" style="margin-bottom: 16px;">
          <h2>Neighboring State HVAC Guides</h2>
          <p style="font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);">Also serving homeowners in states bordering {name}. HVAC licensing, SEER2 requirements, and rebate programs differ by state.</p>
        </div>
        <div class="neighbor-links">
{neighbor_html}
        </div>
      </div>
    </section>''' if neighbor_html.strip() else ''

    # Featured-city callout (striking-distance pages, GSC positions 10-30 as of 14 Apr 2026)
    FEATURED_CITIES = {
        "OR": [("Portland", "portland-or")],
        "MO": [("St. Louis", "st-louis-mo")],
        "CA": [("Sacramento", "sacramento-ca")],
        "OK": [("Tulsa", "tulsa-ok")],
    }
    featured_callout = ""
    if abbr in FEATURED_CITIES:
        items = FEATURED_CITIES[abbr]
        links = ", ".join(f'<a href="{slug}" style="color: var(--orange-dark); font-weight: 600;">{city}</a>' for city, slug in items)
        featured_callout = f'''
        <div class="callout-box callout-info" style="margin-bottom: 24px;">
          <div>
            <strong>Featured city:</strong> {links} &mdash; see detailed HVAC pricing, permit offices, utility rebates, and ZIP-level coverage.
          </div>
        </div>'''

    # Byline + Last reviewed (E-E-A-T). Sticky by default across regens:
    # reuse the date already in the file unless refresh_date is set.
    reviewed_date = None if refresh_date else _read_existing_reviewed_date(out_path)
    if reviewed_date is None:
        reviewed_date = date.today()
    reviewed_iso = reviewed_date.isoformat()
    reviewed_readable = reviewed_date.strftime("%B %#d, %Y") if os.name == 'nt' else reviewed_date.strftime("%B %-d, %Y")
    meta_block = (
        f'<div class="article-meta-modern"><div class="meta-line meta-reviewed">Last reviewed <time datetime="{reviewed_iso}">{reviewed_readable}</time></div></div>\n'
        f'        <div class="author-byline">\n'
        f'          <img src="../images/author-gyane.webp" alt="Gyanesh Gulshan" class="author-avatar-sm" loading="lazy" width="48" height="48" />\n'
        f'          <span>By <a href="../author-gyanesh">Gyanesh Gulshan</a> &mdash; Founder, Cool Call Pro</span>\n'
        f'        </div>'
    )

    # Sources & References footer (E-E-A-T)
    sources_block = '''
    <!-- Sources & References -->
    <section class="section" style="padding: 0 0 48px;">
      <div class="container" style="max-width: 800px;">
        <div class="city-sources">
          <h3>Sources &amp; References</h3>
          <ul>
            <li><a href="https://www.energy.gov/energysaver/heat-and-cool" target="_blank" rel="nofollow noopener">U.S. Department of Energy &mdash; Home Heating &amp; Cooling</a></li>
            <li><a href="https://www.energystar.gov/" target="_blank" rel="nofollow noopener">U.S. ENERGY STAR (EPA)</a></li>
            <li><a href="https://www.dsireusa.org/" target="_blank" rel="nofollow noopener">DSIRE Database of State Incentives</a></li>
            <li><a href="https://www.irs.gov/credits-deductions/energy-efficient-home-improvement-credit" target="_blank" rel="nofollow noopener">IRS Section 25C Energy Efficient Home Improvement Credit</a></li>
          </ul>
        </div>
      </div>
    </section>'''

    # Climate-aware cost and services strings
    cost_para = state_cost_text(name, ac_cost, furnace_cost, state_profile['cost_type'], html=True)
    cost_para += (f" Higher-efficiency units cost more upfront but lower monthly bills, "
                  f"especially given {name}'s electricity rates. See our "
                  f'<a href="../costs" style="color: var(--orange-dark); font-weight: 600;">'
                  f"full HVAC cost guide</a> for detailed pricing breakdowns.")
    schema_cost_plain = state_cost_text(name, ac_cost, furnace_cost, state_profile['cost_type'], html=False)
    faq_cost_para = state_cost_text(name, ac_cost, furnace_cost, state_profile['cost_type'], html=True)
    faq_cost_para += f" All new AC units must meet a minimum <strong>SEER2 {seer_min}</strong> rating."
    svc_items = '\n'.join(f'            <li>{svc.format(name=name)}</li>' for svc in state_profile['services'])

    # Utility names for display
    util_parts = [u.strip() for u in str(utilities).split(',')] if utilities else []
    util_bold = ', '.join(f'<strong>{u}</strong>' for u in util_parts)

    # ---- New card-driven layout prep (25 April 2026) ----------------------
    # Federal IRA Rebate Allocation: real, verifiable per-state HEAR figure
    ira_alloc_raw = str(d.get('Federal IRA Rebate Allocation', '') or '').strip()
    ira_alloc = ira_alloc_raw if ira_alloc_raw and ira_alloc_raw.startswith('$') else ''

    # State Energy Office name + URL — column 28-29 in xlsx (preferred);
    # fall back to state_energy_offices.json if column is empty.
    seo_name = str(d.get('State Energy Office Name', '') or '').strip()
    seo_url = str(d.get('State Energy Office URL', '') or '').strip()
    if not (seo_name and seo_url):
        entry = _ENERGY_OFFICES.get(abbr) or {}
        seo_name = seo_name or entry.get('name', '')
        seo_url = seo_url or entry.get('url', '')
    seo_url_display = _strip_scheme(seo_url) if seo_url else ''

    # Climate hazards prose — used in Climate Profile compliance card.
    hazards_list = [h.strip() for h in str(hazards_csv or '').split(',') if h.strip()]
    if hazards_list:
        if len(hazards_list) == 1:
            hazards_prose = f'<strong>{hazards_list[0]}</strong>'
        elif len(hazards_list) == 2:
            hazards_prose = f'<strong>{hazards_list[0]}</strong> and <strong>{hazards_list[1]}</strong>'
        else:
            hazards_prose = ', '.join(f'<strong>{h}</strong>' for h in hazards_list[:-1]) + f', and <strong>{hazards_list[-1]}</strong>'
    else:
        hazards_prose = ''

    # Climate descriptor for cost-context language. Reuses cost_type from state_profile.
    cost_type = state_profile['cost_type']
    if cost_type == 'ac_tropical':
        climate_descriptor = 'tropical, year-round cooling'
    elif cost_type == 'ac_first':
        climate_descriptor = 'cooling-dominated' if 'b' in str(zones).lower() and not 'a' in str(zones).lower() else 'cooling-dominated, hot-humid to hot-dry'
    elif cost_type == 'furnace_first':
        climate_descriptor = 'heating-dominated, long winter season'
    else:
        climate_descriptor = 'mixed cooling and heating demand'

    # Top-of-card cost summary (compact)
    cost_compact_para = state_cost_text(name, ac_cost, furnace_cost, cost_type, html=True)

    # SEER2 region-specific notes — Southwest gets EER2 11.7 callout.
    seer_region_str = str(seer_region or '').strip()
    if seer_region_str.lower() == 'southwest':
        seer_extra = ' with the additional <strong>EER2 11.7</strong> peak-load standard for the Southwest hot-dry region'
    else:
        seer_extra = ''

    # License name short — used in mid-page CTA banner ("TDLR-licensed network").
    body_short = str(d.get('Licensing Body/Agency', '') or '').strip()
    req_yes = str(d.get('State HVAC License Requirement', '') or '').strip().lower().startswith('yes')
    if req_yes and body_short:
        # Pull the abbr inside parens, e.g. "(TDLR)" or use first token. Defensive defaults.
        import re as _re_lic
        m = _re_lic.search(r'\(([^)]+)\)', body_short)
        if m:
            lic_short = m.group(1).split('—')[0].split('-')[0].strip()
        else:
            lic_short = body_short.split(',')[0].strip()[:30]
        cta_band_credential = f'{lic_short}-licensed network'
    else:
        cta_band_credential = 'Independent contractor network'

    # Common HVAC issues card content (pulled from existing helper).
    common_issues_links = (_COMMON_ISSUES_SUMMER if str(peak).strip() == 'Summer'
                           else _COMMON_ISSUES_WINTER if str(peak).strip() == 'Winter'
                           else _COMMON_ISSUES_BOTH)
    common_issues_intro = _COMMON_ISSUES_INTRO.get(str(peak).strip(),
                                                    _COMMON_ISSUES_INTRO['Both']).format(state=name)
    common_issues_lis = '\n'.join(
        f'              <li><a href="{href}" style="color: var(--orange-dark); font-weight: 600;">{label}</a> &mdash; {desc}</li>'
        for href, label, desc in common_issues_links
    )

    # License lookup URL formatting + DSIRE state link
    license_lookup_url = d.get('License Lookup URL', '') or ''
    has_real_license_url = _url_is_real(license_lookup_url)
    license_link_html = (
        f'<a href="{license_lookup_url}" target="_blank" rel="nofollow noopener" '
        f'style="color: var(--orange-dark); font-weight: 600;">{body_short} license lookup</a>'
        if has_real_license_url and req_yes else
        (f'<strong>{body_short}</strong>' if body_short else 'your local building department')
    )

    # Bold-rendered rebate program list ("<strong>X</strong>, <strong>Y</strong>")
    rebate_progs = [r.strip() for r in str(rebates or '').split(',') if r.strip()]
    if rebate_progs:
        rebates_bold = ', '.join(f'<strong>{r}</strong>' for r in rebate_progs)
    else:
        rebates_bold = '<strong>local utility rebates</strong>'

    # State Energy Office card content
    if seo_name and seo_url:
        seo_short_label = seo_name.split(',')[0].strip()[:60] or f'{name} Energy Office'
        seo_paragraph = (
            f'The <strong>{seo_name}</strong> coordinates HVAC rebates, weatherization '
            f'assistance, and federal Inflation Reduction Act program administration in {name}.'
        )
        seo_link_html = (
            f'<a href="{seo_url}" target="_blank" rel="nofollow noopener" '
            f'style="color: var(--orange-dark); font-weight: 600;">{seo_url_display} &rarr;</a>'
        )
    else:
        seo_short_label = f'{name} Energy Office'
        seo_paragraph = (
            f"{name}'s state energy office coordinates HVAC rebates, weatherization "
            f"assistance, and federal Inflation Reduction Act program administration."
        )
        seo_link_html = (
            '<a href="https://www.energy.gov/scep/wap/weatherization-assistance-program" '
            'target="_blank" rel="nofollow noopener" '
            'style="color: var(--orange-dark); font-weight: 600;">DOE WAP &rarr;</a>'
        )

    # IRA HEAR allocation paragraph
    if ira_alloc:
        ira_paragraph = (
            f'{name} was allocated <strong>{ira_alloc}</strong> for federally-funded '
            f'Home Energy Rebates (HEAR) under the Inflation Reduction Act. The state '
            f'is rolling out applications through 2026 &mdash; ask your installer if '
            f'your project qualifies.'
        )
    else:
        ira_paragraph = (
            f'The federal Inflation Reduction Act funds Home Energy Rebates (HEAR) for '
            f'income-qualified households. {name} is administering its allocation through '
            f'the state energy office &mdash; ask your installer if your project qualifies.'
        )

    return f"""<!DOCTYPE html>
<html lang="en">

<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>HVAC Repair in {name} | Cool Call Pro</title>
  <meta name="description"
    content="HVAC in {name}? Cool Call Pro connects homeowners with 24/7 {state_profile['meta_service']}. Call (844) 582-1795." />
  <link rel="canonical" href="https://coolcallpro.com/locations/{s}" />

  <!-- Open Graph -->
  <meta property="og:type" content="website" />
  <meta property="og:title" content="HVAC Repair in {name} | Cool Call Pro" />
  <meta property="og:description" content="Connect with independent HVAC professionals across {name}. {state_profile['og_service']}" />
  <meta property="og:url" content="https://coolcallpro.com/locations/{s}" />
  <meta property="og:site_name" content="Cool Call Pro" />
  <meta property="og:image" content="https://coolcallpro.com/images/og-homepage.webp" />

  <!-- Twitter Card -->
  <meta name="twitter:card" content="summary_large_image" />
  <meta name="twitter:title" content="HVAC Repair in {name}" />
  <meta name="twitter:description" content="Connect with independent HVAC professionals across {name}." />
  <meta name="twitter:image" content="https://coolcallpro.com/images/og-homepage.webp" />

  <!-- Favicon -->
  <link rel="icon" href="/favicon.ico" type="image/x-icon" />
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />

  <link rel="stylesheet" href="../css/style.min.css" />
  <link rel="preload" href="/fonts/inter-latin.woff2" as="font" type="font/woff2" crossorigin>
  <link rel="preload" href="/fonts/outfit-latin.woff2" as="font" type="font/woff2" crossorigin>
  <!-- Google tag (gtag.js) - GA4 — deferred until first interaction for LCP -->
  <script>
      (function() {{
          var loaded = false;
          function loadGA() {{
              if (loaded) return;
              loaded = true;
              var s = document.createElement('script');
              s.src = 'https://www.googletagmanager.com/gtag/js?id=G-WD0ND0K60Q';
              s.async = true;
              document.head.appendChild(s);
              s.onload = function() {{
                  window.dataLayer = window.dataLayer || [];
                  function gtag(){{dataLayer.push(arguments);}}
                  window.gtag = gtag;
                  gtag('js', new Date());
                  gtag('config', 'G-WD0ND0K60Q');
              }};
          }}
          ['click', 'scroll', 'touchstart', 'keydown'].forEach(function(evt) {{
              window.addEventListener(evt, loadGA, {{ once: true, passive: true }});
          }});
          setTimeout(loadGA, 4000);
      }})();
  </script>

  <!-- Structured Data: BreadcrumbList + FAQPage + Organization + Service -->
  <script type="application/ld+json">
  {{
    "@context": "https://schema.org",
    "@graph": [
      {{
        "@type": "BreadcrumbList",
        "itemListElement": [
          {{"@type": "ListItem", "position": 1, "name": "Home", "item": "https://coolcallpro.com/"}},
          {{"@type": "ListItem", "position": 2, "name": "Locations", "item": "https://coolcallpro.com/locations"}},
          {{"@type": "ListItem", "position": 3, "name": "{name}", "item": "https://coolcallpro.com/locations/{s}"}}
        ]
      }},
      {{
        "@type": "FAQPage",
        "mainEntity": [
          {{
            "@type": "Question",
            "name": "What HVAC license is required in {name}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "{lic_faq_schema}"
            }}
          }},
          {{
            "@type": "Question",
            "name": "What SEER2 rating is required for new AC units in {name}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "{name} falls in the {seer_region} SEER2 region. All new central air conditioning systems must meet a minimum SEER2 rating of {seer_min} as of January 2023."
            }}
          }},
          {{
            "@type": "Question",
            "name": "{state_profile['cost_faq_q']} in {name}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "{schema_cost_plain}"
            }}
          }},
          {{
            "@type": "Question",
            "name": "Are there HVAC rebates available in {name}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "Yes. Check with your local utility provider for current energy efficiency rebates. The federal Section 25C tax credit was terminated for installations after Dec 31, 2025 (OBBBA, P.L. 119-21); check current state HEAR rebates and utility programs for 2026."
            }}
          }},
          {{
            "@type": "Question",
            "name": "What type of HVAC system is most common in {name}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "{sys_faq_schema}"
            }}
          }},
          {{
            "@type": "Question",
            "name": "What HVAC permits are required in {name}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "{permits_faq_schema}"
            }}
          }}
        ]
      }},
      {{
        "@type": "Organization",
        "name": "Cool Call Pro",
        "url": "https://coolcallpro.com",
        "telephone": "+1-844-582-1795",
        "description": "24/7 referral service connecting homeowners with independent HVAC professionals across {name}.",
        "areaServed": {{
          "@type": "State",
          "name": "{name}"
        }}
      }},
      {{
        "@type": "Service",
        "serviceType": "HVAC Repair and Installation Referral",
        "provider": {{
          "@type": "Organization",
          "name": "Cool Call Pro",
          "telephone": "+1-844-582-1795",
          "url": "https://coolcallpro.com"
        }},
        "areaServed": {{
          "@type": "State",
          "name": "{name}"
        }},
        "audience": {{ "@type": "Audience", "audienceType": "Homeowners" }}
      }}
    ]
  }}
  </script>

  <style>
    @font-face{{font-family:'Inter';font-style:normal;font-weight:400 900;font-display:swap;src:url('/fonts/inter-latin.woff2') format('woff2')}}@font-face{{font-family:'Outfit';font-style:normal;font-weight:400 800;font-display:swap;src:url('/fonts/outfit-latin.woff2') format('woff2')}}
    .city-hero {{
      background: linear-gradient(135deg, var(--navy-deep) 0%, var(--navy) 50%, var(--navy-mid) 100%);
      background-size: cover;
      background-position: center;
      color: white;
      padding: 100px 0 80px;
      text-align: center;
    }}
    .city-hero h1 {{
      font-family: var(--font-display);
      font-size: 2.6rem;
      font-weight: 800;
      line-height: 1.2;
      margin-bottom: 20px;
      letter-spacing: -0.02em;
    }}
    .city-hero p {{
      font-size: 1.15rem;
      max-width: 720px;
      margin: 0 auto;
      opacity: 0.9;
      line-height: 1.7;
    }}
    .state-info-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
      gap: 20px;
      max-width: 760px;
      margin: 0 auto 16px;
    }}
    .state-info-card {{
      background: #fff;
      border: 1px solid var(--gray-200);
      border-radius: 12px;
      padding: 24px;
      text-align: center;
      box-shadow: 0 4px 16px rgba(10,22,40,0.04);
    }}
    .state-info-card .info-label {{
      font-size: 0.85rem;
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.05em;
      color: var(--gray-500);
      margin-bottom: 8px;
    }}
    .state-info-card .info-value {{
      font-size: 1.4rem;
      font-weight: 700;
      color: var(--navy);
    }}
    .state-info-card .info-value small {{
      font-size: 0.75rem;
      font-weight: 400;
      color: var(--gray-500);
      display: block;
      margin-top: 4px;
    }}
    .city-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 16px;
      max-width: 1100px;
      margin: 0 auto;
    }}
    .city-grid-card {{
      background: #fff;
      border: 1px solid var(--gray-200);
      border-radius: 12px;
      padding: 20px 24px;
      text-decoration: none;
      color: var(--navy);
      font-weight: 600;
      font-size: 1.05rem;
      transition: border-color 0.2s, box-shadow 0.2s;
      display: flex;
      align-items: center;
      gap: 10px;
    }}
    .city-grid-card:hover {{
      border-color: var(--orange-dark);
      box-shadow: 0 2px 8px rgba(255, 165, 0, 0.15);
    }}
    .city-grid-card .card-icon {{
      font-size: 1.2rem;
    }}
    .neighbor-links {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      max-width: 1100px;
      margin: 16px auto 0;
      justify-content: center;
    }}
    .neighbor-link {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      padding: 10px 18px;
      background: #fff;
      border: 1px solid var(--gray-200);
      border-radius: 8px;
      text-decoration: none;
      color: var(--navy);
      font-weight: 600;
      font-size: 0.95rem;
      transition: border-color 0.2s;
    }}
    .neighbor-link:hover {{
      border-color: var(--orange-dark);
    }}
    .neighbor-link.coming-soon {{
      opacity: 0.72;
      pointer-events: none;
    }}
    @media (max-width: 768px) {{
      .city-hero {{
        padding: 80px 0 60px;
      }}
      .city-hero h1 {{
        font-size: 1.8rem;
      }}
      .state-info-grid {{
        grid-template-columns: repeat(2, 1fr);
      }}
    }}
  </style>
</head>

<body>

  <!-- Skip to Content -->
  <a href="#main-content" class="skip-link">Skip to main content</a>

  <!-- Top Bar -->
  <div class="topbar">
    <div class="container topbar-inner" style="justify-content: center; padding: 10px; font-weight: 600;">
      <span class="topbar-text">24/7 Referral Service &#8212; Connecting Homeowners with Independent HVAC Professionals</span>
    </div>
  </div>

  <!-- Header -->
  <header class="header" id="header">
    <nav class="nav container" role="navigation" aria-label="Main navigation">
      <a href="../index" class="logo">
        <span class="logo-icon">&#10052;&#65039;</span>
        <span class="logo-text">CoolCall<span class="logo-accent">Pro</span></span>
      </a>
      <ul class="nav-links" id="navLinks">
        <li><a href="../emergency" class="nav-link emergency-link">&#128680; Emergency Service</a></li>
        <li><a href="../costs" class="nav-link">Cost Guide</a></li>
        <li><a href="../safety" class="nav-link">Safety Tips</a></li>
        <li><a href="../locations" class="nav-link active">&#128205; Locations</a></li>
        <li><a href="../articles" class="nav-link">Articles</a></li>
        <li><a href="../about" class="nav-link">About</a></li>
        <li><a href="../contact" class="nav-link">Contact</a></li>
      </ul>
      <a href="tel:+18445821795" class="btn btn-cta nav-cta btn-vibrate" id="headerNavCta" aria-hidden="true"
        tabindex="-1" style="opacity: 0; pointer-events: none; transition: opacity 0.3s ease;"><span
          class="phone-icon">&#128222;</span>
        <span class="hide-on-mobile">(844) 582-1795</span></a>
      <button class="hamburger" id="hamburger" aria-label="Menu">&#9776;</button>
    </nav>
  </header>

  <!-- Main Content -->
  <main id="main-content">

    <!-- Hero Section -->
    <section class="city-hero" id="hero">
      <div class="container">
        <span class="section-tag" style="background: rgba(255,255,255,0.1); color: #fff;">&#128205; {name}</span>
        <h1>HVAC Service &amp; Repair in {name}</h1>
        <p>Independent local HVAC professionals serve every city and ZIP code across {name}&rsquo;s {hero_lead}. {state_profile['hero_service']}</p>
        {meta_block}
        <div style="margin-top: 28px;">
          <a href="tel:+18445821795" class="btn btn-primary btn-lg btn-vibrate"><span class="phone-icon">&#128222;</span> Call Now &#8212; (844) 582-1795</a>
        </div>
        <div class="jump-links" style="margin-top: 16px;">
          <a href="#trust" class="jump-link" style="color: rgba(255,255,255,0.8);">At a Glance</a>
          <span class="jump-link-dot" style="color: rgba(255,255,255,0.5);">&#8226;</span>
          <a href="#climate" class="jump-link" style="color: rgba(255,255,255,0.8);">Climate</a>
          <span class="jump-link-dot" style="color: rgba(255,255,255,0.5);">&#8226;</span>
          <a href="#services" class="jump-link" style="color: rgba(255,255,255,0.8);">Services &amp; Licensing</a>
          <span class="jump-link-dot" style="color: rgba(255,255,255,0.5);">&#8226;</span>
          <a href="#rebates" class="jump-link" style="color: rgba(255,255,255,0.8);">Rebates</a>
          <span class="jump-link-dot" style="color: rgba(255,255,255,0.5);">&#8226;</span>
          <a href="#cities" class="jump-link" style="color: rgba(255,255,255,0.8);">Service Areas</a>
          <span class="jump-link-dot" style="color: rgba(255,255,255,0.5);">&#8226;</span>
          <a href="#faqs" class="jump-link" style="color: rgba(255,255,255,0.8);">FAQs</a>
        </div>
      </div>
    </section>

    <!-- Breadcrumb -->
    <div class="breadcrumb-nav">
      <div class="container">
        <nav aria-label="Breadcrumb">
          <ol class="breadcrumb-list">
            <li><a href="../index">Home</a></li>
            <li><a href="../locations">Locations</a></li>
            <li aria-current="page">{name}</li>
          </ol>
        </nav>
      </div>
    </div>

    <!-- STATE AT A GLANCE / TRUST STRIP -->
    <section class="section" id="trust" style="padding: 72px 0; background: var(--gray-50);">
      <div class="container">
        <div style="text-align:center; margin-bottom: 36px;">
          <span style="display:inline-block; font-size:0.78rem; font-weight:800; letter-spacing:0.12em; text-transform:uppercase; color:var(--navy); padding:6px 14px; background:#fff; border:1px solid var(--gray-200); border-radius:999px; margin-bottom:16px;">&#128205; State at a Glance</span>
          <h2 style="font-size: 1.75rem; color: var(--navy); margin: 0 0 14px;">HVAC across {name}</h2>
          <p style="max-width: 720px; margin: 0 auto; font-size: 1.05rem; color: var(--gray-700); line-height: 1.7;">{name} is home to over <strong>{pop_long} residents</strong> with a <strong>{homeown}% homeownership rate</strong>. The state spans IECC <strong>{zones_fmt}</strong>, with summer highs averaging {sh}&deg;F and winter lows near {wl}&deg;F.</p>
        </div>

        {state_figure}

        <div class="state-info-grid">
          <div class="state-info-card">
            <div class="info-label">Population</div>
            <div class="info-value">{pop_short}</div>
          </div>
          <div class="state-info-card">
            <div class="info-label">Homeownership</div>
            <div class="info-value">{homeown}%</div>
          </div>
          <div class="state-info-card">
            <div class="info-label">Climate Zones</div>
            <div class="info-value">{zones}</div>
          </div>
          <div class="state-info-card">
            <div class="info-label">SEER2 Minimum</div>
            <div class="info-value">{seer_min} <small>{seer_region} Region</small></div>
          </div>
          <div class="state-info-card">
            <div class="info-label">Summer High</div>
            <div class="info-value">{sh}&deg;F</div>
          </div>
          <div class="state-info-card">
            <div class="info-label">Winter Low</div>
            <div class="info-value">{wl}&deg;F</div>
          </div>
        </div>
      </div>
    </section>

    <!-- CLIMATE & COMPLIANCE -->
    <section class="section" id="climate" style="padding: 72px 0; background:#fff;">
      <div class="container">
        <div style="max-width: 1100px; margin: 0 auto;">
          <div style="text-align:center; margin-bottom: 32px;">
            <span style="display:inline-block; font-size:0.78rem; font-weight:800; letter-spacing:0.12em; text-transform:uppercase; color:#1557b0; padding:6px 14px; background:#ebf8ff; border:1px solid #bee3f8; border-radius:999px; margin-bottom:16px;">&#127777;&#65039; Climate &amp; Compliance</span>
            <h2 style="font-size: 1.75rem; color: var(--navy); margin: 0;">{name} climate, efficiency code &amp; typical costs</h2>
          </div>

          <div class="precaution-cards" style="display:grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 20px; align-items: stretch;">

            <div class="precaution-card" style="display:flex; flex-direction:column; background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);">
              <div class="precaution-icon" style="color: var(--blue);">&#127777;&#65039; Climate Demands</div>
              <h3>{climate_descriptor}</h3>
              <p style="margin-bottom: 0;">{name} spans IECC <strong>{zones}</strong>. Summer highs avg {sh}&deg;F, winter lows near {wl}&deg;F. Average electricity {elec}&cent;/kWh &mdash; system efficiency directly drives monthly costs. Learn more about <a href="{art_url}" style="color: var(--orange-dark); font-weight: 600;">{art_text}</a>.</p>
            </div>

            <div class="precaution-card" style="display:flex; flex-direction:column; background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);">
              <div class="precaution-icon" style="color: var(--blue);">&#128161; SEER2 Code</div>
              <h3>{seer_region} region: SEER2 {seer_min} minimum</h3>
              <p style="margin-bottom: 0;">As of January 2023, all new central AC systems in {name} must meet <strong>SEER2 {seer_min}</strong>{seer_extra}. Higher-efficiency systems cost more upfront but reduce monthly bills. See our <a href="../article-ac-summer" style="color: var(--orange-dark); font-weight: 600;">summer AC guide</a>.</p>
            </div>

            <div class="precaution-card" style="display:flex; flex-direction:column; background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);">
              <div class="precaution-icon" style="color: var(--blue);">&#128176; Typical Costs</div>
              <h3>{name} HVAC installation costs</h3>
              <p style="margin-bottom: 0;">{cost_compact_para} See our <a href="../costs" style="color: var(--orange-dark); font-weight: 600;">full HVAC cost guide</a>.</p>
            </div>

            <div class="precaution-card" style="display:flex; flex-direction:column; background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);">
              <div class="precaution-icon" style="color: var(--blue);">&#127744;&#65039; Climate Hazards</div>
              <h3>FEMA-tracked exposure</h3>
              <p style="margin-bottom: 0;">Per <a href="https://hazards.fema.gov/nri/" target="_blank" rel="nofollow noopener" style="color: var(--orange-dark); font-weight: 600;">FEMA's National Risk Index</a>, {name} faces elevated {hazards_prose} exposure. Each event stresses HVAC systems &mdash; schedule a technician inspection after major weather before restart.</p>
            </div>

          </div>
        </div>
      </div>
    </section>

    <!-- SERVICES & LICENSING -->
    <section class="section" id="services" style="padding: 72px 0; background: var(--gray-50);">
      <div class="container">
        <div style="text-align:center; margin-bottom: 36px;">
          <span style="display:inline-block; font-size:0.78rem; font-weight:800; letter-spacing:0.12em; text-transform:uppercase; color:var(--navy); padding:6px 14px; background:#fff; border:1px solid var(--gray-200); border-radius:999px; margin-bottom:16px;">&#128295; Services &amp; Licensing</span>
          <h2 style="font-size: 1.75rem; color: var(--navy); margin: 0;">What {name} HVAC contractors do &mdash; and what to verify</h2>
        </div>

        <div class="precaution-cards" style="max-width: 1100px; margin: 0 auto; display:grid; grid-template-columns: repeat(auto-fit, minmax(360px, 1fr)); gap: 20px; align-items: stretch;">

          <div class="precaution-card" style="display:flex; flex-direction:column; background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);" id="services-card">
            <div class="precaution-icon" style="color: var(--navy);">&#128295; HVAC Services in {name}</div>
            <h3>What our network covers</h3>
            <ul style="margin-top: 0; margin-bottom: 0;">
{svc_items}
            </ul>
          </div>

          <div class="precaution-card" style="display:flex; flex-direction:column; background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);" id="licensing">
            <div class="precaution-icon" style="color: var(--navy);">&#128203; Licensing Requirements</div>
            <h3>Verify before you hire</h3>
            <p style="margin-bottom: 0;">{lic_html}</p>
          </div>

        </div>
      </div>
    </section>

    <!-- REBATES & PROGRAMS -->
    <section class="section" id="rebates" style="padding: 72px 0; background:#fff;">
      <div class="container">
        <div style="max-width: 1100px; margin: 0 auto;">
          <div style="text-align:center; margin-bottom: 32px;">
            <span style="display:inline-block; font-size:0.78rem; font-weight:800; letter-spacing:0.12em; text-transform:uppercase; color:#1557b0; padding:6px 14px; background:#ebf8ff; border:1px solid #bee3f8; border-radius:999px; margin-bottom:16px;">&#128176; Rebates &amp; Programs</span>
            <h2 style="font-size: 1.75rem; color: var(--navy); margin: 0;">{name} utility rebates &amp; state programs</h2>
          </div>

          <div style="margin: 0 auto 28px; background: linear-gradient(135deg, #f7fafc, #edf2f7); border: 1px solid var(--gray-200); border-left: 5px solid var(--blue); border-radius: var(--radius); padding: 18px 22px;">
            <div style="display: flex; align-items: flex-start; gap: 14px;">
              <div style="font-size: 1.4rem; line-height: 1; flex-shrink: 0;" aria-hidden="true">&#9889;</div>
              <p style="margin: 0; color: var(--gray-700); line-height: 1.65;">Major utility providers in {name} include {util_bold}. Available rebate programs are listed below. Always confirm current amounts with the provider before scheduling work.</p>
            </div>
          </div>

          <div class="precaution-cards" style="display:grid; grid-template-columns: repeat(auto-fit, minmax(360px, 1fr)); gap: 20px; align-items: stretch;">

            <div class="precaution-card" style="display:flex; flex-direction:column; background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);">
              <div class="precaution-icon" style="color: var(--blue);">&#9889; Utility Rebate Programs</div>
              <h3>Through your local provider</h3>
              <p style="margin-bottom: 0;">Active rebate programs include {rebates_bold}. Programs change &mdash; verify current amounts and eligibility before scheduling work.</p>
              <p style="margin-top: auto; padding-top: 8px; margin-bottom: 0;"><a href="https://www.energystar.gov/products/energy_star_home_upgrade/clean_heating_cooling" target="_blank" rel="nofollow noopener" style="color: var(--orange-dark); font-weight: 600;">ENERGY STAR Heating &amp; Cooling &rarr;</a></p>
            </div>

            <div class="precaution-card" style="display:flex; flex-direction:column; background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);">
              <div class="precaution-icon" style="color: var(--blue);">&#127970; State Energy Office</div>
              <h3>{seo_short_label}</h3>
              <p style="margin-bottom: 0;">{seo_paragraph}</p>
              <p style="margin-top: auto; padding-top: 8px; margin-bottom: 0;">{seo_link_html} &middot; <a href="https://www.dsireusa.org/" target="_blank" rel="nofollow noopener" style="color: var(--orange-dark); font-weight: 600;">DSIRE {name} &rarr;</a></p>
            </div>

            <div class="precaution-card" style="display:flex; flex-direction:column; background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);">
              <div class="precaution-icon" style="color: var(--blue);">&#127970; Federal HEAR Allocation</div>
              <h3>IRA-funded state rebate pool</h3>
              <p style="margin-bottom: 0;">{ira_paragraph}</p>
              <p style="margin-top: auto; padding-top: 8px; margin-bottom: 0;"><a href="https://www.energy.gov/scep/home-energy-rebates" target="_blank" rel="nofollow noopener" style="color: var(--orange-dark); font-weight: 600;">DOE Home Energy Rebates &rarr;</a></p>
            </div>

          </div>

          <div style="max-width: 1100px; margin: 32px auto 0; background: linear-gradient(135deg, #fffbeb, #fef3c7); border: 1px solid var(--yellow); border-left: 5px solid var(--orange-dark); border-radius: var(--radius); padding: 20px 24px;">
            <div style="display: flex; align-items: flex-start; gap: 16px;">
              <div style="font-size: 1.6rem; line-height: 1; flex-shrink: 0;" aria-hidden="true">&#128203;</div>
              <div>
                <div style="font-family: var(--font-display); font-weight: 800; font-size: 1.05rem; color: var(--navy); margin-bottom: 6px;">Federal tax credits &mdash; important update for 2026</div>
                <p style="margin: 0; color: var(--gray-700); line-height: 1.65;">The federal <strong>Section 25C Energy Efficient Home Improvement Credit</strong> was terminated for installations placed in service after Dec 31, 2025 by the One Big Beautiful Bill Act (Public Law 119-21). <strong>State HEAR rebates and utility programs remain in effect.</strong> See our <a href="../article-hvac-financing" style="color: var(--orange-dark); font-weight: 600;">HVAC financing options</a> for what's still available.</p>
              </div>
            </div>
          </div>
        </div>
      </div>
    </section>

    <!-- MID-PAGE CTA -->
    <section style="padding: 64px 0; background: var(--gray-50);">
      <div class="container">
        <div style="max-width: 760px; margin: 0 auto; background:linear-gradient(135deg, var(--navy-deep), var(--navy)); color:#fff; padding: 32px; border-radius: var(--radius-lg); text-align:center; box-shadow: 0 12px 40px rgba(10,22,40,0.18);">
          <p style="font-size: 1.4rem; font-weight: 700; font-family: var(--font-display); color:#fff; margin-bottom: 10px;">Ready to talk to a {name} HVAC pro?</p>
          <p style="margin-bottom: 22px; opacity: 0.85; font-size: 1rem;">Independent technicians &middot; 24/7 dispatch &middot; {cta_band_credential}</p>
          <a href="tel:+18445821795" class="btn btn-primary btn-lg btn-vibrate"><span class="phone-icon">&#128222;</span> Call Now &mdash; (844) 582-1795</a>
          <p style="font-size: 0.72rem; color: rgba(255,255,255,0.5); margin: 16px auto 0; max-width: 460px; line-height: 1.5;">Disclosure: We are a referral service and may receive compensation for qualified calls. Calls may be routed to an independent provider network and may be recorded. Pricing and availability vary by provider and location.</p>
        </div>
      </div>
    </section>

    <!-- COMMON HVAC ISSUES (climate-keyed cluster links) -->
    <section class="section" style="padding: 72px 0; background:#fff;">
      <div class="container">
        <div style="max-width: 1000px; margin: 0 auto;">
          <div style="text-align:center; margin-bottom: 32px;">
            <span style="display:inline-block; font-size:0.78rem; font-weight:800; letter-spacing:0.12em; text-transform:uppercase; color:var(--navy); padding:6px 14px; background:var(--gray-50); border:1px solid var(--gray-200); border-radius:999px; margin-bottom:16px;">&#128295; Common Issues</span>
            <h2 style="font-size: 1.75rem; color: var(--navy); margin: 0 0 14px;">Common HVAC issues in {name}</h2>
            <p style="max-width: 720px; margin: 0 auto; font-size: 1.05rem; color: var(--gray-700); line-height: 1.7;">{common_issues_intro}</p>
          </div>
          <div class="precaution-card" style="background:#fff; border:1px solid var(--gray-200); box-shadow: 0 4px 16px rgba(10,22,40,0.04);">
            <ul style="margin: 0;">
{common_issues_lis}
            </ul>
          </div>
        </div>
      </div>
    </section>

    <!-- Service Areas (Cities in {name}) -->
    <section class="section" id="cities" style="padding: 72px 0; background: var(--gray-50);">
      <div class="container">
        <div style="text-align:center; margin-bottom: 36px;">
          <span style="display:inline-block; font-size:0.78rem; font-weight:800; letter-spacing:0.12em; text-transform:uppercase; color:var(--navy); padding:6px 14px; background:#fff; border:1px solid var(--gray-200); border-radius:999px; margin-bottom:16px;">&#128205; Service Areas</span>
          <h2 style="font-size: 1.75rem; color: var(--navy); margin: 0 0 14px;">HVAC service areas across {name}</h2>
          <p style="max-width: 720px; margin: 0 auto; font-size: 1.05rem; color: var(--gray-700); line-height: 1.7;">Cool Call Pro connects homeowners with independent HVAC professionals across {name}. Browse city pages for local costs, permit offices, and licensing details.</p>
        </div>{featured_callout}
        <div class="city-grid">
{city_grid}
        </div>
      </div>
    </section>

    <!-- FAQs -->
    <section class="section" id="faqs" style="padding: 72px 0; background:#fff;">
      <div class="container" style="max-width: 800px;">
        <div style="text-align:center; margin-bottom: 36px;">
          <span style="display:inline-block; font-size:0.78rem; font-weight:800; letter-spacing:0.12em; text-transform:uppercase; color:var(--navy); padding:6px 14px; background:var(--gray-50); border:1px solid var(--gray-200); border-radius:999px; margin-bottom:16px;">&#10067; Frequently Asked Questions</span>
          <h2 style="font-size: 1.75rem; color: var(--navy); margin: 0;">{name} HVAC &mdash; common questions</h2>
        </div>
        <div class="faq-list" style="margin-top: 8px;">
            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>What HVAC license is required in {name}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>{lic_faq}</p>
                </div>
              </div>
            </div>

            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>What SEER2 rating is required for new AC units in {name}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>{name} is in the <strong>{seer_region} SEER2 region</strong>, requiring a minimum <strong>SEER2 {seer_min}</strong> for all new central AC systems as of January 2023.</p>
                </div>
              </div>
            </div>

            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>{state_profile['cost_faq_q']} in {name}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>{faq_cost_para}</p>
                </div>
              </div>
            </div>

            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>Are there HVAC rebates available in {name}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>Yes. Check with your local utility provider for current energy efficiency rebates on qualifying high-efficiency equipment. The federal Section 25C tax credit was terminated for installations after Dec 31, 2025 by the One Big Beautiful Bill Act (Public Law 119-21); state HEAR rebates and utility programs remain in effect for 2026.</p>
                </div>
              </div>
            </div>

            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>What type of HVAC system is most common in {name}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>{sys_faq}</p>
                </div>
              </div>
            </div>
            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>What HVAC permits are required in {name}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>{permits_faq_html_str}</p>
                </div>
              </div>
            </div>
        </div>

        <div style="text-align: center; margin-top: 32px;">
          <a href="tel:+18445821795" class="btn btn-primary btn-lg btn-vibrate"><span class="phone-icon">&#128222;</span> Call Now &#8212; (844) 582-1795</a>
        </div>
      </div>
    </section>

{neighbor_section}
{sources_block}

  </main>

  <!-- Footer -->
  <footer class="footer">
    <div class="container">
      <div class="footer-grid">
        <div class="footer-brand">
          <a href="../index" class="logo"><span class="logo-icon">&#10052;&#65039;</span><span
              class="logo-text">CoolCall<span class="logo-accent">Pro</span></span></a>
          <p>Connecting homeowners with independent HVAC professionals across the US. Available 24/7 in many areas.</p>
          <div class="footer-contact">
            <a href="tel:+18445821795" class="footer-phone">(844) 582-1795</a>
            <span>24/7 Service Line</span>
          </div>
        </div>
        <div class="footer-col">
          <h3 class="footer-heading">Resources</h3>
          <ul>
            <li><a href="../costs">Cost Guide</a></li>
            <li><a href="../emergency">Emergency Service</a></li>
            <li><a href="../safety">Safety Tips</a></li>
            <li><a href="../articles">All Articles</a></li>
            <li><a href="../locations">Locations</a></li>
          </ul>
        </div>
        <div class="footer-col">
          <h3 class="footer-heading">Company</h3>
          <ul>
            <li><a href="../about">About Us</a></li>
            <li><a href="../contact">Contact</a></li>
            <li><a href="../author-gyanesh">Author</a></li>
            <li><a href="../privacy">Privacy Policy</a></li>
            <li><a href="../advertising-disclosure">Advertising Disclosure</a></li>
          </ul>
        </div>
        <div class="footer-col">
          <h3 class="footer-heading">Service Areas</h3>
          <ul class="service-areas">
            <li><a href="atlanta-ga">Atlanta, GA</a></li>
            <li><a href="chicago-il">Chicago, IL</a></li>
            <li><a href="dallas-tx">Dallas, TX</a></li>
            <li><a href="houston-tx">Houston, TX</a></li>
            <li><a href="phoenix-az">Phoenix, AZ</a></li>
            <li><a href="../locations">All 50 states &#8594;</a></li>
          </ul>
        </div>
      </div>
      <div class="footer-bottom">
        <p><strong style="color: rgba(255,255,255,0.82);">How this site works:</strong> We publish HVAC cost and
          troubleshooting information and may connect callers to independent service providers. We do not perform HVAC
          services ourselves. Pricing, availability, and response times vary by provider and location.</p>
        <p><strong style="color: rgba(255,255,255,0.82);">Editorial standards:</strong> Our content focuses on helpful
          troubleshooting tips and realistic pricing information. We avoid guarantees; actual pricing depends on the
          provider and situation. We clearly disclose compensation for referrals. We update content for clarity and
          accuracy when needed.</p>
        <p><strong style="color: rgba(255,255,255,0.82);">Disclaimer:</strong> Cool Call Pro is a free service to assist homeowners in connecting with local service providers. All contractors/providers are independent and Cool Call Pro does not warrant or guarantee any work performed. It is the responsibility of the homeowner to verify that the hired contractor furnishes the necessary license and insurance required for the work being performed. All persons depicted in a photo or video are actors or models and not contractors listed on Cool Call Pro.</p>
        <p>&copy; <span class="copyright-year">2026</span> Cool Call Pro. All rights reserved. &nbsp;&#183;&nbsp; <a href="../privacy">Privacy
            Policy</a> &nbsp;&#183;&nbsp; <a href="../terms">Terms of Use</a> &nbsp;&#183;&nbsp; <a
            href="../disclaimer">Disclaimer</a> &nbsp;&#183;&nbsp; <a href="../advertising-disclosure">Advertising
            Disclosure</a> &nbsp;&#183;&nbsp; <a href="../image-credits">Image Credits</a></p>
      </div>
    </div>
  </footer>

  <script src="../js/main.min.js" defer></script>

  <!-- Mobile Call Bar -->
  <div class="mobile-call-bar">
    <a href="tel:+18445821795"><svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor" style="width: 1.25em; height: 1.25em; position: relative; top: 0.15em;"><path d="M20.01 15.38c-1.23 0-2.42-.2-3.53-.56-.35-.12-.74-.03-1.01.24l-1.57 1.97c-2.83-1.35-5.48-3.9-6.89-6.83l1.95-1.66c.27-.28.35-.67.24-1.02-.37-1.11-.56-2.3-.56-3.53 0-.54-.45-.99-.99-.99H4.19C3.65 3 3 3.24 3 3.99 3 13.28 10.73 21 20.01 21c.71 0 1.01-.63 1.01-1.18v-3.45c0-.54-.45-.99-.99-.99z"/></svg> Call Now &#8212; (844) 582-1795</a>
  </div>
<script>document.querySelectorAll(".copyright-year").forEach(function(el){{el.textContent=new Date().getFullYear()}});</script>
</body>

</html>"""

# ─── Main ───────────────────────────────────────────────────────────────

def main():
    states = load_states()
    cities = load_cities()

    # Determine which states have at least 1 published city page
    hub_abbrs = set()
    for abbr, city_list in cities.items():
        for c in city_list:
            if c.get('City Page Published') == 'Yes' or c.get('City') in GENERATED_CITIES:
                hub_abbrs.add(abbr)
                break

    # Accept CLI args: specific state abbreviations, or 'all' for all hub states.
    # --refresh-date bumps the "Last reviewed" date to today; otherwise the
    # existing date in the file is preserved (sticky default).
    # --out-dir <path> writes to a non-default directory (e.g. _pilot_test/states).
    raw_args = sys.argv[1:]
    refresh_date = '--refresh-date' in raw_args
    out_dir = 'locations'
    if '--out-dir' in raw_args:
        idx = raw_args.index('--out-dir')
        if idx + 1 < len(raw_args):
            out_dir = raw_args[idx + 1]
            raw_args = raw_args[:idx] + raw_args[idx + 2:]
    targets = [a for a in raw_args if a != '--refresh-date']
    if not targets:
        print("Usage: python generate_state_hubs.py AL  (or: AL TX FL  or: all)")
        print("       add --refresh-date to bump Last-reviewed to today")
        print("       add --out-dir <path> to write to a different directory (default: locations)")
        print(f"\nStates with published cities (eligible for hubs): {sorted(hub_abbrs)}")
        return

    if targets == ['all']:
        targets = sorted(hub_abbrs)

    os.makedirs(out_dir, exist_ok=True)

    for abbr in targets:
        abbr = abbr.upper()
        if abbr not in states:
            print(f"  SKIP: {abbr} not found in states.xlsx")
            continue
        if abbr not in hub_abbrs:
            print(f"  SKIP: {abbr} has no published city pages")
            continue

        d = states[abbr]
        city_list = cities.get(abbr, [])
        filename = slug(d['State Name']) + '.html'
        filepath = os.path.join(out_dir, filename)
        html = generate_html(d, city_list, hub_abbrs, states,
                             out_path=filepath, refresh_date=refresh_date)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"  DONE: {filepath} ({d['State Name']})")

    print(f"\nGenerated {len(targets)} state hub page(s).")

if __name__ == '__main__':
    main()
