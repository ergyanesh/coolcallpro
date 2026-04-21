"""
One-off: add a "Climate Hazards" column to states.xlsx.

Populates each state row with a comma-separated list of 3 climate hazards
drawn from FEMA's National Risk Index hazard vocabulary. Hazard categories
are FEMA NRI types; per-state assignments reflect the hazards FEMA NRI
and NOAA storm data identify as elevated for each state.

After this script runs once, states.xlsx is the canonical source-of-truth
for per-state hazard lists. The hazard vocabulary (impact + pro-action copy
per hazard) lives in climate_hazards.json and is consumed by both the state
hub HTML patcher and generate_state_hubs.py going forward.

Idempotent: if the "Climate Hazards" column already exists, the script
verifies cells match the expected values and exits without rewriting the
file. To force a refresh, delete the column manually first.

Data source: https://hazards.fema.gov/nri/ (state-level hazard exposure)
"""

from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
STATES_XLSX = ROOT / "states.xlsx"
COLUMN_NAME = "Climate Hazards"

# Per-state 3-hazard assignments. Hazards listed roughly in order of
# frequency/severity per FEMA NRI + NOAA storm data for that state.
STATE_HAZARDS = {
    "Alabama":              ["Tornado", "Hurricane", "Severe Thunderstorm"],
    "Alaska":               ["Extreme Cold", "Earthquake", "High Wind"],
    "Arizona":              ["Extreme Heat", "Dust Storm", "Wildfire"],
    "Arkansas":             ["Tornado", "Severe Thunderstorm", "Ice Storm"],
    "California":           ["Wildfire", "Earthquake", "Extreme Heat"],
    "Colorado":             ["Hail", "Wildfire", "Blizzard"],
    "Connecticut":          ["Nor'easter", "Ice Storm", "Severe Thunderstorm"],
    "Delaware":             ["Nor'easter", "Coastal Flooding", "Severe Thunderstorm"],
    "District of Columbia": ["Severe Thunderstorm", "Extreme Heat", "Ice Storm"],
    "Florida":              ["Hurricane", "Severe Thunderstorm", "Flooding"],
    "Georgia":              ["Tornado", "Hurricane", "Ice Storm"],
    "Hawaii":               ["Tropical Storm", "Volcanic Activity", "Flooding"],
    "Idaho":                ["Wildfire", "Extreme Cold", "Winter Storm"],
    "Illinois":             ["Tornado", "Extreme Cold", "Severe Thunderstorm"],
    "Indiana":              ["Tornado", "Severe Thunderstorm", "Ice Storm"],
    "Iowa":                 ["Tornado", "Severe Thunderstorm", "Blizzard"],
    "Kansas":               ["Tornado", "Hail", "Severe Thunderstorm"],
    "Kentucky":             ["Tornado", "Ice Storm", "Severe Thunderstorm"],
    "Louisiana":            ["Hurricane", "Flooding", "Tornado"],
    "Maine":                ["Nor'easter", "Ice Storm", "Extreme Cold"],
    "Maryland":             ["Nor'easter", "Severe Thunderstorm", "Hurricane"],
    "Massachusetts":        ["Nor'easter", "Blizzard", "Coastal Flooding"],
    "Michigan":             ["Lake-Effect Snow", "Extreme Cold", "Severe Thunderstorm"],
    "Minnesota":            ["Extreme Cold", "Blizzard", "Tornado"],
    "Mississippi":          ["Tornado", "Hurricane", "Severe Thunderstorm"],
    "Missouri":             ["Tornado", "Severe Thunderstorm", "Ice Storm"],
    "Montana":              ["Wildfire", "Winter Storm", "Extreme Cold"],
    "Nebraska":             ["Tornado", "Blizzard", "Hail"],
    "Nevada":               ["Wildfire", "Extreme Heat", "Earthquake"],
    "New Hampshire":        ["Nor'easter", "Ice Storm", "Extreme Cold"],
    "New Jersey":           ["Nor'easter", "Coastal Flooding", "Hurricane"],
    "New Mexico":           ["Wildfire", "Severe Thunderstorm", "Extreme Heat"],
    "New York":             ["Nor'easter", "Lake-Effect Snow", "Ice Storm"],
    "North Carolina":       ["Hurricane", "Tornado", "Ice Storm"],
    "North Dakota":         ["Extreme Cold", "Blizzard", "Tornado"],
    "Ohio":                 ["Tornado", "Lake-Effect Snow", "Ice Storm"],
    "Oklahoma":             ["Tornado", "Severe Thunderstorm", "Ice Storm"],
    "Oregon":               ["Wildfire", "Earthquake", "Winter Storm"],
    "Pennsylvania":         ["Ice Storm", "Nor'easter", "Severe Thunderstorm"],
    "Rhode Island":         ["Nor'easter", "Coastal Flooding", "Ice Storm"],
    "South Carolina":       ["Hurricane", "Tornado", "Ice Storm"],
    "South Dakota":         ["Blizzard", "Tornado", "Hail"],
    "Tennessee":            ["Tornado", "Ice Storm", "Severe Thunderstorm"],
    "Texas":                ["Hurricane", "Tornado", "Extreme Heat"],
    "Utah":                 ["Wildfire", "Earthquake", "Winter Storm"],
    "Vermont":              ["Nor'easter", "Ice Storm", "Extreme Cold"],
    "Virginia":             ["Hurricane", "Tornado", "Ice Storm"],
    "Washington":           ["Wildfire", "Earthquake", "Winter Storm"],
    "West Virginia":        ["Ice Storm", "Severe Thunderstorm", "Flooding"],
    "Wisconsin":            ["Extreme Cold", "Lake-Effect Snow", "Tornado"],
    "Wyoming":              ["Winter Storm", "Extreme Cold", "Wildfire"],
}


def main():
    wb = openpyxl.load_workbook(STATES_XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    if COLUMN_NAME in headers:
        col_idx = headers.index(COLUMN_NAME) + 1
        print(f"'{COLUMN_NAME}' already exists at column {col_idx}. Verifying values...")
        state_col = headers.index("State Name") + 1
        mismatches = []
        for row_i in range(2, ws.max_row + 1):
            state = ws.cell(row=row_i, column=state_col).value
            if not state:
                continue
            expected = ", ".join(STATE_HAZARDS.get(state, []))
            actual = ws.cell(row=row_i, column=col_idx).value or ""
            if expected and actual.strip() != expected:
                mismatches.append((state, actual, expected))
        if mismatches:
            print(f"  {len(mismatches)} row(s) differ from expected values:")
            for state, actual, expected in mismatches[:5]:
                print(f"    {state}: have '{actual}', expect '{expected}'")
            print("  Not overwriting. Delete the column manually to regenerate.")
            return 1
        print(f"  All {len(STATE_HAZARDS)} rows match expected values. No change.")
        return 0

    new_col_idx = len(headers) + 1
    print(f"Adding '{COLUMN_NAME}' as column {new_col_idx} ({openpyxl.utils.get_column_letter(new_col_idx)})")

    ws.cell(row=1, column=new_col_idx, value=COLUMN_NAME)
    state_col = headers.index("State Name") + 1
    written = 0
    missing = []
    for row_i in range(2, ws.max_row + 1):
        state = ws.cell(row=row_i, column=state_col).value
        if not state:
            continue
        hazards = STATE_HAZARDS.get(state)
        if not hazards:
            missing.append(state)
            continue
        ws.cell(row=row_i, column=new_col_idx, value=", ".join(hazards))
        written += 1

    wb.save(STATES_XLSX)
    print(f"  wrote {written} rows")
    if missing:
        print(f"  missing hazards for: {missing}")
        return 1
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
