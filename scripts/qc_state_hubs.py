"""Comprehensive QC for state hub pages — mirrors scripts/qc_city_pages.py
but with state-hub-specific checks. Run after every batch regeneration.

Usage:
  python scripts/qc_state_hubs.py [--dir locations] [--states texas.html ...]
  python scripts/qc_state_hubs.py --pilot     # check _pilot_test/states/

Exit code is 1 if any state hub fails any check, 0 otherwise (so it can
gate CI / pre-commit). Failures are grouped by check type so a single
generator bug surfaces as a pattern, not 51 separate noise lines.

State-hub specific checks (different from cities):
  - Required section markers: STATE AT A GLANCE, CLIMATE & COMPLIANCE,
    SERVICES & LICENSING, REBATES & PROGRAMS, MID-PAGE CTA, FAQs,
    Service Areas, Common HVAC Issues
  - OBBBA-aware: no banned tax-credit phrases
  - Clean URLs (no internal .html), canonical at https://coolcallpro.com
  - Valid JSON-LD with Organization + Service + FAQPage + BreadcrumbList
  - No LocalBusiness or LocalService schema
  - Hero subhead has no leaked placeholders ({name}, {sh}, etc.)
  - Climate Profile compliance card cites FEMA NRI primary source
  - License Requirements card cites a primary-source URL when xlsx says
    licensing == 'Yes'
  - State Energy Office card cites a primary-source URL
  - Mobile call bar present
  - Footer uses .footer-col wrappers
  - Card-driven design discipline: no H3+P-island anti-patterns
"""

import os
import re
import sys
import json
import argparse

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

# ============================================================
# Required markers and banned phrases
# ============================================================

REQUIRED_SECTION_MARKERS = [
    "STATE AT A GLANCE",          # comment marker on Trust Strip section
    "CLIMATE & COMPLIANCE",       # comment marker on climate section
    "SERVICES & LICENSING",       # comment marker on services section
    "REBATES & PROGRAMS",         # comment marker on rebates section
    "MID-PAGE CTA",               # comment marker on mid-page CTA
    "FAQs",                       # comment marker on FAQs section
    "Service Areas",              # comment marker on cities section
    "COMMON HVAC ISSUES",         # comment marker on common issues
]

# OBBBA-stale claims that must NEVER appear on state hubs (terminated 25C)
STALE_OBBBA_PATTERNS = [
    "may also receive up to $3,200/year",
    "stack with the federal",
    "stack with 25C",
    "Federal tax credits of up to $2,000 for qualifying heat pumps",
    "Federal ENERGY STAR tax credits may also apply",
    # Generic "may also apply" framing (use OBBBA-aware copy instead)
    "Federal tax credits may also apply for qualifying high-efficiency",
    "All local incentives stack with the federal",
]

STATE_FILE_RE = re.compile(r"^[a-z][a-z\-]+\.html$")  # e.g. texas.html, new-york.html

# Cities pages match this (we want to EXCLUDE them from the state-hub QC sweep)
CITY_FILE_RE = re.compile(r"^[a-z0-9\-]+-[a-z]{2}\.html$")


def load_states_xlsx(path: str = "states.xlsx") -> dict:
    """slug -> dict(state_name, abbr, license_required, license_url, seo_url, hazards)"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        d = dict(zip(headers, row))
        slug = str(d["State Name"]).lower().replace(" ", "-")
        out[slug] = {
            "name": d["State Name"],
            "abbr": d["State Abbreviation"],
            "license_req": str(d.get("State HVAC License Requirement", "") or "").strip(),
            "license_url": str(d.get("License Lookup URL", "") or "").strip(),
            "seo_url": str(d.get("State Energy Office URL", "") or "").strip(),
            "hazards": str(d.get("Climate Hazards", "") or "").strip(),
            "ira_alloc": str(d.get("Federal IRA Rebate Allocation", "") or "").strip(),
            "published": str(d.get("State Hub Published", "") or "").strip(),
        }
    return out


# ============================================================
# Per-page checks
# ============================================================
def check_required_sections(html):
    failures = []
    for marker in REQUIRED_SECTION_MARKERS:
        if marker not in html:
            failures.append(f"section marker missing: '{marker}'")
    return failures


def check_obbba_clean(html):
    failures = []
    for pattern in STALE_OBBBA_PATTERNS:
        if pattern in html:
            failures.append(f"stale OBBBA claim: '{pattern[:60]}...'")
    # Must have the OBBBA-aware termination clause somewhere
    if not re.search(r"Section 25C.*?terminated for installations after Dec 31, 2025", html):
        failures.append("missing OBBBA-aware Section 25C termination clause")
    return failures


def check_clean_urls(html):
    failures = []
    internal_html = re.findall(
        r'href="(?!https?://|mailto:|tel:|javascript:|#)[^"]+\.html(?:#[^"]*)?"',
        html,
    )
    if internal_html:
        failures.append(
            f"{len(internal_html)} internal href(s) end in .html — first: {internal_html[0]}"
        )
    canon = re.search(r'<link rel="canonical"\s+href="([^"]+)"', html)
    if not canon:
        failures.append("missing canonical")
    else:
        url = canon.group(1)
        if not url.startswith("https://coolcallpro.com"):
            failures.append(f"canonical wrong domain: {url}")
        if url.endswith(".html"):
            failures.append(f"canonical ends in .html: {url}")
    return failures


def check_jsonld(html):
    failures = []
    blocks = re.findall(
        r'<script type="application/ld\+json">(.*?)</script>', html, re.DOTALL
    )
    if not blocks:
        failures.append("no JSON-LD blocks")
        return failures
    parsed = []
    for i, raw in enumerate(blocks):
        try:
            parsed.append(json.loads(raw))
        except json.JSONDecodeError as e:
            failures.append(f"JSON-LD block {i + 1} invalid: {e}")
    # Check the @graph contains required entities
    required_types = {"BreadcrumbList", "FAQPage", "Organization", "Service"}
    found_types = set()
    for obj in parsed:
        graph = obj.get("@graph", [obj])
        for entity in graph if isinstance(graph, list) else [graph]:
            t = entity.get("@type") if isinstance(entity, dict) else None
            if isinstance(t, str):
                found_types.add(t)
    missing = required_types - found_types
    if missing:
        failures.append(f"JSON-LD missing schema types: {sorted(missing)}")
    return failures


def check_no_local_business(html):
    failures = []
    if '"@type": "LocalBusiness"' in html or '"@type":"LocalBusiness"' in html:
        failures.append("LocalBusiness schema present (banned for referral site)")
    if '"@type": "LocalService"' in html or '"@type":"LocalService"' in html:
        failures.append("LocalService schema present (banned for referral site)")
    return failures


def check_hero_subhead(html):
    m = re.search(r'<section class="city-hero".*?</section>', html, re.DOTALL)
    if not m:
        return ["missing city-hero section"]
    section = m.group(0)
    failures = []
    sub = re.search(r"<h1>.*?</h1>\s*<p>(.*?)</p>", section, re.DOTALL)
    if not sub:
        return ["no hero subhead <p> found"]
    text = re.sub(r"<[^>]+>", "", sub.group(1))
    # Placeholders that should never leak
    for placeholder in ("{name}", "{sh}", "{wl}", "{hero_lead}", "{state}", "{state_profile"):
        if placeholder in text:
            failures.append(f"subhead leaked {placeholder} placeholder")
    if len(text) < 30:
        failures.append(f"subhead too short ({len(text)} chars)")
    if len(text) > 320:
        failures.append(f"subhead too long ({len(text)} chars) — risk of orphan-word wrap")
    return failures


def check_meta_basics(html):
    failures = []
    h1_matches = re.findall(r"<h1[\s>]", html)
    if len(h1_matches) != 1:
        failures.append(f"{len(h1_matches)} <h1> tags (expected 1)")
    title = re.search(r"<title>([^<]+)</title>", html)
    if not title:
        failures.append("missing <title>")
    elif len(title.group(1)) > 70:
        failures.append(f"<title> length {len(title.group(1))} > 70")
    desc = re.search(r'<meta name="description"[^>]*content="([^"]+)"', html)
    if not desc:
        failures.append("missing meta description")
    elif len(desc.group(1)) > 160:
        failures.append(f"meta description length {len(desc.group(1))} > 160")
    return failures


def check_mobile_call_bar(html):
    if '<div class="mobile-call-bar">' not in html:
        return ["missing mobile-call-bar"]
    return []


def check_footer_col(html):
    if 'class="footer"' in html and 'class="footer-col"' not in html:
        return ["footer present but missing footer-col wrappers (renders invisible)"]
    return []


def check_climate_card_primary_source(html):
    """The Climate Hazards card must cite FEMA's National Risk Index URL.
    Required by 'every numerical claim or institution name must link to its
    primary source via <a href>'."""
    if "hazards.fema.gov/nri" not in html:
        return ["Climate Hazards card missing FEMA NRI primary-source link (hazards.fema.gov/nri)"]
    return []


def check_license_card_primary_source(html, license_req, license_url):
    """When xlsx license requirement == Yes AND License Lookup URL is real,
    the Licensing Requirements card MUST link to that URL. Otherwise (Varies
    or no real URL), no link is required."""
    if license_req.lower().startswith("yes") and license_url.lower().startswith(("http://", "https://")):
        if license_url not in html:
            return [f"Licensing card does not link to License Lookup URL ({license_url[:60]})"]
    return []


def check_state_energy_office_link(html, seo_url):
    """The Rebates section must link to the State Energy Office URL when xlsx provides one."""
    if seo_url and seo_url.lower().startswith(("http://", "https://")):
        if seo_url.rstrip("/") not in html:
            return [f"State Energy Office card does not link to {seo_url[:60]}"]
    return []


def check_dsire_link(html):
    if "dsireusa.org" not in html:
        return ["Rebates section missing DSIRE primary-source link"]
    return []


def check_h2_count(html):
    """Card-driven layout has ~7-9 H2s per state hub. Fewer means the layout
    regressed; more means we likely double-printed a section."""
    h2_count = len(re.findall(r"<h2[\s>]", html))
    if h2_count < 6:
        return [f"only {h2_count} <h2> sections — layout incomplete"]
    if h2_count > 12:
        return [f"{h2_count} <h2> sections — possibly duplicated content"]
    return []


def check_ira_allocation(html, ira_alloc):
    """If xlsx provides an IRA allocation, the Rebates section must reference it."""
    if ira_alloc and ira_alloc.startswith("$"):
        if ira_alloc not in html:
            return [f"IRA HEAR allocation {ira_alloc} from xlsx not present in page body"]
    return []


def check_obbba_callout_banner(html):
    """The Rebates section must include the warning-orange OBBBA banner."""
    if "Federal tax credits &mdash; important update for 2026" not in html:
        return ["missing OBBBA termination warning banner in Rebates section"]
    return []


def check_card_design_discipline(html):
    """No section should have 2+ direct-child <h3> elements (H3+P island anti-pattern).
    Mirrors the spirit of audit_page.py's audit_h3_islands but is a coarser
    text-level check that catches the common regression."""
    # Find all <section> blocks. If a section has >=3 <h3> tags as direct children
    # (not inside .precaution-card / .wh-step / .faq-item / .city-services),
    # flag it. This is a loose heuristic — audit_page.py is the strict gate.
    failures = []
    sections = re.findall(r"<section[^>]*>(.*?)</section>", html, re.DOTALL)
    for s in sections:
        # Strip out cards
        cleaned = re.sub(r'<div class="precaution-card[^"]*"[^>]*>.*?</div>\s*</div>',
                         "", s, flags=re.DOTALL)
        cleaned = re.sub(r'<div class="faq-item"[^>]*>.*?</div>\s*</div>',
                         "", cleaned, flags=re.DOTALL)
        loose_h3 = re.findall(r"<h3[\s>]", cleaned)
        if len(loose_h3) >= 3:
            failures.append(f"section has {len(loose_h3)} loose <h3> elements after stripping cards (H3+P-island risk)")
            break  # one per page is enough signal
    return failures


# ============================================================
# Main
# ============================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default="locations", help="Directory to scan (default: locations)")
    ap.add_argument("--pilot", action="store_true", help="Scan _pilot_test/states/ instead")
    ap.add_argument("--states", nargs="*", help="Specific filenames to check (e.g. texas.html)")
    args = ap.parse_args()

    if args.pilot:
        target_dir = "_pilot_test/states"
    else:
        target_dir = args.dir

    if not os.path.isdir(target_dir):
        print(f"ERROR: directory not found: {target_dir}")
        sys.exit(2)

    xlsx_data = load_states_xlsx()

    if args.states:
        files = [f for f in args.states if os.path.exists(os.path.join(target_dir, f))]
    else:
        files = sorted(
            f for f in os.listdir(target_dir)
            if STATE_FILE_RE.match(f) and not CITY_FILE_RE.match(f)
        )

    if not files:
        print(f"No state-hub HTML files found in {target_dir}")
        sys.exit(0)

    print(f"Running QC on {len(files)} state hubs in {target_dir}/")
    print("=" * 100)

    failure_pattern = {}  # check_name -> [(slug, msg)]
    perfect = []
    failed = []

    for fn in files:
        slug = fn.replace(".html", "")
        path = os.path.join(target_dir, fn)
        with open(path, "r", encoding="utf-8") as fh:
            html = fh.read()

        meta = xlsx_data.get(slug)
        if not meta:
            failure_pattern.setdefault("xlsx_lookup", []).append(
                (slug, f"no xlsx row for slug: {slug}")
            )
            failed.append(slug)
            continue

        license_req = meta["license_req"]
        license_url = meta["license_url"]
        seo_url = meta["seo_url"]
        ira_alloc = meta["ira_alloc"]

        runs = [
            ("required_sections",  lambda h: check_required_sections(h)),
            ("obbba_clean",        lambda h: check_obbba_clean(h)),
            ("clean_urls",         lambda h: check_clean_urls(h)),
            ("jsonld",             lambda h: check_jsonld(h)),
            ("no_local_biz",       lambda h: check_no_local_business(h)),
            ("hero_subhead",       lambda h: check_hero_subhead(h)),
            ("meta_basics",        lambda h: check_meta_basics(h)),
            ("mobile_call_bar",    lambda h: check_mobile_call_bar(h)),
            ("footer_col",         lambda h: check_footer_col(h)),
            ("fema_primary",       lambda h: check_climate_card_primary_source(h)),
            ("license_primary",    lambda h: check_license_card_primary_source(h, license_req, license_url)),
            ("seo_primary",        lambda h: check_state_energy_office_link(h, seo_url)),
            ("dsire_link",         lambda h: check_dsire_link(h)),
            ("h2_count",           lambda h: check_h2_count(h)),
            ("ira_allocation",     lambda h: check_ira_allocation(h, ira_alloc)),
            ("obbba_banner",       lambda h: check_obbba_callout_banner(h)),
            ("card_discipline",    lambda h: check_card_design_discipline(h)),
        ]

        page_failures = []
        for name, fn_check in runs:
            for f in fn_check(html):
                page_failures.append((name, f))

        if not page_failures:
            perfect.append(slug)
        else:
            failed.append(slug)
            for name, msg in page_failures:
                failure_pattern.setdefault(name, []).append((slug, msg))

    print(f"\n{'=' * 100}\nRESULT: {len(perfect)} clean / {len(failed)} with failures, "
          f"out of {len(files)} total\n{'=' * 100}\n")

    if failure_pattern:
        for name, fails in sorted(failure_pattern.items(), key=lambda kv: -len(kv[1])):
            print(f"\n[{name}] {len(fails)} state hubs failed:")
            for slug, msg in fails[:5]:
                print(f"  {slug}: {msg}")
            if len(fails) > 5:
                print(f"  ... and {len(fails) - 5} more")

    if failed:
        print(f"\n{len(failed)} state hubs failed. Exit 1.")
        sys.exit(1)
    else:
        print("\nAll state hubs pass QC. Exit 0.")
        sys.exit(0)


if __name__ == "__main__":
    main()
