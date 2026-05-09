"""
Phase 1 / Step 2 of the programmatic-uniqueness uplift plan.

Populates 5 Census ACS-sourced columns in cities_updated.xlsx for all 115
published cities, using the U.S. Census Bureau's free ACS 5-year API
(no token required).

Dataset: ACS 5-year estimates 2019-2023 (most recent available in 2026).
Endpoint: https://api.census.gov/data/2023/acs/acs5

Fields fetched per city:
  acs_median_home_year       — B25035_001E   (median year structure built)
  acs_heating_fuel_pct_gas   — B25040_002E / _001E (utility gas heat %)
  acs_heating_fuel_pct_electric — B25040_004E / _001E (electric heat %)
  acs_owner_occupied_pct     — B25003_002E / _001E (owner-occupied %)
  acs_median_home_value      — B25077_001E   (median home value, $)

Mapping source: scripts/city_data_map.json (census_place_fips + census_state_fips).

Usage:
  python scripts/fetch_census_acs.py                     # all 115 cities
  python scripts/fetch_census_acs.py --slug atlanta-ga   # one city
  python scripts/fetch_census_acs.py --dry-run           # fetch but do not write xlsx
  python scripts/fetch_census_acs.py --verify atlanta-ga # show fetched values + URL

The Census API's /data/2023/acs/acs5 endpoint returns JSON arrays. First row is
header names, subsequent rows are data. Single-place queries return one row.
"""
from __future__ import annotations

import argparse
import json
import ssl
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
CITY_MAP = REPO / "scripts" / "city_data_map.json"
XLSX = REPO / "cities_updated.xlsx"
USER_AGENT = "CoolCallPro-Census/1.0 (gyanesh.gulshan@gmail.com)"

try:
    import certifi
    _SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except ImportError:
    _SSL_CTX = ssl.create_default_context()


CENSUS_API = "https://api.census.gov/data/2023/acs/acs5"

# Variables we request from the Census API for each place
CENSUS_VARIABLES = [
    "B25035_001E",  # Median year structure built
    "B25077_001E",  # Median home value
    "B25003_001E",  # Total occupied housing units (tenure denominator)
    "B25003_002E",  # Owner-occupied
    "B25040_001E",  # Total housing units with heating (fuel denominator)
    "B25040_002E",  # Utility gas
    "B25040_004E",  # Electricity
]

# Output column names (xlsx) — order matches downstream prose generator
OUTPUT_COLUMNS = [
    "acs_median_home_year",
    "acs_heating_fuel_pct_gas",
    "acs_heating_fuel_pct_electric",
    "acs_owner_occupied_pct",
    "acs_median_home_value",
]


def fetch_place(state_fips: str, place_fips: str, allow_unverified: bool = False) -> dict | None:
    """Fetch ACS variables for one Census place. Returns dict {var: raw_value}
    or None if no data."""
    state_padded = state_fips.zfill(2)
    place_padded = place_fips.zfill(5)
    url = (
        f"{CENSUS_API}?get=NAME,{','.join(CENSUS_VARIABLES)}"
        f"&for=place:{place_padded}&in=state:{state_padded}"
    )
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    contexts = [_SSL_CTX]
    if allow_unverified:
        contexts.append(ssl._create_unverified_context())
    last_err = None
    for ctx in contexts:
        try:
            with urllib.request.urlopen(req, timeout=30, context=ctx) as r:
                data = json.loads(r.read())
            if not isinstance(data, list) or len(data) < 2:
                return None
            header, row = data[0], data[1]
            return dict(zip(header, row))
        except (urllib.error.URLError, ssl.SSLError) as e:
            last_err = e
            continue
        except (json.JSONDecodeError, urllib.error.HTTPError) as e:
            return None  # API returned non-JSON or 404 → genuine miss
    raise RuntimeError(f"fetch failed for state={state_fips} place={place_fips}: {last_err}")


def normalize_for_city(raw: dict | None) -> dict:
    """Compute the 5 output columns from the Census API response.

    Census sentinels:
      "-666666666" / "" / None: data not available for this geography
      Negative dollar amounts (rare): suppressed for privacy

    Returns dict with all 5 columns; values may be None where Census data is
    unavailable.
    """
    out = {col: None for col in OUTPUT_COLUMNS}
    if not raw:
        return out

    def parse_int(v):
        if v is None or v == "" or v == "-666666666":
            return None
        try:
            n = int(v)
        except (ValueError, TypeError):
            return None
        return n if n >= 0 else None

    median_year = parse_int(raw.get("B25035_001E"))
    median_value = parse_int(raw.get("B25077_001E"))
    tenure_total = parse_int(raw.get("B25003_001E"))
    owner = parse_int(raw.get("B25003_002E"))
    fuel_total = parse_int(raw.get("B25040_001E"))
    fuel_gas = parse_int(raw.get("B25040_002E"))
    fuel_elec = parse_int(raw.get("B25040_004E"))

    out["acs_median_home_year"] = median_year
    out["acs_median_home_value"] = median_value

    if tenure_total and tenure_total > 0 and owner is not None:
        out["acs_owner_occupied_pct"] = round(owner / tenure_total * 100, 1)
    if fuel_total and fuel_total > 0:
        if fuel_gas is not None:
            out["acs_heating_fuel_pct_gas"] = round(fuel_gas / fuel_total * 100, 1)
        if fuel_elec is not None:
            out["acs_heating_fuel_pct_electric"] = round(fuel_elec / fuel_total * 100, 1)

    return out


def fetch_for_city(slug: str, mapping: dict, allow_unverified: bool = False) -> dict:
    entry = mapping[slug]
    state_fips = entry.get("census_state_fips")
    place_fips = entry.get("census_place_fips")
    if not (state_fips and place_fips):
        raise ValueError(f"{slug} missing census_state_fips or census_place_fips")
    raw = fetch_place(state_fips, place_fips, allow_unverified=allow_unverified)
    out = normalize_for_city(raw)
    out["__name__"] = (raw or {}).get("NAME", "(no Census data)")
    out["__place__"] = f"state:{state_fips} place:{place_fips}"
    return out


def write_xlsx_columns(results: dict[str, dict]) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    header = [c.value for c in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(header) if h}

    next_col = ws.max_column + 1
    for col in OUTPUT_COLUMNS:
        if col not in col_index:
            ws.cell(row=1, column=next_col, value=col)
            col_index[col] = next_col
            next_col += 1

    slug_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        city = ws.cell(row=row_idx, column=1).value
        state = ws.cell(row=row_idx, column=2).value
        if not city or not state:
            continue
        slug = f"{str(city).lower().replace(' ', '-').replace('.', '').replace(chr(39), '')}-{str(state).lower()}"
        slug_to_row[slug] = row_idx

    written, missing = 0, []
    for slug, vals in results.items():
        row = slug_to_row.get(slug)
        if not row:
            missing.append(slug)
            continue
        for col_name in OUTPUT_COLUMNS:
            ws.cell(row=row, column=col_index[col_name], value=vals.get(col_name))
        written += 1

    wb.save(XLSX)
    print(f"  wrote {written} rows; missing {len(missing)}: {missing[:5]}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--slug", help="Fetch only this city")
    parser.add_argument("--verify", help="Show fetched values + source URL for one city")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--allow-unverified-ssl", action="store_true")
    parser.add_argument("--sleep", type=float, default=0.4)
    args = parser.parse_args()

    if not CITY_MAP.exists():
        print(f"ERROR: {CITY_MAP} not found.", file=sys.stderr)
        return 2
    mapping = json.loads(CITY_MAP.read_text(encoding="utf-8"))

    if args.verify:
        slug = args.verify
        if slug not in mapping:
            print(f"ERROR: {slug} not in city_data_map.json", file=sys.stderr)
            return 2
        e = mapping[slug]
        url = (
            f"{CENSUS_API}?get=NAME,{','.join(CENSUS_VARIABLES)}"
            f"&for=place:{(e.get('census_place_fips') or '').zfill(5)}"
            f"&in=state:{(e.get('census_state_fips') or '').zfill(2)}"
        )
        print(f"City:        {slug}")
        print(f"State FIPS:  {e.get('census_state_fips')}")
        print(f"Place FIPS:  {e.get('census_place_fips')}")
        print(f"Source URL:  {url}")
        print()
        vals = fetch_for_city(slug, mapping, allow_unverified=args.allow_unverified_ssl)
        print(f"Place name:  {vals.pop('__name__')}")
        vals.pop("__place__", None)
        for k, v in vals.items():
            print(f"  {k:35s} = {v}")
        return 0

    slugs = [args.slug] if args.slug else list(mapping.keys())
    print(f"Fetching Census ACS for {len(slugs)} cities...")
    results = {}
    failures = []
    for i, slug in enumerate(slugs, start=1):
        try:
            vals = fetch_for_city(slug, mapping, allow_unverified=args.allow_unverified_ssl)
            results[slug] = vals
            present = sum(1 for k, v in vals.items() if not k.startswith("__") and v is not None)
            print(f"  [{i:3d}/{len(slugs)}] {slug:25s} {present}/5 fields")
        except Exception as e:
            failures.append((slug, str(e)))
            print(f"  [{i:3d}/{len(slugs)}] {slug:25s} FAILED: {e}")
        if i < len(slugs):
            time.sleep(args.sleep)

    print(f"\nFetched: {len(results)} / {len(slugs)}")
    print(f"Failed:  {len(failures)}")
    for slug, err in failures[:10]:
        print(f"  {slug}: {err}")

    if args.dry_run:
        print("\n--dry-run: not writing xlsx")
        return 0 if not failures else 1

    if results:
        print(f"\nWriting columns to {XLSX} ...")
        write_xlsx_columns(results)
    return 0 if not failures else 1


if __name__ == "__main__":
    sys.exit(main())
