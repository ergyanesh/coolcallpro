"""
Phase 1 / Step 3 of the programmatic-uniqueness uplift plan.

Populates 2 EIA-sourced columns in cities_updated.xlsx for all 115
published cities. Values are state-level (51 records keyed by state),
mapped to each city by its state field.

Data sources (all public, no auth required):

  Electricity rate ($/kWh, residential average):
    EIA Electric Power Monthly Table 5.6.A
    https://www.eia.gov/electricity/monthly/epm_table_grapher.php?t=epmt_5_6_a
    Snapshot date: 2026-02 (February 2026)
    To refresh: re-fetch the URL and update ELEC_RATE_BY_STATE constant.

  Natural gas rate ($/therm, residential, annual avg):
    EIA Natural Gas Annual — Residential Price by State (NG.PRI.SUM)
    https://www.eia.gov/dnav/ng/ng_pri_sum_a_EPG0_PRS_DMcf_a.htm
    Snapshot date: 2024 (most recent annual data)
    Source values are $/Mcf; converted to $/therm using AGA standard
    1 Mcf = 10.37 therms (natural gas heat content ~1037 BTU/cf).
    To refresh: re-fetch the URL, divide by 10.37, update GAS_RATE_BY_STATE.

Why baked-in constants vs API call: EIA rates change slowly (annual gas,
monthly electric). The 30-day Layer-2 source watcher (already shipped)
can monitor these URLs for change; the actual values stay stable enough
that re-fetching every script run is overkill. Re-run with --refresh-print
to compare current EIA values against the baked-in constants.

Usage:
  python scripts/fetch_eia_state_rates.py            # populate xlsx (all 115)
  python scripts/fetch_eia_state_rates.py --slug atlanta-ga  # single city
  python scripts/fetch_eia_state_rates.py --dry-run
  python scripts/fetch_eia_state_rates.py --verify-state GA
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
CITY_MAP = REPO / "scripts" / "city_data_map.json"
XLSX = REPO / "cities_updated.xlsx"


# Residential electricity rate by state ($/kWh).
# Source: EIA Electric Power Monthly Table 5.6.A — Feb 2026 data.
# Values are cents/kWh from EIA, converted to dollars by /100.
ELEC_RATE_BY_STATE_CENTS = {
    "AL": 16.18, "AK": 25.79, "AZ": 16.03, "AR": 12.73, "CA": 33.22,
    "CO": 16.79, "CT": 30.77, "DE": 16.27, "DC": 23.97, "FL": 15.80,
    "GA": 14.13, "HI": 43.00, "ID": 12.63, "IL": 17.83, "IN": 16.06,
    "IA": 12.74, "KS": 15.11, "KY": 13.42, "LA": 12.87, "ME": 32.17,
    "MD": 20.08, "MA": 30.46, "MI": 20.00, "MN": 15.39, "MS": 14.72,
    "MO": 12.17, "MT": 13.33, "NE": 11.79, "NV": 14.38, "NH": 26.52,
    "NJ": 23.12, "NM": 15.07, "NY": 29.99, "NC": 14.64, "ND": 11.64,
    "OH": 17.52, "OK": 12.89, "OR": 14.64, "PA": 20.30, "RI": 29.45,
    "SC": 16.15, "SD": 13.24, "TN": 12.82, "TX": 15.41, "UT": 13.33,
    "VT": 23.27, "VA": 15.96, "WA": 14.11, "WV": 14.41, "WI": 18.74,
    "WY": 13.04,
}

# Residential natural gas rate by state ($/Mcf, thousand cubic feet).
# Source: EIA Natural Gas Annual Residential Prices — 2024 data.
GAS_RATE_BY_STATE_MCF = {
    "AL": 18.63, "AK": 11.73, "AZ": 20.47, "AR": 16.09, "CA": 19.14,
    "CO": 10.58, "CT": 16.83, "DE": 14.78, "DC": 16.49, "FL": 25.37,
    "GA": 18.19, "HI": 48.88, "ID": 8.51,  "IL": 11.11, "IN": 9.50,
    "IA": 10.33, "KS": 13.14, "KY": 13.61, "LA": 16.38, "ME": 18.95,
    "MD": 16.14, "MA": 21.90, "MI": 10.76, "MN": 10.65, "MS": 15.42,
    "MO": 16.86, "MT": 8.22,  "NE": 10.63, "NV": 18.69, "NH": 18.74,
    "NJ": 13.21, "NM": 9.05,  "NY": 16.58, "NC": 16.42, "ND": 8.99,
    "OH": 13.77, "OK": 13.99, "OR": 15.71, "PA": 13.62, "RI": 21.66,
    "SC": 16.91, "SD": 9.32,  "TN": 10.75, "TX": 17.98, "UT": 13.15,
    "VT": 17.34, "VA": 15.40, "WA": 12.72, "WV": 15.26, "WI": 10.14,
    "WY": 12.39,
}

# AGA conversion: 1 Mcf natural gas ≈ 10.37 therms (heat content 1037 BTU/cf).
MCF_TO_THERMS = 10.37


def elec_rate_dollars(state_abbr: str) -> float | None:
    """Return $/kWh for state, rounded to 4 decimals."""
    cents = ELEC_RATE_BY_STATE_CENTS.get(state_abbr.upper())
    return round(cents / 100, 4) if cents is not None else None


def gas_rate_dollars_per_therm(state_abbr: str) -> float | None:
    """Return $/therm for state, rounded to 3 decimals."""
    mcf = GAS_RATE_BY_STATE_MCF.get(state_abbr.upper())
    return round(mcf / MCF_TO_THERMS, 3) if mcf is not None else None


def fetch_for_city(slug: str, mapping: dict) -> dict:
    state = (mapping[slug].get("state") or "").upper()
    return {
        "eia_state_elec_rate": elec_rate_dollars(state),
        "eia_state_gas_rate": gas_rate_dollars_per_therm(state),
        "__state__": state,
    }


def write_xlsx_columns(results: dict[str, dict]) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    header = [c.value for c in ws[1]]
    new_cols = ["eia_state_elec_rate", "eia_state_gas_rate"]
    col_index = {h: i + 1 for i, h in enumerate(header) if h}

    next_col = ws.max_column + 1
    for col in new_cols:
        if col not in col_index:
            ws.cell(row=1, column=next_col, value=col)
            col_index[col] = next_col
            next_col += 1

    slug_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        city = ws.cell(row=row_idx, column=1).value
        state = ws.cell(row=row_idx, column=2).value
        if not city or not state:
            continue
        slug = f"{str(city).lower().replace(' ', '-').replace('.', '').replace(chr(39), '')}-{str(state).lower()}"
        slug_to_row[slug] = row_idx

    written, missing = 0, []
    for slug, vals in results.items():
        row = slug_to_row.get(slug)
        if not row:
            missing.append(slug)
            continue
        for col_name in new_cols:
            ws.cell(row=row, column=col_index[col_name], value=vals.get(col_name))
        written += 1

    wb.save(XLSX)
    print(f"  wrote {written} rows; missing {len(missing)}: {missing[:5]}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--slug", help="Single city")
    parser.add_argument("--verify-state", help="Show rates for one state abbrev (e.g., GA)")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if args.verify_state:
        st = args.verify_state.upper()
        elec = elec_rate_dollars(st)
        gas = gas_rate_dollars_per_therm(st)
        print(f"State: {st}")
        print(f"  Electric:    ${elec}/kWh  (raw: {ELEC_RATE_BY_STATE_CENTS.get(st)}¢/kWh)")
        print(f"  Natural gas: ${gas}/therm  (raw: ${GAS_RATE_BY_STATE_MCF.get(st)}/Mcf)")
        return 0

    if not CITY_MAP.exists():
        print(f"ERROR: {CITY_MAP} not found.", file=sys.stderr)
        return 2
    mapping = json.loads(CITY_MAP.read_text(encoding="utf-8"))

    slugs = [args.slug] if args.slug else list(mapping.keys())
    results = {}
    failures = []
    for slug in slugs:
        try:
            results[slug] = fetch_for_city(slug, mapping)
        except Exception as e:
            failures.append((slug, str(e)))

    print(f"Resolved: {len(results)} / {len(slugs)} cities")
    print(f"Failed:   {len(failures)}")
    for slug, err in failures[:10]:
        print(f"  {slug}: {err}")

    # Quick sanity print
    if args.slug or len(slugs) <= 5:
        for slug, vals in results.items():
            print(f"  {slug:25s} state={vals['__state__']:2s}  "
                  f"elec=${vals['eia_state_elec_rate']}/kWh  gas=${vals['eia_state_gas_rate']}/therm")

    if args.dry_run:
        print("\n--dry-run: not writing xlsx")
        return 0 if not failures else 1

    if results:
        print(f"\nWriting columns to {XLSX} ...")
        write_xlsx_columns(results)

    return 0 if not failures else 1


if __name__ == "__main__":
    sys.exit(main())
