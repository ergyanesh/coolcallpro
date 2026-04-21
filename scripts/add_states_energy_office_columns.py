"""
One-off: add "State Energy Office Name" and "State Energy Office URL" columns
to states.xlsx.

All 51 values are read from state_energy_offices.json (see that file for
provenance: NASEO state-energy-offices directory, fetched 2026-04-21, with
URL corrections for 3 stale entries documented in _notes).

After this script runs once, states.xlsx is the canonical source-of-truth
for per-state energy office name + URL. The JSON file remains the vocabulary
/ notes reference consumed by the HTML patcher and generator template.

Idempotent: if both columns already exist, the script verifies values match
and exits without rewriting the file. To force a refresh, delete the columns
manually first.
"""

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
STATES_XLSX = ROOT / "states.xlsx"
OFFICES_JSON = ROOT / "state_energy_offices.json"

NAME_COL = "State Energy Office Name"
URL_COL = "State Energy Office URL"


def main():
    with OFFICES_JSON.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    offices = data["offices"]  # {abbr: {name, url}}

    wb = openpyxl.load_workbook(STATES_XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    abbr_idx = headers.index("State Abbreviation") + 1
    name_col_exists = NAME_COL in headers
    url_col_exists = URL_COL in headers

    if name_col_exists and url_col_exists:
        print(f"Both columns already exist. Verifying values...")
        n_col = headers.index(NAME_COL) + 1
        u_col = headers.index(URL_COL) + 1
        mismatches = []
        for row_i in range(2, ws.max_row + 1):
            abbr = ws.cell(row=row_i, column=abbr_idx).value
            if not abbr:
                continue
            expected = offices.get(abbr, {})
            actual_name = ws.cell(row=row_i, column=n_col).value or ""
            actual_url = ws.cell(row=row_i, column=u_col).value or ""
            if actual_name.strip() != expected.get("name", "").strip():
                mismatches.append((abbr, "name", actual_name, expected.get("name")))
            if actual_url.strip() != expected.get("url", "").strip():
                mismatches.append((abbr, "url", actual_url, expected.get("url")))
        if mismatches:
            print(f"  {len(mismatches)} mismatches:")
            for abbr, field, have, want in mismatches[:5]:
                print(f"    {abbr}/{field}: have='{have}' want='{want}'")
            print("  Not overwriting. Delete the columns manually to regenerate.")
            return 1
        print(f"  All values match. No change.")
        return 0

    if name_col_exists or url_col_exists:
        print("ERROR: Only one of the two columns exists. Manual cleanup required.")
        return 2

    n_col = len(headers) + 1
    u_col = len(headers) + 2
    print(f"Adding '{NAME_COL}' at column {n_col}, '{URL_COL}' at column {u_col}")

    ws.cell(row=1, column=n_col, value=NAME_COL)
    ws.cell(row=1, column=u_col, value=URL_COL)

    written = 0
    missing = []
    for row_i in range(2, ws.max_row + 1):
        abbr = ws.cell(row=row_i, column=abbr_idx).value
        if not abbr:
            continue
        entry = offices.get(abbr)
        if not entry:
            missing.append(abbr)
            continue
        ws.cell(row=row_i, column=n_col, value=entry["name"])
        ws.cell(row=row_i, column=u_col, value=entry["url"])
        written += 1

    wb.save(STATES_XLSX)
    print(f"  wrote {written} rows")
    if missing:
        print(f"  missing data for: {missing}")
        return 1
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
