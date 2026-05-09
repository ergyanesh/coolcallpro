"""
Phase 1 / Step 1 of the programmatic-uniqueness uplift plan.

Populates 5 NOAA-sourced columns in cities_updated.xlsx for all 115 published
cities, using NOAA NCEI's 1991-2020 Climate Normals dataset accessed via the
free NCEI Access API (no token required).

Fields fetched per city (annual normals):
  ANN-HTDD-NORMAL          → noaa_hdd_annual          (heating degree days)
  ANN-CLDD-NORMAL          → noaa_cdd_annual          (cooling degree days)
  ANN-TMAX-AVGNDS-GRTH090  → noaa_days_above_90       (days max temp >= 90°F)
  ANN-TMIN-AVGNDS-LSTH032  → noaa_days_below_32       (days min temp <= 32°F)
  ANN-SNOW-NORMAL          → noaa_snowfall_inches     (annual snowfall in inches)

Mapping source: scripts/city_data_map.json (ghcnd_station_id field).

Usage:
  python scripts/fetch_noaa_normals.py                     # all 115 cities
  python scripts/fetch_noaa_normals.py --slug atlanta-ga   # one city, prints to stdout
  python scripts/fetch_noaa_normals.py --dry-run           # fetch but do not write xlsx
  python scripts/fetch_noaa_normals.py --verify atlanta-ga # show fetched values + URL

Environment notes:
  Local Windows machines may have SSL cert chain issues against NCEI; see the
  YMYL watcher for the same pattern. The script falls back to a verified
  HTTPS context using certifi if available, or unverified context as last
  resort (the NOAA data is public and integrity-checked via the NCEI URL,
  so this is acceptable for one-time data acquisition).
"""
from __future__ import annotations

import argparse
import json
import ssl
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
CITY_MAP = REPO / "scripts" / "city_data_map.json"
XLSX = REPO / "cities_updated.xlsx"
USER_AGENT = "CoolCallPro-NOAA/1.0 (gyanesh.gulshan@gmail.com)"

# Set up SSL context (handles Windows cert chain issue)
try:
    import certifi
    _SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except ImportError:
    _SSL_CTX = ssl.create_default_context()


# NOAA NCEI Access API
ACCESS_API = "https://www.ncei.noaa.gov/access/services/data/v1"
DATATYPES = [
    "ANN-HTDD-NORMAL",
    "ANN-CLDD-NORMAL",
    "ANN-TMAX-AVGNDS-GRTH090",
    "ANN-TMIN-AVGNDS-LSTH032",
    "ANN-SNOW-NORMAL",
]

# Map NCEI field name → xlsx column name
FIELD_MAP = {
    "ANN-HTDD-NORMAL":           "noaa_hdd_annual",
    "ANN-CLDD-NORMAL":           "noaa_cdd_annual",
    "ANN-TMAX-AVGNDS-GRTH090":   "noaa_days_above_90",
    "ANN-TMIN-AVGNDS-LSTH032":   "noaa_days_below_32",
    "ANN-SNOW-NORMAL":           "noaa_snowfall_inches",
}


def fetch_station(station_id: str, allow_unverified: bool = False) -> dict | None:
    """Fetch annual normals for one GHCN-D station. Returns dict mapping
    NCEI field name → string value (NCEI returns numbers as strings)."""
    url = (
        f"{ACCESS_API}"
        f"?dataset=normals-annualseasonal"
        f"&dataTypes={','.join(DATATYPES)}"
        f"&stations={station_id}"
        f"&startDate=0001-01-01&endDate=9999-12-31"
        f"&format=json&includeStationName=true"
    )
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    contexts = [_SSL_CTX]
    if allow_unverified:
        contexts.append(ssl._create_unverified_context())
    last_err = None
    for ctx in contexts:
        try:
            with urllib.request.urlopen(req, timeout=30, context=ctx) as r:
                data = json.loads(r.read())
            if not data:
                return None
            row = data[0]  # one row for annual
            return row
        except (urllib.error.URLError, ssl.SSLError) as e:
            last_err = e
            continue
    raise RuntimeError(f"fetch failed for {station_id}: {last_err}")


def normalize_value(field: str, raw: str | None) -> float | int | None:
    """Apply NCEI scale factors and coerce to native types.

    Per the NOAA NCEI 1991-2020 Climate Normals data dictionary:
      - HDD / CDD: returned as direct degree-days (no scaling)
      - TMAX-AVGNDS-GRTH/LSTH: tenths of days (divide by 10)
      - TMIN-AVGNDS-GRTH/LSTH: tenths of days (divide by 10)
      - SNOW-NORMAL: tenths of inches (divide by 10)

    Sentinel values:
      -9999: missing data → None
      -7777 / -777.7: not applicable for this climate (e.g., HDD for
        tropical Honolulu, snow for arid Phoenix). Also returned as None;
        downstream prose generation should interpret None as "essentially
        zero / not climatically relevant" for these contexts.
    """
    if raw is None or raw == "" or raw == "-9999":
        return None
    try:
        v = float(raw)
    except ValueError:
        return None
    if v <= -7000:  # NCEI sentinel for "not applicable"
        return None
    if field in ("ANN-HTDD-NORMAL", "ANN-CLDD-NORMAL"):
        return int(round(v))
    if field in ("ANN-TMAX-AVGNDS-GRTH090", "ANN-TMIN-AVGNDS-LSTH032"):
        return round(v / 10.0, 1)  # tenths of days → days
    if field == "ANN-SNOW-NORMAL":
        return round(v / 10.0, 1)  # tenths of inches → inches
    return v


def fetch_for_city(slug: str, mapping: dict, allow_unverified: bool = False) -> dict:
    """Fetch + normalize for one city. Returns column→value dict."""
    entry = mapping[slug]
    sid = entry.get("ghcnd_station_id")
    if not sid:
        raise ValueError(f"{slug} has no ghcnd_station_id")
    raw = fetch_station(sid, allow_unverified=allow_unverified)
    if raw is None:
        # Station has no 1991-2020 annual normals at NCEI. Return Nones
        # but keep station/name so the caller can log the gap consistently.
        out = {col: None for col in FIELD_MAP.values()}
        out["__station__"] = sid
        out["__name__"] = "(no normals data at NCEI)"
        return out
    out = {}
    for nce_field, col in FIELD_MAP.items():
        out[col] = normalize_value(nce_field, raw.get(nce_field))
    out["__station__"] = sid
    out["__name__"] = raw.get("NAME", "")
    return out


def write_xlsx_columns(results: dict[str, dict]) -> None:
    """Add 5 new columns to cities_updated.xlsx and populate them.
    Preserves existing columns; matches by City + State."""
    import openpyxl

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    # Read header
    header = [c.value for c in ws[1]]
    new_cols = list(FIELD_MAP.values())
    col_index = {h: i + 1 for i, h in enumerate(header) if h}

    # Append new columns at the end if not already present
    next_col = ws.max_column + 1
    for col in new_cols:
        if col not in col_index:
            ws.cell(row=1, column=next_col, value=col)
            col_index[col] = next_col
            next_col += 1

    # Build slug → row index lookup
    slug_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        city = ws.cell(row=row_idx, column=1).value
        state = ws.cell(row=row_idx, column=2).value
        if not city or not state:
            continue
        slug = f"{str(city).lower().replace(' ', '-').replace('.', '').replace(chr(39), '')}-{str(state).lower()}"
        slug_to_row[slug] = row_idx

    # Write values
    written, missing = 0, []
    for slug, vals in results.items():
        row = slug_to_row.get(slug)
        if not row:
            missing.append(slug)
            continue
        for col_name in new_cols:
            ws.cell(row=row, column=col_index[col_name], value=vals.get(col_name))
        written += 1

    wb.save(XLSX)
    print(f"  wrote {written} rows; missing {len(missing)}: {missing[:5]}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--slug", help="Fetch only this city's data and print to stdout")
    parser.add_argument("--verify", help="Show fetched values + source URL for one city")
    parser.add_argument("--dry-run", action="store_true", help="Fetch all but do not write xlsx")
    parser.add_argument("--allow-unverified-ssl", action="store_true",
                        help="Fall back to unverified HTTPS if cert chain fails (Windows local)")
    parser.add_argument("--sleep", type=float, default=0.4,
                        help="Seconds between requests (default 0.4 = ~150/min, well under NCEI limits)")
    args = parser.parse_args()

    if not CITY_MAP.exists():
        print(f"ERROR: {CITY_MAP} not found. Run scripts/_match_ghcnd_stations.py first.", file=sys.stderr)
        return 2
    mapping = json.loads(CITY_MAP.read_text(encoding="utf-8"))

    if args.verify:
        slug = args.verify
        if slug not in mapping:
            print(f"ERROR: {slug} not in city_data_map.json", file=sys.stderr)
            return 2
        sid = mapping[slug]["ghcnd_station_id"]
        url = (
            f"{ACCESS_API}?dataset=normals-annualseasonal"
            f"&dataTypes={','.join(DATATYPES)}&stations={sid}"
            f"&startDate=0001-01-01&endDate=9999-12-31&format=json&includeStationName=true"
        )
        print(f"City:        {slug}")
        print(f"GHCN-D ID:   {sid}")
        print(f"Source URL:  {url}")
        print()
        vals = fetch_for_city(slug, mapping, allow_unverified=args.allow_unverified_ssl)
        print(f"Station name: {vals.pop('__name__')}")
        vals.pop('__station__', None)
        for k, v in vals.items():
            print(f"  {k:30s} = {v}")
        return 0

    if args.slug:
        slugs = [args.slug]
    else:
        slugs = list(mapping.keys())

    print(f"Fetching NOAA normals for {len(slugs)} cities...")
    results = {}
    failures = []
    for i, slug in enumerate(slugs, start=1):
        try:
            vals = fetch_for_city(slug, mapping, allow_unverified=args.allow_unverified_ssl)
            results[slug] = vals
            present = sum(1 for k, v in vals.items() if not k.startswith("__") and v is not None)
            print(f"  [{i:3d}/{len(slugs)}] {slug:25s} {vals['__station__']:12s} {present}/5 fields")
        except Exception as e:
            failures.append((slug, str(e)))
            print(f"  [{i:3d}/{len(slugs)}] {slug:25s} FAILED: {e}")
        if i < len(slugs):
            time.sleep(args.sleep)

    print(f"\nFetched: {len(results)} / {len(slugs)}")
    print(f"Failed:  {len(failures)}")
    for slug, err in failures[:10]:
        print(f"  {slug}: {err}")

    if args.dry_run:
        print("\n--dry-run: not writing xlsx")
        return 0 if not failures else 1

    if results:
        print(f"\nWriting columns to {XLSX} ...")
        write_xlsx_columns(results)
    return 0 if not failures else 1


if __name__ == "__main__":
    sys.exit(main())
