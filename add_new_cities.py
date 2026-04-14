"""Add 20 new cities (12 remaining states) to cities_updated.xlsx"""
import openpyxl

wb = openpyxl.load_workbook('cities_updated.xlsx')
ws = wb.active

# [City, State, AC Cost, Furnace Cost, Neighborhoods, ZIPs, Permit Office, Utility, Pop, Summer High, Winter Low, SEER, Rebates, License, Climate Zone, Published, Research Done]
new_cities = [
    # Hawaii
    ["Honolulu", "HI", "$5,000\u2013$10,000", "$2,500\u2013$4,500",
     "Kahala, Kaimuki, Manoa, Hawaii Kai, Kakaako",
     "96817, 96818, 96822, 96816, 96819",
     "City and County of Honolulu Department of Planning and Permitting",
     "Hawaiian Electric (HECO)", "344967", "89", "66",
     "15 SEER2 (South Region)",
     "Hawaii Energy: rebates on cooling systems, smart HVAC controls (up to 50% of cost, capped at $5,000)",
     "Hawaii DCCA Contractors License Board \u2014 C-52 Air Conditioning and Warm Air Heating license required; 4 years supervisory experience + Trade and Business & Law exams",
     "Zone 1A (Very Hot-Humid)", "Yes", "Yes"],

    # South Carolina
    ["Charleston", "SC", "$3,500\u2013$8,000", "$3,500\u2013$7,000",
     "South of Broad, Mount Pleasant, James Island, West Ashley, Harleston Village",
     "29407, 29412, 29414, 29403, 29406",
     "City of Charleston Permit Center",
     "Dominion Energy South Carolina", "162499", "91", "39",
     "15 SEER2 (South Region)",
     "Dominion Energy SC: $400\u2013$500 rebate on ENERGY STAR central AC/heat pumps, up to $300 for ductwork improvements",
     "SC Dept. of Labor, Licensing and Regulation \u2014 Residential Builders Commission (residential) or Contractor\u2019s Licensing Board (commercial); exam + 1\u20132 years experience required",
     "Zone 3A (Warm-Humid)", "Yes", "Yes"],

    ["Columbia", "SC", "$3,500\u2013$8,000", "$3,500\u2013$7,000",
     "Shandon, Forest Acres, Rosewood, Downtown Columbia, Seven Oaks",
     "29205, 29206, 29209, 29210, 29212",
     "City of Columbia Planning & Development, Development Center",
     "Dominion Energy South Carolina", "138019", "92", "38",
     "15 SEER2 (South Region)",
     "Dominion Energy SC: $400\u2013$500 rebate on ENERGY STAR central AC/heat pumps, up to $300 for ductwork improvements",
     "SC Dept. of Labor, Licensing and Regulation \u2014 same statewide requirement",
     "Zone 3A (Warm-Humid)", "Yes", "Yes"],

    ["Greenville", "SC", "$3,500\u2013$7,500", "$3,500\u2013$7,000",
     "Alta Vista, Augusta Road, North Main, Parkins Mill, Overbrook",
     "29601, 29605, 29607, 29609, 29615",
     "City of Greenville Building & Permit Center",
     "Duke Energy Carolinas", "74371", "90", "33",
     "15 SEER2 (South Region)",
     "Duke Energy: up to $1,000 on qualifying heat pump or AC units, up to $500 for older system replacement",
     "SC Dept. of Labor, Licensing and Regulation \u2014 same statewide requirement",
     "Zone 3A (Warm-Humid)", "Yes", "Yes"],

    # Mississippi
    ["Jackson", "MS", "$3,500\u2013$7,500", "$3,000\u2013$6,500",
     "Belhaven, Fondren, Northeast Jackson, Meadowbrook, Belhaven Heights",
     "39206, 39211, 39212, 39209, 39202",
     "City of Jackson Building Permit Office",
     "Entergy Mississippi", "141449", "92", "41",
     "15 SEER2 (South Region)",
     "Entergy Mississippi: up to $1,100 on high-efficiency heat pump or AC, free AC tune-up per account, duct sealing every 5 years",
     "Mississippi State Board of Contractors (MSBOC) \u2014 license required for HVAC work $10,000+; law/business and HVAC exams required",
     "Zone 3A (Warm-Humid)", "Yes", "Yes"],

    ["Gulfport", "MS", "$3,500\u2013$7,500", "$3,000\u2013$6,500",
     "Bayou View, Orange Grove, North Gulfport, Magnolia Grove, The Island",
     "39501, 39503, 39507, 39505, 39506",
     "City of Gulfport Urban Development Building Code Services",
     "Mississippi Power", "74621", "90", "46",
     "15 SEER2 (South Region)",
     "Mississippi Power: energy efficiency rebates for qualifying HVAC equipment",
     "Mississippi State Board of Contractors (MSBOC) \u2014 same statewide requirement",
     "Zone 2A (Hot-Humid)", "Yes", "Yes"],

    # Maine
    ["Portland", "ME", "$4,000\u2013$8,000", "$4,000\u2013$8,500",
     "Munjoy Hill, West End, Deering Center, Back Cove, Woodfords Corner",
     "04103, 04102, 04101, 04106, 04108",
     "City of Portland Permitting and Inspections Department",
     "Central Maine Power (CMP)", "69568", "79", "13",
     "14 SEER2 (North Region)",
     "Efficiency Maine Trust: $1,000 per heat pump unit (standard income, max 3 units), $2,000/unit (moderate income), $3,000/unit (low income)",
     "Maine Fuel Board licenses technicians for oil, solid fuel, propane, and natural gas heating; companies must have licensed Master Technician",
     "Zone 6A (Cold-Humid)", "Yes", "Yes"],

    ["Bangor", "ME", "$4,000\u2013$8,000", "$4,000\u2013$8,500",
     "Downtown Bangor, Fairmount, Essex Street, Union Street, Broadway",
     "04401, 04402, 04410, 04412, 04414",
     "City of Bangor Code Enforcement Office",
     "Versant Power", "31753", "77", "10",
     "14 SEER2 (North Region)",
     "Efficiency Maine Trust: $1,000\u2013$3,000 per heat pump unit depending on income",
     "Maine Fuel Board \u2014 same statewide requirement",
     "Zone 6A (Cold-Humid)", "Yes", "Yes"],

    # West Virginia
    ["Charleston", "WV", "$3,500\u2013$7,500", "$3,500\u2013$7,500",
     "South Hills, Kanawha City, East End, Edgewood, Greenbrier",
     "25304, 25314, 25311, 25301, 25309",
     "Charleston Building Commission",
     "Appalachian Power (AEP)", "46482", "86", "25",
     "14 SEER2 (North Region)",
     "Appalachian Power TakeCharge WV: rebates on ENERGY STAR appliances, heat pump conversion program",
     "WV Division of Labor / Contractor Licensing Board \u2014 HVAC Technician Certification required (2,000 hours experience + exam); Contractor License for projects $5,000+",
     "Zone 4A (Mixed-Humid)", "Yes", "Yes"],

    ["Morgantown", "WV", "$3,500\u2013$7,500", "$3,500\u2013$7,500",
     "Suncrest, South Park, Woodburn, Evansdale, Sabraton",
     "26505, 26501, 26508, 26504, 26502",
     "City of Morgantown Code Enforcement/Building Inspection Division",
     "Mon Power (FirstEnergy)", "30490", "83", "22",
     "14 SEER2 (North Region)",
     "Mon Power/FirstEnergy: residential energy efficiency rebate programs",
     "WV Division of Labor / Contractor Licensing Board \u2014 same statewide requirement",
     "Zone 5A (Cool-Humid)", "Yes", "Yes"],

    # Montana
    ["Billings", "MT", "$4,000\u2013$8,000", "$3,500\u2013$7,500",
     "West End, The Heights, Rimrock, Central Park, South Side",
     "59102, 59101, 59106, 59105, 59103",
     "City of Billings Codes & Information Division",
     "NorthWestern Energy", "118321", "88", "18",
     "14 SEER2 (North Region)",
     "NorthWestern Energy: rebates for heat pumps, insulation, smart thermostats",
     "No state-level HVAC contractor license; must register with MT Dept. of Labor and Industry; EPA 608 required for refrigerant work",
     "Zone 6B (Cold-Dry)", "Yes", "Yes"],

    ["Missoula", "MT", "$4,000\u2013$8,000", "$3,500\u2013$7,500",
     "University District, South Hills, Rattlesnake, Downtown Missoula, Farviews",
     "59801, 59802, 59803, 59804, 59808",
     "City of Missoula Building Division, Permits and Inspections",
     "NorthWestern Energy", "78204", "87", "16",
     "14 SEER2 (North Region)",
     "NorthWestern Energy: heat pump, insulation, and smart thermostat rebates",
     "No state-level HVAC contractor license; register with MT Dept. of Labor and Industry",
     "Zone 6B (Cold-Dry)", "Yes", "Yes"],

    # North Dakota
    ["Fargo", "ND", "$3,500\u2013$7,500", "$3,500\u2013$7,500",
     "Downtown Fargo, Horace Mann, Longfellow, Hawthorne, Island Park",
     "58103, 58104, 58102, 58078, 58047",
     "City of Fargo Inspections Department",
     "Xcel Energy", "136285", "83", "-2",
     "14 SEER2 (North Region)",
     "Xcel Energy: rebates on qualifying furnaces, ACs, heat pumps, smart thermostats; Cass County Electric: up to $1,500 for qualifying off-peak HVAC",
     "No state-level HVAC license; municipal licensing required in Fargo; state contractor license from ND Secretary of State for projects $4,000+",
     "Zone 7 (Very Cold)", "Yes", "Yes"],

    ["Bismarck", "ND", "$3,500\u2013$7,500", "$3,500\u2013$7,500",
     "Cathedral District, North Hills, Cottonwood, South Bismarck, Wachter",
     "58501, 58503, 58504, 58505, 58502",
     "City of Bismarck Building Inspections Division",
     "Montana-Dakota Utilities (MDU)", "77772", "85", "-1",
     "14 SEER2 (North Region)",
     "Montana-Dakota Utilities: Conservation Improvement Program rebates for efficient thermostats and furnaces",
     "No state-level HVAC license; municipal licensing required; state contractor license for projects $4,000+",
     "Zone 6A (Cold-Humid)", "Yes", "Yes"],

    # South Dakota
    ["Sioux Falls", "SD", "$3,500\u2013$7,500", "$3,500\u2013$7,500",
     "McKennan Park, Downtown, Prairie Green, East Sioux Falls, Southwest Sioux Falls",
     "57106, 57103, 57104, 57105, 57108",
     "City of Sioux Falls Building Services",
     "Xcel Energy (electric), MidAmerican Energy (gas)", "209289", "85", "10",
     "14 SEER2 (North Region)",
     "Xcel Energy: rebates on qualifying furnaces, ACs, heat pumps; MidAmerican Energy: rebates for energy-efficient HVAC upgrades",
     "No state-level HVAC license; municipal licensing required in Sioux Falls (must pass mechanical exam)",
     "Zone 6A (Cold-Humid)", "Yes", "Yes"],

    ["Rapid City", "SD", "$3,500\u2013$7,500", "$3,500\u2013$7,500",
     "Southwest Rapid City, Chapel Valley, Sheridan Lake Road, Elks Country Estates, West Boulevard",
     "57701, 57702, 57703, 57706, 57709",
     "City of Rapid City Building Services Division",
     "Black Hills Energy", "79894", "86", "11",
     "14 SEER2 (North Region)",
     "Black Hills Energy: residential energy efficiency rebates for HVAC equipment",
     "No state-level HVAC license; municipal licensing required in Rapid City (2 years journeyman experience)",
     "Zone 5B (Cool-Dry)", "Yes", "Yes"],

    # Delaware
    ["Wilmington", "DE", "$3,500\u2013$7,500", "$3,500\u2013$7,000",
     "Trolley Square, The Highlands, Wawaset Park, Brandywine Hills, Forty Acres",
     "19802, 19806, 19805, 19801, 19803",
     "City of Wilmington Dept. of Licenses & Inspections",
     "Delmarva Power (Exelon)", "73176", "85", "25",
     "14 SEER2 (North Region)",
     "Delmarva Power: rebates on energy-efficient HVAC systems, Home Performance with ENERGY STAR up to $10,000, up to $15,000 for electrification upgrades",
     "Delaware Board of Plumbing, Heating, Ventilation, Air Conditioning and Refrigeration Examiners \u2014 Master HVACR license required",
     "Zone 4A (Mixed-Humid)", "Yes", "Yes"],

    # Vermont
    ["Burlington", "VT", "$4,000\u2013$8,000", "$4,000\u2013$8,500",
     "Hill Section, New North End, Old North End, South End, Five Sisters",
     "05401, 05408, 05405, 05403, 05404",
     "City of Burlington Permitting and Inspections Department",
     "Burlington Electric Department", "44743", "81", "12",
     "14 SEER2 (North Region)",
     "Efficiency Vermont: rebates on heat pumps, insulation, weatherization; Burlington Electric Dept. offers additional local incentives",
     "No state-level HVAC contractor license in Vermont; local requirements may apply",
     "Zone 6A (Cold-Humid)", "Yes", "Yes"],

    # Wyoming
    ["Cheyenne", "WY", "$3,500\u2013$7,500", "$3,500\u2013$7,500",
     "Saddle Ridge, Fox Farm-College, Thomas Heights, Whitney Ranch, Sun Valley",
     "82001, 82007, 82009, 82003, 82005",
     "City of Cheyenne Building Department",
     "Cheyenne Light Fuel and Power (Black Hills Energy)", "65132", "84", "15",
     "14 SEER2 (North Region)",
     "Black Hills Energy: incentives for energy-efficient HVAC equipment for gas and electric customers",
     "No state-level HVAC license in Wyoming; check local requirements",
     "Zone 6B (Cold-Dry)", "Yes", "Yes"],

    # New Hampshire
    ["Manchester", "NH", "$4,000\u2013$8,000", "$3,500\u2013$7,500",
     "North End, Pinardville, South End, Downtown/Millyard, Amoskeag",
     "03103, 03102, 03104, 03101, 03109",
     "City of Manchester Planning & Community Development Department",
     "Eversource Energy", "116386", "82", "18",
     "14 SEER2 (North Region)",
     "NHSaves (funded by Eversource, Liberty, Unitil, NH Electric Co-op): mail-in rebates on heat pumps and central AC, $250/ton heat pump rebate, Wi-Fi thermostat rebate",
     "No state-level HVAC license in NH (gas fitters excepted); local requirements may apply",
     "Zone 5A (Cool-Humid)", "Yes", "Yes"],
]

added = 0
for city in new_cities:
    ws.append(city)
    added += 1

wb.save('cities_updated.xlsx')
print(f"Added {added} cities to cities_updated.xlsx")

# Verify
total = sum(1 for _ in ws.iter_rows(min_row=2))
published = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if str(row[15] or '').lower() == 'yes')
print(f"Total cities: {total}, Published: {published}")
