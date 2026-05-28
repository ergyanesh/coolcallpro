#!/usr/bin/env python3
"""
City Page Generator v3 — 20 March 2026
Reads from cities_updated.xlsx + project_city_build_list.md
Generates HTML city pages matching the exact Houston template.

Usage:
    python generate_city_pages_v3.py                 # Generate all unpublished cities from build list
    python generate_city_pages_v3.py --batch 1       # Generate Batch 1 only (15 high-priority)
    python generate_city_pages_v3.py --city "Austin"  # Generate a single city
    python generate_city_pages_v3.py --dry-run        # Preview without writing files
"""

import json
import openpyxl
import os
import sys
import re
from datetime import date
from html import escape
from pathlib import Path

# ============================================================
# IMAGE ATTRIBUTIONS (loaded once)
# ============================================================
_ATTRIBUTIONS_PATH = Path(__file__).parent / "images" / "_attributions.json"
try:
    _IMAGE_ATTRIBUTIONS = json.loads(_ATTRIBUTIONS_PATH.read_text(encoding="utf-8"))
except (FileNotFoundError, json.JSONDecodeError):
    _IMAGE_ATTRIBUTIONS = {}


def get_location_image(kind, slug):
    """Return attribution dict for a location image, or None if missing."""
    return _IMAGE_ATTRIBUTIONS.get(f"{kind}/{slug}")


def location_figure_html(kind, slug, label, in_subdir=True):
    """Render an attributed <figure> block for a city or state image.

    Returns empty string if no image is available — caller need not check.
    `in_subdir` controls relative path: True for /locations/*.html (uses ../images/),
    False for root-level pages (uses images/).
    """
    rec = get_location_image(kind, slug)
    if not rec:
        return ""
    file_rel = rec["file"]  # e.g. "images/cities/austin-tx.webp"
    src = f"../{file_rel}" if in_subdir else file_rel
    width = rec.get("width") or 1600
    height = rec.get("height") or 900
    artist = escape(rec.get("artist") or "Unknown photographer")
    license_short = escape(rec.get("license") or "")
    license_url = escape(rec.get("license_url") or "")
    commons_url = escape(rec.get("commons_url") or rec.get("image_source_url") or "")
    description = (rec.get("description") or "").strip()
    # Alt text: use Wikipedia description if meaningful, else fall back to label
    if description and len(description) > 12:
        # Truncate at first sentence boundary or 140 chars
        alt_text = description.split(". ")[0].strip()[:140]
        if not alt_text.endswith(('.', '!', '?')):
            alt_text = alt_text + f" — {label}"
    else:
        alt_text = f"{label} — local landmark photo"
    alt_text = escape(alt_text)
    credits_href = "../image-credits" if in_subdir else "image-credits.html"
    license_html = (
        f'<a href="{license_url}" target="_blank" rel="nofollow noopener">{license_short}</a>'
        if license_url and license_short else license_short
    )
    return f'''
        <figure class="location-figure" style="margin: 0 auto 40px; max-width: 1100px;">
          <img src="{src}" alt="{alt_text}" width="{width}" height="{height}" loading="lazy" decoding="async" style="width: 100%; height: auto; border-radius: var(--radius); display: block;" />
          <figcaption style="display: inline-block; margin-top: 12px; padding: 8px 14px; background: var(--gray-50); border: 1px solid var(--gray-100); border-radius: 999px; font-size: 0.78rem; color: var(--gray-600); line-height: 1.5;">
            <span aria-hidden="true" style="margin-right: 6px;">&#128247;</span>{artist} &middot; {license_html} via <a href="{commons_url}" target="_blank" rel="nofollow noopener" style="color: var(--gray-700); text-decoration: underline;">Wikimedia Commons</a> &middot; <a href="{credits_href}" style="color: var(--gray-700); text-decoration: underline;">credits</a>
          </figcaption>
        </figure>
'''

# ============================================================
# CONFIGURATION
# ============================================================

XLSX_FILE = "cities_updated.xlsx"
STATES_XLSX = "states.xlsx"
OUTPUT_DIR = "locations"

# Already published — do NOT regenerate
PUBLISHED_CITIES = {
    "Birmingham", "Phoenix", "Denver", "Tampa", "Atlanta", "Chicago",
    "New Orleans", "Boston", "Detroit", "Minneapolis", "Kansas City",
    "St. Louis", "Charlotte", "Las Vegas", "Oklahoma City", "Philadelphia",
    "Dallas", "Houston", "San Antonio", "Milwaukee",
    # Batch 1 (published 20 March 2026)
    "Los Angeles", "New York City", "San Jose", "Austin", "Jacksonville",
    "Fort Worth", "San Diego", "San Francisco", "Columbus", "Indianapolis",
    "Nashville", "Washington", "Seattle", "El Paso", "Louisville",
    # Batch 2 (published 29 March 2026)
    "Portland", "Memphis", "Baltimore", "Albuquerque", "Tucson",
    "Fresno", "Sacramento", "Mesa", "Omaha", "Miami",
    "Colorado Springs", "Raleigh", "Virginia Beach", "Oakland",
    "Tulsa", "Cleveland", "Wichita",
    # Batch 3 (deployed 8 April 2026)
    "Allentown", "Anchorage", "Augusta", "Baton Rouge", "Boise",
    "Bridgeport", "Brownsville", "Buffalo", "Chattanooga", "Dayton",
    "Des Moines", "Eugene", "Fayetteville", "Fort Wayne", "Glendale",
    "Grand Rapids", "Hartford", "Huntsville", "Irvine", "Knoxville",
    "Laredo", "Lexington", "Lincoln", "Little Rock", "Lubbock",
    "Madison", "Mobile", "Modesto", "Newark", "Peoria",
    "Pittsburgh", "Providence", "Reno", "Richmond", "Salt Lake City",
    "Shreveport", "Spokane", "Springfield", "Syracuse", "Tallahassee",
    "Toledo", "Winston-Salem", "Worcester"
}

# Removed — too small, no page
REMOVED_CITIES = {"Fairbanks", "Binghamton", "Johnstown", "Danville"}

# Metro absorption map: absorbed_city -> anchor_city
ABSORBED = {
    "Gilbert": "Phoenix", "Chula Vista": "San Diego", "Fremont": "San Jose",
    "Santa Cruz": "San Jose", "Santa Rosa": "San Francisco",
    "Aurora": "Denver", "Boulder": "Denver", "Stamford": "Bridgeport",
    "Hialeah": "Miami", "St. Petersburg": "Tampa", "Davenport": "Des Moines",
    "Rockford": "Chicago", "Pittsfield": "Springfield", "Flint": "Detroit",
    "Jersey City": "Newark", "Trenton": "Newark", "Henderson": "Las Vegas",
    "Utica": "Syracuse", "Akron": "Cleveland", "Canton": "Cleveland",
    "Youngstown": "Cleveland", "Scranton": "Allentown",
    "Arlington": "Dallas", "Plano": "Dallas", "Irving": "Dallas",
    "Grand Prairie": "Dallas", "Garland": "Dallas",
    "Chesapeake": "Virginia Beach", "Tacoma": "Seattle",
    "Rochester": "Buffalo", "Elmira": "Binghamton"
}

# Reverse: anchor -> list of absorbed cities
ANCHOR_ABSORBS = {}
for absorbed, anchor in ABSORBED.items():
    ANCHOR_ABSORBS.setdefault(anchor, []).append(absorbed)

# Batch definitions (city names)
BATCH_1 = [
    "Los Angeles", "New York City", "San Jose", "Austin", "Jacksonville",
    "Fort Worth", "San Diego", "San Francisco", "Columbus", "Indianapolis",
    "Nashville", "Washington", "Seattle", "El Paso", "Louisville"
]

BATCH_2 = [
    "Portland", "Memphis", "Baltimore", "Albuquerque", "Tucson",
    "Fresno", "Sacramento", "Mesa", "Omaha", "Miami",
    "Colorado Springs", "Raleigh", "Virginia Beach", "Oakland",
    "Tulsa", "Cleveland", "Wichita"
]

# ============================================================
# CLIMATE MAPPING
# ============================================================

CLIMATE_MAP = {
    "Zone 1A (Very Hot-Humid)": "tropical",
    "Zone 2A (Hot-Humid)": "hot-humid",
    "Zone 2B (Hot-Dry)": "hot-dry",
    "Zone 3A (Warm-Humid)": "mixed-humid",
    "Zone 3B (Warm-Dry)": "hot-dry",
    "Zone 3C (Warm-Marine)": "coastal",
    "Zone 4A (Mixed-Humid)": None,  # needs manual decision
    "Zone 4B (Mixed-Dry)": "mountain",
    "Zone 4C (Mixed-Marine)": "coastal",
    "Zone 5A (Cool-Humid)": "cold",
    "Zone 5B (Cold-Dry)": "mountain",
    "Zone 5B (Cool-Dry)": "mountain",
    "Zone 6A (Cold-Humid)": "cold",
    "Zone 6B (Cold-Dry)": "mountain",
    "Zone 7 (Very Cold)": "subarctic",
    "Zone 8 (Subarctic)": "subarctic",
}

# Zone 4A manual decisions
ZONE_4A_DECISIONS = {
    "Kansas City": "cold", "Philadelphia": "cold", "St. Louis": "mixed-humid",
    "Charlotte": "mixed-humid", "Baltimore": "cold", "Washington": "mixed-humid",
    "Nashville": "mixed-humid", "Louisville": "mixed-humid", "Lexington": "mixed-humid",
    "Raleigh": "mixed-humid", "Winston-Salem": "mixed-humid", "Wichita": "cold",
    "Chattanooga": "mixed-humid", "Knoxville": "mixed-humid", "Newark": "cold",
    "New York City": "cold", "Richmond": "mixed-humid", "Virginia Beach": "mixed-humid",
    "Dayton": "cold", "Fort Wayne": "cold", "Indianapolis": "cold",
    "Omaha": "cold", "Lincoln": "cold", "Des Moines": "cold",
    "Toledo": "cold", "Columbus": "cold", "Cleveland": "cold",
    "Akron": "cold", "Buffalo": "cold", "Syracuse": "cold",
    "Rochester": "cold",
    "Wilmington": "cold",
    "Charleston": "mixed-humid",
}

HERO_IMAGES = {
    "hot-humid": "hot-humid-climate-hvac-houston-hero.webp",
    "hot-dry": "desert-climate-hvac-phoenix-hero.webp",
    "mixed-humid": "mixed-humid-climate-hvac-atlanta-hero.webp",
    "cold": "cold-climate-furnace-minneapolis-hero.webp",
    "mountain": "mountain-climate-hvac-denver-hero.webp",
    "tropical": "hot-humid-climate-hvac-houston-hero.webp",
    "coastal": "mixed-humid-climate-hvac-atlanta-hero.webp",
    "subarctic": "cold-climate-furnace-minneapolis-hero.webp",
}

# Climate profiles — drive all conditional text across the template
CLIMATE_PROFILES = {
    "tropical": {
        "title_service": "AC &amp; Cooling",
        "h1_service": "24/7 AC Repair &amp; Emergency Cooling Service",
        "hero_service": "Emergency AC repair, heat pump service, and system installation available 24/7.",
        "og_service": "Emergency AC and cooling service available 24/7.",
        "meta_desc_service": "AC repair",
        "permit_faq_q": "Do I need a permit to replace my AC",
        "cost_faq_q": "How much does AC replacement cost",
        "furnace_label": "mini-split/heat pump installations",
        "cost_order": "ac_first_tropical",
    },
    "hot-humid": {
        "title_service": "AC &amp; HVAC Repair",
        "h1_service": "24/7 AC Repair &amp; Emergency HVAC Service",
        "hero_service": "Emergency AC repair, furnace service, and system installation available 24/7.",
        "og_service": "Emergency AC repair and HVAC service available 24/7.",
        "meta_desc_service": "AC or HVAC repair",
        "permit_faq_q": "Do I need a permit to replace my AC",
        "cost_faq_q": "How much does HVAC replacement cost",
        "furnace_label": "furnace installations",
        "cost_order": "ac_first",
    },
    "hot-dry": {
        "title_service": "AC &amp; HVAC Repair",
        "h1_service": "24/7 AC Repair &amp; Emergency HVAC Service",
        "hero_service": "Emergency AC repair, furnace service, and system installation available 24/7.",
        "og_service": "Emergency AC repair and HVAC service available 24/7.",
        "meta_desc_service": "AC repair",
        "permit_faq_q": "Do I need a permit to replace my AC",
        "cost_faq_q": "How much does HVAC replacement cost",
        "furnace_label": "furnace installations",
        "cost_order": "ac_first",
    },
    "mixed-humid": {
        "title_service": "AC &amp; HVAC Repair",
        "h1_service": "24/7 AC Repair &amp; Emergency HVAC Service",
        "hero_service": "Emergency AC repair, furnace service, and system installation available 24/7.",
        "og_service": "Emergency AC and furnace service available 24/7.",
        "meta_desc_service": "AC or furnace repair",
        "permit_faq_q": "Do I need a permit to replace my AC or furnace",
        "cost_faq_q": "How much does HVAC replacement cost",
        "furnace_label": "furnace installations",
        "cost_order": "ac_first",
    },
    "cold": {
        "title_service": "Furnace &amp; HVAC Repair",
        "h1_service": "24/7 Furnace Repair &amp; Emergency HVAC Service",
        "hero_service": "Emergency furnace repair, AC service, and system installation available 24/7.",
        "og_service": "Emergency furnace repair and HVAC service available 24/7.",
        "meta_desc_service": "furnace or HVAC repair",
        "permit_faq_q": "Do I need a permit to replace my furnace",
        "cost_faq_q": "How much does HVAC replacement cost",
        "furnace_label": "furnace installations",
        "cost_order": "furnace_first",
    },
    "mountain": {
        "title_service": "HVAC Repair &amp; Service",
        "h1_service": "24/7 HVAC Repair &amp; Emergency Service",
        "hero_service": "Emergency furnace repair, AC service, and high-altitude HVAC installation available 24/7.",
        "og_service": "Emergency HVAC repair and high-altitude service available 24/7.",
        "meta_desc_service": "HVAC repair",
        "permit_faq_q": "Do I need a permit to replace my HVAC system",
        "cost_faq_q": "How much does HVAC replacement cost",
        "furnace_label": "furnace installations",
        "cost_order": "ac_first",
    },
    "coastal": {
        "title_service": "Heat Pump &amp; HVAC",
        "h1_service": "24/7 Heat Pump &amp; HVAC Service",
        "hero_service": "Emergency heat pump repair, AC service, and system installation available 24/7.",
        "og_service": "Emergency HVAC and heat pump service available 24/7.",
        "meta_desc_service": "heat pump or HVAC repair",
        "permit_faq_q": "Do I need a permit to replace my HVAC system",
        "cost_faq_q": "How much does HVAC replacement cost",
        "furnace_label": "furnace installations",
        "cost_order": "ac_first",
    },
    "subarctic": {
        "title_service": "Furnace &amp; Heating",
        "h1_service": "24/7 Furnace Repair &amp; Emergency Heating Service",
        "hero_service": "Emergency furnace repair, heating service, and system installation available 24/7.",
        "og_service": "Emergency furnace repair and heating service available 24/7.",
        "meta_desc_service": "furnace repair",
        "permit_faq_q": "Do I need a permit to replace my furnace",
        "cost_faq_q": "How much does HVAC replacement cost",
        "furnace_label": "furnace installations",
        "cost_order": "furnace_first",
    },
}

CLIMATE_ARTICLE_LINKS = {
    "hot-humid": ('../article-ac-summer', 'preparing your AC for summer'),
    "hot-dry": ('../article-spring-ac', 'preparing your AC before the heat arrives'),
    "mixed-humid": ('../article-heat-pump', 'heat pump guide'),
    "cold": ('../article-furnace', 'what to do when your furnace fails during a cold snap'),
    "mountain": ('../article-winter-storm', 'preparing for winter storms'),
    "tropical": ('../article-ac-summer', 'keeping your AC running through peak summer heat'),
    "coastal": ('../article-heat-pump', 'heat pump guide'),
    "subarctic": ('../article-furnace', 'what to do when your furnace fails during a cold snap'),
}

# ============================================================
# STATE LOOKUPS
# ============================================================

STATE_FULL = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DC": "District of Columbia",
    "DE": "Delaware", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii",
    "IA": "Iowa", "ID": "Idaho", "IL": "Illinois", "IN": "Indiana",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "MA": "Massachusetts",
    "MD": "Maryland", "ME": "Maine", "MI": "Michigan", "MN": "Minnesota",
    "MO": "Missouri", "MS": "Mississippi", "MT": "Montana", "NC": "North Carolina",
    "ND": "North Dakota", "NE": "Nebraska", "NH": "New Hampshire", "NJ": "New Jersey",
    "NM": "New Mexico", "NV": "Nevada", "NY": "New York", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island",
    "SC": "South Carolina", "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas",
    "UT": "Utah", "VA": "Virginia", "VT": "Vermont", "WA": "Washington",
    "WI": "Wisconsin", "WV": "West Virginia", "WY": "Wyoming",
}

def state_slug(abbr):
    """Convert state abbreviation to URL slug: TX -> texas, NC -> north-carolina"""
    full = STATE_FULL.get(abbr, abbr)
    return full.lower().replace(" ", "-")

def city_slug(city, state_abbr):
    """Convert city name + state to URL slug: New York City, NY -> new-york-city-ny"""
    slug = city.lower().replace(" ", "-").replace(".", "")
    return f"{slug}-{state_abbr.lower()}"

def get_climate_type(city_name, climate_zone):
    """Determine climate_type from zone, with Zone 4A manual lookup."""
    ct = CLIMATE_MAP.get(climate_zone)
    if ct is None and "4A" in str(climate_zone):
        ct = ZONE_4A_DECISIONS.get(city_name, "mixed-humid")
    if ct is None:
        ct = "mixed-humid"  # fallback
    return ct

# ============================================================
# NEARBY CITY LOGIC
# ============================================================

# State-based nearby preferences (will pick from same state first, then these)
CROSS_STATE_NEARBY = {
    "AK": ["Seattle"],
    "AR": ["Memphis", "Dallas"],
    "DC": ["Baltimore", "Richmond", "Philadelphia"],
    "ID": ["Salt Lake City", "Spokane"],
    "KS": ["Oklahoma City", "Kansas City"],
    "MD": ["Washington", "Philadelphia"],
    "NM": ["El Paso", "Phoenix"],
    "RI": ["Boston", "Worcester"],
    "UT": ["Denver", "Boise"],
}

def get_nearby_cities(city_data, all_cities_by_state, all_cities_lookup):
    """Return list of (city_name, state_abbr, slug) for nearby links."""
    city = city_data['city']
    state = city_data['state']
    nearby = []

    # Same-state cities first
    same_state = [c for c in all_cities_by_state.get(state, [])
                  if c['city'] != city and c['city'] not in REMOVED_CITIES
                  and c['city'] not in ABSORBED]
    same_state.sort(key=lambda c: c['pop'], reverse=True)
    for c in same_state[:2]:
        nearby.append((c['city'], c['state'], city_slug(c['city'], c['state'])))

    # Cross-state if needed
    if len(nearby) < 3 and state in CROSS_STATE_NEARBY:
        for cross_city in CROSS_STATE_NEARBY[state]:
            if cross_city in all_cities_lookup and len(nearby) < 3:
                cs = all_cities_lookup[cross_city]
                nearby.append((cs['city'], cs['state'], city_slug(cs['city'], cs['state'])))

    # If still < 2, just use some big cities
    if len(nearby) < 2:
        fallback = ["Chicago", "Houston", "Phoenix", "Philadelphia", "Dallas"]
        for fb in fallback:
            if fb != city and fb in all_cities_lookup and len(nearby) < 3:
                cs = all_cities_lookup[fb]
                nearby.append((cs['city'], cs['state'], city_slug(cs['city'], cs['state'])))

    return nearby[:3]

# ============================================================
# CONTENT GENERATORS
# ============================================================

def climate_h2_and_paragraph(c, climate_type):
    """Generate the climate-specific H2 and paragraph."""
    city = c['city']
    summer = c['summer_high']
    winter = c['winter_low']
    seer = c['seer']
    utility = c['utility']
    climate_zone = c['climate']
    article_href, article_text = CLIMATE_ARTICLE_LINKS[climate_type]

    templates = {
        "hot-humid": (
            f"{escape(city)}'s Hot-Humid Climate &amp; Your HVAC System",
            f'With summer highs averaging {summer}&deg;F and humidity levels that strain even the best systems, {escape(city)} homeowners know their AC isn\'t optional. Winter lows rarely dip below {winter}&deg;F, so heating costs stay manageable. As a {escape(climate_zone)} climate zone, federal regulations require all new AC units meet a minimum {escape(seer)} rating, which translates to lower energy bills with {escape(utility)}.'
        ),
        "hot-dry": (
            f"Desert Heat and Your HVAC: What {escape(city)} Homeowners Need to Know",
            f'When {escape(city)} summers push past {summer}&deg;F, your air conditioner works harder than almost anywhere else in the country. The desert climate means low humidity but extreme heat, making proper system sizing critical&mdash;an undersized unit will run nonstop and fail prematurely. Winter lows around {winter}&deg;F mean you still need reliable heating for a few months each year. Located in {escape(climate_zone)}, all new installations must meet {escape(seer)} efficiency standards.'
        ),
        "mixed-humid": (
            f"Year-Round Comfort in {escape(city)}: Dual-Season HVAC Demands",
            f'{escape(city)}\'s {escape(climate_zone)} climate means your HVAC system works year-round&mdash;battling {summer}&deg;F summer heat and winter lows near {winter}&deg;F. This dual demand makes efficiency essential, and many {escape(city)} homeowners are switching to heat pumps that handle both heating and cooling in a single system. All new installations must meet {escape(seer)} efficiency standards, and {escape(utility)} serves as the primary utility for most {escape(city)} ZIP codes.'
        ),
        "cold": (
            f"{escape(city)} Winters Mean Business: Is Your Furnace Ready?",
            f'{escape(city)} winters are serious, with temperatures dropping to {winter}&deg;F during the coldest stretches. Your furnace is not a luxury&mdash;it is a lifeline. While summer highs reach {summer}&deg;F and AC still matters, the real investment for {escape(city)} homeowners is reliable, efficient heating. Located in {escape(climate_zone)}, heating efficiency standards are critical here. {escape(utility)} serves most {escape(city)} residents, and upgrading an aging furnace before it fails mid-January is always cheaper than an emergency replacement.'
        ),
        "mountain": (
            f"Mile-High HVAC: Why {escape(city)} Needs Specialized Equipment",
            f'{escape(city)}\'s high altitude creates HVAC challenges you will not find at sea level. Summer highs reach {summer}&deg;F, but winter lows plunge to {winter}&deg;F, and the thin air reduces equipment efficiency by 4&ndash;5% per 1,000 feet of elevation. As a {escape(climate_zone)} zone, you need contractors who understand high-altitude combustion adjustments and properly sized blower motors. {escape(utility)} serves most {escape(city)} residents.'
        ),
        "tropical": (
            f"Year-Round AC in {escape(city)}: Why Cooling Never Stops",
            f'In {escape(city)}, air conditioning is not seasonal&mdash;it is essential 365 days a year. With summer highs averaging {summer}&deg;F and winter lows rarely dropping below {winter}&deg;F, your AC system runs nearly nonstop. The {escape(climate_zone)} classification means extreme humidity that accelerates corrosion, mold growth in ductwork, and refrigerant pressure issues. All new installations must meet {escape(seer)} efficiency standards, and higher-SEER units make a real difference on your {escape(utility)} bill when the system runs 10+ months per year.'
        ),
        "coastal": (
            f"Mild Climate, Hidden HVAC Challenges in {escape(city)}",
            f'{escape(city)}\'s {escape(climate_zone)} marine climate means moderate temperatures year-round&mdash;summer highs around {summer}&deg;F and winter lows near {winter}&deg;F. While the mild weather means lower energy bills than most U.S. cities, the persistent moisture and coastal air create unique HVAC demands: mold prevention, corrosion-resistant equipment, and proper ventilation matter more here than raw heating or cooling power. All new installations must meet {escape(seer)} efficiency standards. {escape(utility)} serves most {escape(city)} residents.'
        ),
        "subarctic": (
            f"Extreme Cold HVAC: Surviving {escape(city)} Winters Below Zero",
            f'{escape(city)} faces some of the most extreme heating demands in the United States. Winter lows plunge to {winter}&deg;F, and temperatures can stay below freezing for months at a time. Your furnace is your most critical home system&mdash;a failure in January is a genuine emergency that can cause frozen pipes and structural damage within hours. Summer highs reach {summer}&deg;F, making AC useful but secondary to heating. Located in {escape(climate_zone)}, you need contractors experienced with cold-climate equipment. {escape(utility)} serves most {escape(city)} residents.'
        ),
    }

    h2, para = templates[climate_type]
    para += f' Learn more about <a href="{article_href}" style="color: var(--orange-dark); font-weight: 600;">{article_text}</a>.'
    return h2, para


def services_list(city, climate_type):
    """Return list of 5 service strings for the climate type."""
    c = escape(city)
    lists = {
        "hot-humid": [
            f"Emergency AC Repair in {c}",
            "Humidity Control &amp; Dehumidification",
            "Central AC Installation &amp; Replacement",
            "HVAC System Maintenance &amp; Tune-Ups",
            "Ductwork Inspection, Cleaning &amp; Mold Prevention",
        ],
        "hot-dry": [
            f"Emergency AC Repair in {c}",
            "Desert-Climate AC Sizing &amp; Installation",
            "Evaporative-to-Refrigerated Cooling Conversion",
            f"Furnace Repair &amp; Winter Heating Service in {c}",
            "Ductwork Inspection, Cleaning &amp; Sealing",
        ],
        "mixed-humid": [
            f"Emergency AC Repair in {c}",
            f"Furnace Repair &amp; Heating Service in {c}",
            "Heat Pump Installation &amp; Dual-Fuel Systems",
            "Central Air Conditioning Installation &amp; Replacement",
            "HVAC System Maintenance &amp; Seasonal Tune-Ups",
        ],
        "cold": [
            f"Emergency Furnace Repair in {c}",
            f"High-Efficiency Furnace Installation in {c}",
            "Central Air Conditioning Repair &amp; Replacement",
            "Boiler Service &amp; Radiant Heating",
            "Ductwork Inspection, Cleaning &amp; Insulation",
        ],
        "mountain": [
            f"High-Altitude Furnace Installation in {c}",
            f"Emergency HVAC Repair in {c}",
            "Central Air Conditioning Installation &amp; Replacement",
            "Heat Pump Systems for Mountain Climates",
            "Ductwork Inspection &amp; High-Altitude Combustion Testing",
        ],
        "tropical": [
            f"24/7 Emergency AC Repair in {c}",
            "Corrosion-Resistant AC Installation for Coastal Climates",
            f"Duct Cleaning &amp; Mold Remediation in {c}",
            "High-Efficiency Heat Pump Installation",
            "Annual HVAC Maintenance &amp; Hurricane-Season Prep",
        ],
        "coastal": [
            f"Emergency AC &amp; Heating Repair in {c}",
            f"Heat Pump Installation in {c}",
            "Corrosion-Resistant HVAC Systems for Marine Climates",
            "Ductwork Inspection, Mold Prevention &amp; Sealing",
            "HVAC System Maintenance &amp; Seasonal Tune-Ups",
        ],
        "subarctic": [
            f"Emergency Furnace Repair in {c}",
            f"Cold-Climate Furnace Installation in {c}",
            "Boiler Service &amp; Radiant Heating Systems",
            "Backup Heating &amp; Generator Integration",
            "Ductwork Insulation &amp; Frozen Pipe Prevention",
        ],
    }
    return lists.get(climate_type, lists["mixed-humid"])


def license_faq_sentence(c, html=True):
    """Return a clean single-sentence answer for the "What HVAC license is
    required in {state}?" FAQ. Handles "no license" states gracefully —
    never dumps raw xlsx into a sentence template that expects a license name.

    Used in BOTH the visible FAQ accordion (html=True) and the FAQPage
    JSON-LD schema (html=False, no tags).
    """
    state_full = STATE_FULL.get(c['state'], c['state'])
    license_req_raw = (c['license'] or '').strip()
    is_no_license = license_req_raw.lower().startswith('no ')

    if is_no_license:
        alt = ""
        if ';' in license_req_raw:
            part = license_req_raw.split(';', 1)[1].strip().rstrip('.')
            if part:
                if part[0].isupper() and not (len(part) > 1 and part[1].isupper()):
                    part = part[0].lower() + part[1:]
                alt = f" Instead, {part}."
        if html:
            return f"{escape(state_full)} does not require a statewide HVAC contractor license.{escape(alt)}"
        return f"{state_full} does not require a statewide HVAC contractor license.{alt}"
    else:
        if html:
            return f"In {escape(state_full)}, HVAC contractors should hold a <strong>{escape(license_req_raw)}</strong>."
        return f"In {state_full}, HVAC contractors should hold a {license_req_raw}."


def licensing_paragraph(c, state_info=None):
    """Generate the licensing/permits paragraph.

    Handles two cases:
    (1) State DOES issue an HVAC license — dumps the license name as-is.
    (2) State does NOT issue a statewide HVAC license — detects ANY variant
        beginning with "No " (e.g. "No state license; $25,000 surety bond
        required") and parses the clause after `;` as an alternative
        requirement ("Instead, contractors must carry a $25,000 surety bond.").

    state_info is a dict from states.xlsx with 'Licensing Body/Agency' and
    'License Lookup URL'. When present, appends a verified lookup link.
    """
    city = escape(c['city'])
    state_full = escape(STATE_FULL.get(c['state'], c['state']))
    license_req_raw = (c['license'] or '').strip()
    license_req = escape(license_req_raw)
    permit = escape(c['permit'])
    state_s = state_slug(c['state'])

    # Verified license lookup link (from states.xlsx)
    lookup_html = ""
    if state_info:
        body = (state_info.get('Licensing Body/Agency') or '').strip()
        url = (state_info.get('License Lookup URL') or '').strip()
        if body and url and url.upper() != 'N/A':
            lookup_html = (f' Verify any contractor\'s license at the '
                           f'<a href="{escape(url)}" target="_blank" rel="nofollow noopener" '
                           f'style="color: var(--orange-dark); font-weight: 600;">{escape(body)}</a>.')

    # Detect "no license" states — any string starting with "No " (case-insensitive).
    # Covers: "No state license; ...", "No statewide license; ...", "No mandatory
    # statewide license; ...", "No statewide HVAC contractor license", etc.
    is_no_license = license_req_raw.lower().startswith('no ')

    if is_no_license:
        # Parse the alternative requirement after the first ';' if present.
        # E.g. "No state license; $25,000 surety bond required" ->
        # alternative = "$25,000 surety bond required"
        alt_sentence = ""
        if ';' in license_req_raw:
            alt = license_req_raw.split(';', 1)[1].strip().rstrip('.')
            if alt:
                # Lowercase the leading character so it reads naturally after
                # "Instead,"; preserve it if it's a proper-noun/acronym like EPA.
                if alt[0].isupper() and not (len(alt) > 1 and alt[1].isupper()):
                    alt = alt[0].lower() + alt[1:]
                alt_sentence = f' Instead, {escape(alt)}.'

        return (
            f"HVAC Licensing &amp; Permits in {city}, {escape(c['state'])}",
            f'{state_full} does not require a statewide HVAC contractor license.{alt_sentence} {city} enforces local permit requirements through <strong>{permit}</strong>.{lookup_html} Always verify your contractor carries liability insurance and workers\' compensation coverage. Any major installation or system replacement should have a permit on file&mdash;skipping permits can void manufacturer warranties and create problems when selling your home. See our <a href="../costs" style="color: var(--orange-dark); font-weight: 600;">HVAC cost guide</a> for detailed pricing breakdowns. For full statewide regulations, see our <a href="{state_s}" style="color: var(--orange-dark); font-weight: 600;">{state_full} HVAC guide</a>.'
        )
    else:
        return (
            f"HVAC Licensing &amp; Permits in {city}, {escape(c['state'])}",
            f'Before hiring any HVAC contractor in {state_full}, verify they hold the proper <strong>{license_req}</strong>.{lookup_html} Licensed contractors carry insurance, pull permits correctly, and stand behind their work. For {city} residents, your local permit office is <strong>{permit}</strong>&mdash;any major installation should have a permit on file. See our <a href="../costs" style="color: var(--orange-dark); font-weight: 600;">HVAC cost guide</a> for detailed pricing breakdowns. For full statewide licensing details, see our <a href="{state_s}" style="color: var(--orange-dark); font-weight: 600;">{state_full} HVAC guide</a>.'
        )


def rebates_paragraph(c):
    """Generate the rebates/incentives paragraph with DSIRE / ENERGY STAR / IRS §25C citations."""
    city = escape(c['city'])
    state_abbr = escape(c['state'])
    state_full = escape(STATE_FULL.get(c['state'], c['state']))
    utility = escape(c['utility'])
    rebates = escape(c['rebates'])
    return (
        f"HVAC Rebates &amp; Incentives in {city}, {state_abbr}",
        f'{city} homeowners served by <strong>{utility}</strong> may qualify for savings through <strong>{rebates}</strong> when installing qualifying high-efficiency equipment. For all {state_full} incentives including state and federal programs, see the <a href="https://www.dsireusa.org/" target="_blank" rel="nofollow noopener" style="color: var(--orange-dark); font-weight: 600;">DSIRE database (N.C. State University / U.S. DOE)</a> or the <a href="https://www.energystar.gov/rebate-finder" target="_blank" rel="nofollow noopener" style="color: var(--orange-dark); font-weight: 600;">ENERGY STAR Rebate Finder (EPA)</a>. The federal Section 25C credit (<a href="https://www.irs.gov/credits-deductions/energy-efficient-home-improvement-credit" target="_blank" rel="nofollow noopener" style="color: var(--orange-dark); font-weight: 600;">Energy Efficient Home Improvement Credit</a>) was terminated for installations placed in service after Dec 31, 2025 by the One Big Beautiful Bill Act (P.L. 119-21); for 2026 installs focus on state HEAR rebates and utility programs. Learn about <a href="../article-hvac-financing" style="color: var(--orange-dark); font-weight: 600;">HVAC financing options</a>.'
    )


def page_meta_html(reviewed_iso, reviewed_readable):
    """Hero byline pill + last-reviewed line, styled to match article pages.

    Uses the existing `.article-meta-modern` (for the timestamp, white-on-dark)
    and `.author-byline` + `.author-avatar-sm` pill used on article templates.
    Rendered inside the dark hero section of each city/state page.
    """
    return (
        f'<div class="article-meta-modern"><div class="meta-line meta-reviewed">Last reviewed <time datetime="{reviewed_iso}">{reviewed_readable}</time></div></div>\n'
        f'        <div class="author-byline">\n'
        f'          <img src="../images/author-gyane.webp" alt="Gyanesh Gulshan" class="author-avatar-sm" loading="lazy" width="48" height="48" />\n'
        f'          <span>By <a href="../author-gyanesh">Gyanesh Gulshan</a> &mdash; Founder, Cool Call Pro</span>\n'
        f'        </div>'
    )


def neighborhoods_section_html(c):
    """Surface neighborhoods + ZIPs from xlsx as a visible list.

    Both neighborhoods and zips are comma-separated in xlsx. We pair them
    positionally; when counts differ, we zip to the shorter length and show
    any remaining ZIPs on a final line. Fully factual — no invented details.
    """
    city_e = escape(c['city'])
    st_e = escape(c['state'])
    raw_n = [n.strip() for n in str(c['neighborhoods']).split(',') if n.strip()]
    raw_z = [z.strip() for z in str(c['zips']).split(',') if z.strip()]

    if not raw_n and not raw_z:
        return ""

    items = []
    pairs = min(len(raw_n), len(raw_z))
    for i in range(pairs):
        items.append(f'<li><strong>{escape(raw_n[i])}</strong> &mdash; ZIP {escape(raw_z[i])}</li>')
    # Extra neighborhoods with no paired ZIP
    for n in raw_n[pairs:]:
        items.append(f'<li><strong>{escape(n)}</strong></li>')
    # Extra ZIPs with no paired neighborhood
    extra_z = raw_z[pairs:]
    if extra_z:
        items.append(f'<li>Additional ZIPs: {escape(", ".join(extra_z))}</li>')

    list_html = "\n            ".join(items)
    return f'''
        <!-- ZIP Codes & Neighborhoods (surfaced from xlsx data) -->
        <div class="city-services" id="neighborhoods" style="margin-bottom: 40px;">
          <h2>ZIP Codes &amp; Neighborhoods Served in {city_e}, {st_e}</h2>
          <p style="font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);">Our referral network connects homeowners across {city_e}&rsquo;s neighborhoods with independent HVAC professionals. Service verification available via <a href="tel:+18445821795" style="color: var(--orange-dark); font-weight: 600;">(844) 582-1795</a>.</p>
          <ul class="neighborhood-grid">
            {list_html}
          </ul>
        </div>
'''


def climate_by_numbers_html(c):
    """Phase 3 / uniqueness uplift: render NOAA Climate Normals as data-driven
    factual prose. Each paragraph picks a sentence variant based on data ranges
    so that two cities with similar climate types still produce distinct prose.

    Drives ~200 words of city-specific content per page when all NOAA fields
    are populated. Skips fields where NCEI returned a sentinel (None for
    tropical-no-HDD, arid-no-snow, etc.) and adapts prose around what's missing.

    Source: NOAA NCEI 1991-2020 Climate Normals (verified 2026-05-09).
    """
    city_e = escape(c['city'])
    st_e = escape(c['state'])
    hdd = c.get('noaa_hdd_annual')
    cdd = c.get('noaa_cdd_annual')
    days_90 = c.get('noaa_days_above_90')
    days_32 = c.get('noaa_days_below_32')
    snow = c.get('noaa_snowfall_inches')

    # If we have nothing, render nothing (Laredo and similar gaps)
    if not any(v is not None for v in (hdd, cdd, days_90, days_32, snow)):
        return ""

    # Sentence 1 — opening, by HDD/CDD ratio
    if hdd is not None and cdd is not None and cdd > 0:
        ratio = hdd / cdd if cdd else 999
        if ratio > 5:
            opener = (f"{city_e}'s climate is heating-dominated: NOAA's 1991-2020 Climate Normals "
                      f"record <strong>{hdd:,} annual heating degree days</strong> against just "
                      f"{cdd:,} cooling degree days &mdash; a {ratio:.1f}-to-1 heating-to-cooling load.")
        elif ratio > 2:
            opener = (f"NOAA's 1991-2020 Climate Normals record <strong>{hdd:,} heating degree days</strong> "
                      f"and {cdd:,} cooling degree days for {city_e} &mdash; a heating-leaning climate "
                      f"where furnaces and heat pumps run more hours than central AC.")
        elif ratio > 0.7:
            opener = (f"{city_e} sits in a balanced climate band: NOAA records <strong>{hdd:,} heating "
                      f"degree days and {cdd:,} cooling degree days</strong> annually, meaning HVAC "
                      f"systems work both seasons roughly equally.")
        else:
            opener = (f"{city_e}'s climate is cooling-dominated: NOAA records <strong>{cdd:,} annual "
                      f"cooling degree days</strong> against just {hdd:,} heating degree days &mdash; "
                      f"AC capacity and refrigerant integrity matter more than furnace size here.")
    elif cdd is not None and hdd is None:
        opener = (f"{city_e}'s climate is so warm that NOAA's 1991-2020 Climate Normals record no "
                  f"meaningful heating load. Cooling degree days run <strong>{cdd:,} per year</strong>, "
                  f"placing {city_e} among the most cooling-intensive markets in the U.S.")
    elif hdd is not None:
        opener = (f"NOAA's 1991-2020 Climate Normals record <strong>{hdd:,} annual heating degree days</strong> "
                  f"for {city_e} &mdash; a heating load HVAC equipment must be sized to handle.")
    else:
        opener = (f"NOAA's 1991-2020 Climate Normals capture {city_e}'s long-run heating and cooling load.")

    # Sentence 2 — temperature extremes
    extremes_parts = []
    if days_90 is not None and days_90 >= 1:
        if days_90 > 100:
            extremes_parts.append(f"{days_90} days per year exceed 90&deg;F &mdash; sustained AC duty cycles "
                                  f"that drive compressor wear and refrigerant stress")
        elif days_90 > 30:
            extremes_parts.append(f"{days_90} days per year top 90&deg;F, putting consistent strain on "
                                  f"central AC equipment through summer")
        else:
            extremes_parts.append(f"{days_90} days per year hit 90&deg;F &mdash; short hot stretches "
                                  f"rather than sustained heat")
    if days_32 is not None and days_32 >= 1:
        if days_32 > 120:
            extremes_parts.append(f"{days_32} days drop below freezing &mdash; cold-climate heat pump "
                                  f"capacity matters here, and gas furnace reliability is non-negotiable")
        elif days_32 > 30:
            extremes_parts.append(f"{days_32} days fall below 32&deg;F, requiring competent heating "
                                  f"capacity through a real winter season")
        else:
            extremes_parts.append(f"only {days_32} days dip below freezing, so cold-climate equipment "
                                  f"specs aren't a primary concern")
    extremes = (" In an average year, " + " and ".join(extremes_parts) + ".") if extremes_parts else ""

    # Sentence 3 — snowfall
    snow_part = ""
    if snow is not None:
        if snow > 80:
            snow_part = (f" Annual snowfall averages <strong>{snow}&quot; per year</strong> &mdash; outdoor "
                         f"unit ice-pack damage and snow-load on heat pump coils are real spring service drivers.")
        elif snow > 30:
            snow_part = (f" Annual snowfall averages {snow}&quot; per year, enough that outdoor units need "
                         f"clearance from snowdrifts and gutter runoff.")
        elif snow > 5:
            snow_part = (f" Snowfall is modest at {snow}&quot; per year &mdash; occasional accumulation "
                         f"rather than a structural design factor.")
        elif snow > 0:
            snow_part = (f" Snowfall is rare at just {snow}&quot; in a typical year.")
        # snow == 0: omit (already implied by climate)

    paragraph = opener + extremes + snow_part

    return f'''
        <!-- Phase 3: Climate by the Numbers (NOAA NCEI 1991-2020 Normals) -->
        <div class="city-services" id="climate-numbers" style="margin-bottom: 40px;">
          <h2>{city_e} Climate by the Numbers</h2>
          <p style="font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);">{paragraph}</p>
        </div>
'''


def housing_stock_html(c):
    """Phase 3: render Census ACS housing stock data as HVAC-relevant prose.

    Drives ~200 words tying median home age + heating fuel mix + ownership rate
    to HVAC implications (older systems, fuel availability, owner motivation).
    Sentence variants chosen by data ranges so similar cities still differ.

    Source: Census ACS 5-year 2019-2023 (verified 2026-05-09).
    """
    city_e = escape(c['city'])
    state_full = escape(STATE_FULL.get(c['state'], c['state']))
    yr = c.get('acs_median_home_year')
    pct_gas = c.get('acs_heating_fuel_pct_gas')
    pct_elec = c.get('acs_heating_fuel_pct_electric')
    owner_pct = c.get('acs_owner_occupied_pct')
    home_value = c.get('acs_median_home_value')

    if not yr and not pct_gas and not pct_elec and not owner_pct and not home_value:
        return ""

    # Sentence 1 — housing age, by year built bucket
    age_part = ""
    if yr:
        age_now = 2026 - yr
        if yr < 1960:
            age_part = (f"{city_e}'s housing stock is among the oldest in {state_full}: per the U.S. Census ACS "
                        f"5-year estimates, the median home was built in <strong>{yr}</strong> &mdash; "
                        f"{age_now} years old on average. Original ductwork from this era often "
                        f"leaks 25-35% of conditioned air through joint failures, and many homes still "
                        f"have undersized electrical panels.")
        elif yr < 1980:
            age_part = (f"The median {city_e} home was built in <strong>{yr}</strong> per the U.S. Census ACS "
                        f"5-year estimates &mdash; about {age_now} years old. Homes from this era typically "
                        f"have second-generation HVAC equipment by now, and replacement ductwork sealing is "
                        f"often the highest-ROI HVAC upgrade.")
        elif yr < 2000:
            age_part = (f"{city_e}'s median home was built in <strong>{yr}</strong> &mdash; about {age_now} "
                        f"years old per the U.S. Census ACS 5-year estimates. Original HVAC equipment from "
                        f"this era is at or beyond expected lifespan; replacements are a normal expense.")
        else:
            age_part = (f"{city_e} has relatively new housing stock: the median home was built in "
                        f"<strong>{yr}</strong> per the U.S. Census ACS 5-year estimates, just {age_now} years old. "
                        f"Original equipment is usually still serviceable, with first-generation replacements "
                        f"on the horizon.")

    # Sentence 2 — heating fuel mix
    fuel_part = ""
    if pct_gas is not None and pct_elec is not None:
        if pct_gas > 70:
            fuel_part = (f" Natural gas heat dominates: <strong>{pct_gas}%</strong> of homes burn utility gas "
                         f"and only {pct_elec}% heat with electricity. Heat pump conversions face cheap-gas "
                         f"economics that often favor staying with a high-efficiency furnace.")
        elif pct_elec > 70:
            fuel_part = (f" Electricity dominates home heating at <strong>{pct_elec}%</strong> of households, "
                         f"with only {pct_gas}% on utility gas. Heat pump installations meet existing "
                         f"infrastructure; the upgrade path is straightforward.")
        elif pct_gas > pct_elec:
            fuel_part = (f" Heating fuel splits {pct_gas}% utility gas to {pct_elec}% electricity &mdash; a "
                         f"mixed-market where heat pump conversion economics depend on local utility rates "
                         f"and individual home setups.")
        else:
            fuel_part = (f" Heating fuel splits {pct_elec}% electricity to {pct_gas}% utility gas, with "
                         f"electric and gas roughly comparable across the housing stock.")

    # Sentence 3 — ownership + value
    ownership_part = ""
    if owner_pct is not None:
        if owner_pct > 65:
            ownership_part = (f" Owner-occupancy is high at <strong>{owner_pct}%</strong>")
        elif owner_pct > 45:
            ownership_part = (f" Owner-occupancy runs {owner_pct}%")
        else:
            ownership_part = (f" Owner-occupancy is low at <strong>{owner_pct}%</strong>")
    if home_value:
        if home_value < 200000:
            value_word = "modest"
        elif home_value < 500000:
            value_word = "mid-range"
        elif home_value < 1000000:
            value_word = "elevated"
        else:
            value_word = "high"
        if ownership_part:
            ownership_part += f" with a {value_word} median home value of <strong>${home_value:,}</strong>"
        else:
            ownership_part = f" Median home value runs <strong>${home_value:,}</strong>"
        # Pivot to HVAC implication
        if home_value > 500000:
            ownership_part += " &mdash; homeowners here typically invest in higher-efficiency replacements rather than minimum-spec equipment."
        elif home_value < 200000:
            ownership_part += " &mdash; replacement budgets tend to favor reliable mid-tier equipment over premium-tier upgrades."
        else:
            ownership_part += "."
    elif ownership_part:
        ownership_part += "."

    paragraph = (age_part + fuel_part + ownership_part).strip()
    if not paragraph:
        return ""

    return f'''
        <!-- Phase 3: Housing Stock & HVAC Considerations (Census ACS 2019-2023) -->
        <div class="city-services" id="housing-stock" style="margin-bottom: 40px;">
          <h2>{city_e} Housing Stock &amp; HVAC Considerations</h2>
          <p style="font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);">{paragraph}</p>
        </div>
'''


def operating_cost_snapshot_html(c):
    """Phase 3: render a per-city heat-pump-vs-furnace operating cost snapshot
    using EIA state rates + NOAA climate load.

    Math (transparent so a homeowner can verify):
      1 MMBtu = 293.07 kWh; 1 MMBtu = 10 therms
      Cost per million BTU of DELIVERED heat:
        - Heat pump (COP 3.0): elec_rate * 293.07 / 3.0   ($/MMBtu)
        - Gas furnace (AFUE 96%): gas_rate * 10 / 0.96    ($/MMBtu)
      Annual heating cost for a 1,500 sqft home (~50 MMBtu/yr at HDD 4500):
        scaled by city HDD ÷ 4500.

    Source: EIA Electric Power Monthly + Natural Gas Annual (verified 2026-02 / 2024).
    """
    city_e = escape(c['city'])
    elec = c.get('eia_state_elec_rate')   # $/kWh
    gas = c.get('eia_state_gas_rate')      # $/therm
    hdd = c.get('noaa_hdd_annual')

    if not (elec and gas):
        return ""

    # Cost per million BTU of delivered heat
    hp_per_mbtu = round(elec * 293.07 / 3.0, 2)
    gas_per_mbtu = round(gas * 10 / 0.96, 2)

    # Wins/loses framing
    winner_word = "heat pump" if hp_per_mbtu < gas_per_mbtu else "gas furnace"
    margin_pct = abs(hp_per_mbtu - gas_per_mbtu) / max(hp_per_mbtu, gas_per_mbtu) * 100

    # Annual cost estimate if we have HDD
    annual_part = ""
    if hdd:
        # Assume 50 MMBtu/year for ~1500 sqft home in a 4500 HDD climate; scale by city HDD
        load_mmbtu = round(50 * (hdd / 4500), 1)
        hp_annual = round(hp_per_mbtu * load_mmbtu / 50) * 50  # round to nearest $50
        gas_annual = round(gas_per_mbtu * load_mmbtu / 50) * 50
        annual_part = (f" For a typical 1,500 sqft home in {city_e}'s climate ({hdd:,} HDD), "
                       f"that maps to roughly <strong>${hp_annual:,}/year</strong> for a heat pump "
                       f"vs <strong>${gas_annual:,}/year</strong> for a gas furnace at current rates &mdash; "
                       f"actual usage varies with insulation, thermostat habits, and equipment efficiency.")

    # Headline
    if margin_pct < 8:
        verdict = (f"At {city_e}'s state-average rates ({elec*100:.1f}&cent;/kWh electricity, "
                   f"${gas:.2f}/therm natural gas), heat pump and gas furnace operating costs run "
                   f"close: <strong>${hp_per_mbtu}/MMBtu for the heat pump vs ${gas_per_mbtu}/MMBtu "
                   f"for the gas furnace</strong>, a {margin_pct:.0f}% gap.")
    else:
        verdict = (f"At {city_e}'s state-average rates ({elec*100:.1f}&cent;/kWh electricity, "
                   f"${gas:.2f}/therm natural gas), the <strong>{winner_word} wins on operating cost</strong>: "
                   f"${hp_per_mbtu}/MMBtu for the heat pump vs ${gas_per_mbtu}/MMBtu for the gas furnace, "
                   f"a {margin_pct:.0f}% margin.")

    return f'''
        <!-- Phase 3: Local Operating Cost Snapshot (EIA + NOAA, transparent math) -->
        <div class="city-services" id="operating-cost" style="margin-bottom: 40px;">
          <h2>Heating Operating Cost in {city_e}: Heat Pump vs Gas Furnace</h2>
          <p style="font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);">{verdict}{annual_part} The math, transparently: 1 MMBtu = 293.07 kWh = 10 therms. Heat pump cost-per-MMBtu = (electricity rate &times; 293.07) &divide; COP; gas furnace cost-per-MMBtu = (gas rate &times; 10) &divide; AFUE. Modern heat pumps achieve a coefficient of performance (COP) of 2.5 to 4 across most U.S. climates per the <a href="https://www.energy.gov/energysaver/heat-pump-systems" target="_blank" rel="noopener noreferrer" style="color: var(--orange-dark); font-weight: 600;">U.S. Department of Energy</a>; modern gas furnaces hit 95-98% AFUE.</p>
        </div>
'''


def local_context_section_html(c):
    """Phase 6: render the researched factual block, if populated.

    `local_context_html` from xlsx is pre-authored HTML (paragraphs with inline
    citations). Returns empty string when the column is empty, so pages that
    haven't been researched yet render unchanged.
    """
    raw = (c.get('local_context_html') or '').strip()
    if not raw:
        return ""
    # Normalize legal-section symbol — some fonts render &sect; oddly.
    # Spell out "Section" so it reads cleanly on every browser.
    raw = raw.replace('&sect; ', 'Section ').replace('&sect;', 'Section ')
    city_e = escape(c['city'])
    st_e = escape(c['state'])
    return f'''
        <!-- Phase 6: Researched per-city factual block (primary-source citations) -->
        <div class="city-services" id="local-data" style="margin-bottom: 40px;">
          <h2>HVAC in {city_e}, {st_e}: Local Data &amp; Context</h2>
          {raw}
        </div>
'''


def sources_footer_html():
    """Sources & References block — authoritative federal sources only."""
    return '''
        <!-- Sources & References (E-E-A-T) -->
        <div class="city-sources">
          <h3>Sources &amp; References</h3>
          <ul>
            <li><a href="https://www.energy.gov/energysaver/heat-and-cool" target="_blank" rel="nofollow noopener">U.S. Department of Energy &mdash; Home Heating &amp; Cooling</a></li>
            <li><a href="https://www.energystar.gov/" target="_blank" rel="nofollow noopener">U.S. ENERGY STAR (EPA)</a></li>
            <li><a href="https://www.dsireusa.org/" target="_blank" rel="nofollow noopener">DSIRE Database of State Incentives</a></li>
            <li><a href="https://www.irs.gov/credits-deductions/energy-efficient-home-improvement-credit" target="_blank" rel="nofollow noopener">IRS Section 25C Energy Efficient Home Improvement Credit</a></li>
          </ul>
        </div>
'''


def also_serving_section(anchor_city, absorbed_cities_data):
    """Generate the 'Also Serving' section for metro anchor pages."""
    if not absorbed_cities_data:
        return ""

    city_names = ", ".join(escape(c['city']) for c in absorbed_cities_data)

    html = f'''
    <!-- Also Serving Metro Area -->
    <section class="section" style="padding: 0 0 48px;">
      <div class="container" style="max-width: 760px;">
        <div class="city-services">
          <h2>Also Serving the Greater {escape(anchor_city)} Metro Area</h2>
          <p style="font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);">Our HVAC referral network extends beyond {escape(anchor_city)} to cover the surrounding metro area, including {city_names}.</p>
'''
    for ac in absorbed_cities_data:
        neighborhoods = escape(ac['neighborhoods'])
        zips = escape(ac['zips'])
        permit = escape(ac['permit'])
        html += f'''          <div style="margin-top: 20px; padding: 16px; background: var(--gray-50); border-radius: 8px;">
            <h3 style="font-size: 1.15rem; color: var(--navy); margin-bottom: 8px;">{escape(ac['city'])}, {escape(ac['state'])}</h3>
            <p style="font-size: 0.95rem; line-height: 1.75; color: var(--gray-700);">Neighborhoods: <strong>{neighborhoods}</strong>. ZIP codes served: <strong>{zips}</strong>. Local permits through <strong>{permit}</strong>.</p>
          </div>
'''
    html += '''        </div>
      </div>
    </section>
'''
    return html


# ============================================================
# CLIMATE-AWARE COST HELPERS
# ============================================================

def cost_context_sentence(city_e, ac_e, furnace_e, climate_type):
    """Cost sentence for local context paragraph — climate-aware."""
    profile = CLIMATE_PROFILES[climate_type]
    fl = profile['furnace_label']
    if climate_type == "tropical":
        return (f"If your system is beyond repair, a standard AC replacement in the "
                f"{city_e} area typically runs <strong>{ac_e}</strong>, and "
                f"{fl} average <strong>{furnace_e}</strong>.")
    elif profile['cost_order'] == "furnace_first":
        return (f"If your system is beyond repair, a furnace replacement in the "
                f"{city_e} area typically runs <strong>{furnace_e}</strong>, and "
                f"AC replacements average <strong>{ac_e}</strong>.")
    else:
        return (f"If your system is beyond repair, a standard AC replacement in the "
                f"{city_e} area typically runs <strong>{ac_e}</strong>, and "
                f"{fl} average <strong>{furnace_e}</strong>.")


def schema_cost_faq_text(city, ac_cost, furnace_cost, climate_type):
    """Schema-safe (no HTML) cost FAQ answer text."""
    profile = CLIMATE_PROFILES[climate_type]
    fl = profile['furnace_label']
    base = "Costs vary based on system size, efficiency rating, and installation complexity."
    if climate_type == "tropical":
        return (f"A standard AC replacement in {city} typically costs {ac_cost}, and "
                f"{fl} run {furnace_cost}. {base}")
    elif profile['cost_order'] == "furnace_first":
        return (f"A furnace replacement in {city} typically costs {furnace_cost}, and "
                f"AC replacements run {ac_cost}. {base}")
    else:
        return (f"A standard AC replacement in {city} typically costs {ac_cost}, and "
                f"{fl} run {furnace_cost}. {base}")


def faq_cost_answer_html(city_e, ac_e, furnace_e, seer_e, state_e, climate_type):
    """HTML cost FAQ answer — climate-aware."""
    profile = CLIMATE_PROFILES[climate_type]
    fl = profile['furnace_label']
    seer_note = f" In {state_e}, new AC units must meet a minimum <strong>{seer_e}</strong> rating."
    if climate_type == "tropical":
        return (f"A standard AC replacement in {city_e} typically costs <strong>{ac_e}</strong>, "
                f"and {fl} run <strong>{furnace_e}</strong>. "
                f"Costs vary based on system size, efficiency rating, and installation complexity."
                f"{seer_note}")
    elif profile['cost_order'] == "furnace_first":
        return (f"A furnace replacement in {city_e} typically costs <strong>{furnace_e}</strong>, "
                f"and AC replacements run <strong>{ac_e}</strong>. "
                f"Costs vary based on system size, efficiency rating, and installation complexity."
                f"{seer_note}")
    else:
        return (f"A standard AC replacement in {city_e} typically costs <strong>{ac_e}</strong>, "
                f"and {fl} run <strong>{furnace_e}</strong>. "
                f"Costs vary based on system size, efficiency rating, and installation complexity."
                f"{seer_note}")


# ============================================================
# MAIN PAGE TEMPLATE
# ============================================================

def generate_page(c, climate_type, nearby, absorbed_data, all_cities_lookup, state_info=None, reviewed_iso=None, reviewed_readable=None):
    """Generate complete HTML for a city page."""
    profile = CLIMATE_PROFILES[climate_type]
    city = c['city']
    st = c['state']
    state_full = STATE_FULL.get(st, st)
    slug = city_slug(city, st)
    s_slug = state_slug(st)
    pop = c['pop']
    pop_str = f"{pop:,}"
    neighborhoods = c['neighborhoods']
    zips = c['zips']
    ac_cost = c['ac_cost']
    furnace_cost = c['furnace_cost']
    permit = c['permit']
    utility = c['utility']
    seer = c['seer']
    rebates = c['rebates']
    license_req = c['license']
    climate_zone = c['climate']

    # Escaped versions
    e_city = escape(city)
    e_st = escape(st)
    e_state = escape(state_full)
    e_neighborhoods = escape(neighborhoods)
    e_zips = escape(zips)
    e_ac = escape(ac_cost)
    e_furnace = escape(furnace_cost)
    e_permit = escape(permit)
    e_utility = escape(utility)
    e_seer = escape(seer)
    e_rebates = escape(rebates)
    e_license = escape(license_req)
    e_climate = escape(climate_zone)

    # Title (max 60 chars — Google truncates beyond this) — climate-aware
    # len check unescapes &amp; so it counts as 1 char (matches audit_page.py)
    title_svc = profile['title_service']
    title_text = f"24/7 {title_svc} in {e_city}, {e_st} | Cool Call Pro"
    if len(title_text.replace("&amp;", "&")) > 60:
        title_text = f"HVAC Repair in {e_city}, {e_st} | Cool Call Pro"

    # Meta description (max 155 chars — Google truncates beyond this).
    # Template chosen so worst-case inputs still fit: longest meta_svc (~50 chars)
    # + longest city (~20 chars) + 2-char state abbr + 68-char fixed = ~140 chars.
    meta_svc = profile['meta_desc_service']
    meta_desc = f"Need {meta_svc} in {e_city}, {e_st}? Cool Call Pro connects you 24/7. Call (844) 582-1795."

    # Climate H2 and paragraph
    climate_h2, climate_para = climate_h2_and_paragraph(c, climate_type)

    # Services
    svcs = services_list(city, climate_type)

    # Climate-aware cost strings
    cost_sentence = cost_context_sentence(e_city, e_ac, e_furnace, climate_type)
    schema_cost_text = schema_cost_faq_text(city, ac_cost, furnace_cost, climate_type)
    faq_cost_html = faq_cost_answer_html(e_city, e_ac, e_furnace, e_seer, e_state, climate_type)
    permit_faq_q = profile['permit_faq_q']
    cost_faq_q = profile['cost_faq_q']

    # Licensing
    lic_h2, lic_para = licensing_paragraph(c, state_info)

    # Page meta (byline + last reviewed) — defaults to today if caller didn't pass
    if not reviewed_iso or not reviewed_readable:
        today = date.today()
        reviewed_iso = today.isoformat()
        reviewed_readable = today.strftime("%B %-d, %Y") if os.name != 'nt' else today.strftime("%B %#d, %Y")
    meta_block = page_meta_html(reviewed_iso, reviewed_readable)
    neighborhoods_block = neighborhoods_section_html(c)
    local_context_block = local_context_section_html(c)
    # Phase 3 / uniqueness uplift: new data-driven blocks. Each one renders
    # only if the underlying xlsx fields are populated; degrades gracefully
    # to empty string for cities with sentinel/missing values.
    climate_numbers_block = climate_by_numbers_html(c)
    housing_stock_block = housing_stock_html(c)
    operating_cost_block = operating_cost_snapshot_html(c)
    sources_block = sources_footer_html()
    license_faq_html = license_faq_sentence(c, html=True)
    license_faq_plain = license_faq_sentence(c, html=False).replace('"', '\\"')

    # Rebates
    reb_h2, reb_para = rebates_paragraph(c)

    # Per-city Wikimedia photo + attribution caption (returns "" if missing)
    figure_block = location_figure_html("city", slug, f"{city}, {st}", in_subdir=True)

    # Also Serving
    also_serving_html = also_serving_section(city, absorbed_data)

    # Nearby cities
    nearby_html = ""
    for nc_name, nc_st, nc_slug in nearby:
        nearby_html += f'            <li><a href="{nc_slug}">{escape(nc_name)}, {escape(nc_st)}</a></li>\n'
    nearby_html += f'            <li><a href="{s_slug}">{e_state} HVAC Guide &#8594;</a></li>\n'
    nearby_html += f'            <li><a href="../locations">View all HVAC service areas &#8594;</a></li>\n'

    # Schema areaServed
    area_served_schema = f'''"areaServed": {{
            "@type": "City",
            "name": "{city}",
            "containedInPlace": {{
              "@type": "State",
              "name": "{state_full}"
            }}
          }}'''

    if absorbed_data:
        areas = [f'''{{
            "@type": "City",
            "name": "{city}",
            "containedInPlace": {{ "@type": "State", "name": "{state_full}" }}
          }}''']
        for ac in absorbed_data:
            areas.append(f'''{{
            "@type": "City",
            "name": "{ac['city']}",
            "containedInPlace": {{ "@type": "State", "name": "{state_full}" }}
          }}''')
        area_served_schema = '"areaServed": [\n          ' + ',\n          '.join(areas) + '\n          ]'

    # ZIP coverage for FAQ (include absorbed ZIPs if metro anchor)
    all_zips = zips
    if absorbed_data:
        for ac in absorbed_data:
            all_zips += ", " + ac['zips']
        # Just show first 5-7 representative ZIPs
        zip_list = [z.strip() for z in all_zips.split(",")]
        all_zips = ", ".join(zip_list[:7])

    # Service schema — first 5 ZIPs as postalCode array (all from xlsx)
    service_zip_list = [z.strip() for z in zips.split(",") if z.strip()][:5]
    service_zip_json = ", ".join(f'"{z}"' for z in service_zip_list)
    service_schema = f''',
      {{
        "@type": "Service",
        "serviceType": "HVAC Repair and Installation Referral",
        "provider": {{
          "@type": "Organization",
          "name": "Cool Call Pro",
          "telephone": "+1-844-582-1795",
          "url": "https://coolcallpro.com"
        }},
        "areaServed": {{
          "@type": "City",
          "name": "{city}",
          "containedInPlace": {{ "@type": "State", "name": "{state_full}" }},
          "postalCode": [{service_zip_json}]
        }},
        "audience": {{ "@type": "Audience", "audienceType": "Homeowners" }}
      }}'''

    # ============================================================
    # HTML OUTPUT
    # ============================================================

    html = f'''<!DOCTYPE html>
<html lang="en">

<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{title_text}</title>
  <meta name="description"
    content="{meta_desc}" />
  <link rel="canonical" href="https://coolcallpro.com/locations/{slug}" />

  <!-- Open Graph -->
  <meta property="og:type" content="website" />
  <meta property="og:title" content="{title_text}" />
  <meta property="og:description" content="Connect with independent HVAC professionals in {e_city}, {e_state}. {profile['og_service']}" />
  <meta property="og:url" content="https://coolcallpro.com/locations/{slug}" />
  <meta property="og:site_name" content="Cool Call Pro" />
  <meta property="og:image" content="https://coolcallpro.com/images/og-homepage.webp" />

  <!-- Twitter Card -->
  <meta name="twitter:card" content="summary_large_image" />
  <meta name="twitter:title" content="24/7 {title_svc} in {e_city}, {e_st}" />
  <meta name="twitter:description" content="Connect with independent HVAC professionals in {e_city}, {e_state}." />
  <meta name="twitter:image" content="https://coolcallpro.com/images/og-homepage.webp" />

  <!-- Favicon -->
  <link rel="icon" href="/favicon.ico" type="image/x-icon" />
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />

  <link rel="stylesheet" href="../css/style.min.css" />
  <link rel="preload" href="/fonts/inter-latin.woff2" as="font" type="font/woff2" crossorigin>
  <link rel="preload" href="/fonts/outfit-latin.woff2" as="font" type="font/woff2" crossorigin>

  <!-- Google tag (gtag.js) - GA4 — deferred until first interaction for LCP -->
  <script>
    (function() {{
      var loaded = false;
      function loadGA() {{
        if (loaded) return;
        loaded = true;
        var s = document.createElement('script');
        s.src = 'https://www.googletagmanager.com/gtag/js?id=G-WD0ND0K60Q';
        s.async = true;
        document.head.appendChild(s);
        s.onload = function() {{
          window.dataLayer = window.dataLayer || [];
          function gtag(){{dataLayer.push(arguments);}}
          window.gtag = gtag;
          gtag('js', new Date());
          gtag('config', 'G-WD0ND0K60Q');
        }};
      }}
      ['click', 'scroll', 'touchstart', 'keydown'].forEach(function(evt) {{
        window.addEventListener(evt, loadGA, {{ once: true, passive: true }});
      }});
      setTimeout(loadGA, 4000);
    }})();
  </script>

  <!-- Structured Data: BreadcrumbList + FAQPage + Organization -->
  <script type="application/ld+json">
  {{
    "@context": "https://schema.org",
    "@graph": [
      {{
        "@type": "BreadcrumbList",
        "itemListElement": [
          {{"@type": "ListItem", "position": 1, "name": "Home", "item": "https://coolcallpro.com/"}},
          {{"@type": "ListItem", "position": 2, "name": "Locations", "item": "https://coolcallpro.com/locations"}},
          {{"@type": "ListItem", "position": 3, "name": "{state_full}", "item": "https://coolcallpro.com/locations/{s_slug}"}},
          {{"@type": "ListItem", "position": 4, "name": "{city}, {st}", "item": "https://coolcallpro.com/locations/{slug}"}}
        ]
      }},
      {{
        "@type": "FAQPage",
        "mainEntity": [
          {{
            "@type": "Question",
            "name": "{permit_faq_q} in {city}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "Yes, ensure your contractor files a mechanical permit with the {permit}."
            }}
          }},
          {{
            "@type": "Question",
            "name": "Are there HVAC rebates in {city}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "{city} homeowners served by {utility} may qualify for rebates through {rebates}. The federal Section 25C tax credit was terminated for installations placed in service after Dec 31, 2025 (OBBBA, P.L. 119-21); check current state HEAR rebates and utility programs for what's still available in 2026."
            }}
          }},
          {{
            "@type": "Question",
            "name": "What ZIP codes do you serve in {city}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "Our network covers {city} and surrounding areas including {zips}. Call (844) 582-1795 to verify service availability for your specific ZIP code."
            }}
          }},
          {{
            "@type": "Question",
            "name": "{cost_faq_q} in {city}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "{schema_cost_text}"
            }}
          }},
          {{
            "@type": "Question",
            "name": "What HVAC license is required in {state_full}?",
            "acceptedAnswer": {{
              "@type": "Answer",
              "text": "{license_faq_plain} Always verify your contractor credentials before authorizing work."
            }}
          }}
        ]
      }},
      {{
        "@type": "Organization",
        "@id": "https://coolcallpro.com/#organization",
        "name": "Cool Call Pro",
        "url": "https://coolcallpro.com",
        "description": "24/7 referral service connecting homeowners with independent HVAC professionals in {city}, {st}.",
        "contactPoint": {{
          "@type": "ContactPoint",
          "telephone": "+1-844-582-1795",
          "contactType": "customer service",
          {area_served_schema},
          "availableLanguage": "English"
        }}
      }}{service_schema}
    ]
  }}
  </script>

  <style>
    @font-face{{font-family:'Inter';font-style:normal;font-weight:400 900;font-display:swap;src:url('/fonts/inter-latin.woff2') format('woff2')}}@font-face{{font-family:'Outfit';font-style:normal;font-weight:400 800;font-display:swap;src:url('/fonts/outfit-latin.woff2') format('woff2')}}
    .city-hero {{
      background: linear-gradient(135deg, var(--navy-deep) 0%, var(--navy) 50%, var(--navy-mid) 100%);
      background-size: cover;
      background-position: center;
      color: white;
      padding: 100px 0 80px;
      text-align: center;
    }}
    .city-hero h1 {{
      font-family: var(--font-display);
      font-size: 2.6rem;
      font-weight: 800;
      line-height: 1.2;
      margin-bottom: 20px;
      letter-spacing: -0.02em;
    }}
    .city-hero p {{
      font-size: 1.15rem;
      max-width: 620px;
      margin: 0 auto;
      opacity: 0.9;
      line-height: 1.7;
    }}
    .city-context {{
      font-size: 1.1rem;
      line-height: 1.85;
      color: var(--gray-700);
      max-width: 760px;
      margin: 0 auto 40px;
    }}
    .city-services {{
      max-width: 760px;
      margin: 0 auto 48px;
    }}
    .city-services h2 {{
      font-size: 1.6rem;
      margin-bottom: 16px;
      color: var(--navy);
    }}
    .city-services ul {{
      list-style: none;
      padding: 0;
    }}
    .city-services ul li {{
      padding: 12px 0 12px 28px;
      position: relative;
      font-size: 1.05rem;
      color: var(--gray-700);
      border-bottom: 1px solid var(--gray-100);
    }}
    .city-services ul li::before {{
      content: "\\2714";
      color: var(--orange-dark);
      font-weight: 700;
      position: absolute;
      left: 0;
      top: 12px;
    }}
    .city-faq {{
      max-width: 760px;
      margin: 0 auto 48px;
    }}
    .city-faq h2 {{
      font-size: 1.6rem;
      margin-bottom: 20px;
      color: var(--navy);
    }}
    @media (max-width: 768px) {{
      .city-hero {{
        padding: 80px 0 60px;
      }}
      .city-hero h1 {{
        font-size: 1.8rem;
      }}
    }}
  </style>
</head>

<body>

  <!-- Skip to Content -->
  <a href="#main-content" class="skip-link">Skip to main content</a>

  <!-- Top Bar -->
  <div class="topbar">
    <div class="container topbar-inner" style="justify-content: center; padding: 10px; font-weight: 600;">
      <span class="topbar-text">24/7 Referral Service &#8212; Connecting Homeowners with Independent HVAC Professionals</span>
    </div>
  </div>

  <!-- Header -->
  <header class="header" id="header">
    <nav class="nav container" role="navigation" aria-label="Main navigation">
      <a href="../index" class="logo">
        <span class="logo-icon">&#10052;&#65039;</span>
        <span class="logo-text">CoolCall<span class="logo-accent">Pro</span></span>
      </a>
      <ul class="nav-links" id="navLinks">
        <li><a href="../emergency" class="nav-link emergency-link">&#128680; Emergency Service</a></li>
        <li><a href="../costs" class="nav-link">Cost Guide</a></li>
        <li><a href="../safety" class="nav-link">Safety Tips</a></li>
        <li><a href="../locations" class="nav-link active">&#128205; Locations</a></li>
        <li><a href="../articles" class="nav-link">Articles</a></li>
        <li><a href="../about" class="nav-link">About</a></li>
        <li><a href="../contact" class="nav-link">Contact</a></li>
      </ul>
      <a href="tel:+18445821795" class="btn btn-cta nav-cta btn-vibrate" id="headerNavCta" aria-hidden="true"
        tabindex="-1" style="opacity: 0; pointer-events: none; transition: opacity 0.3s ease;"><span
          class="phone-icon">&#128222;</span>
        <span class="hide-on-mobile">(844) 582-1795</span></a>
      <button class="hamburger" id="hamburger" aria-label="Menu">&#9776;</button>
    </nav>
  </header>

  <!-- Main Content -->
  <main id="main-content">

    <!-- Hero Section -->
    <section class="city-hero" id="hero">
      <div class="container">
        <span class="section-tag" style="background: rgba(255,255,255,0.1); color: #fff;">&#128205; {e_city}, {e_st}</span>
        <h1>{profile['h1_service']} in {e_city}, {e_st}</h1>
        <p>Connect with independent local HVAC professionals in {e_city}, {e_state}. {profile['hero_service']}</p>
        {meta_block}
        <div style="margin-top: 28px;">
          <a href="tel:+18445821795" class="btn btn-primary btn-lg btn-vibrate"><span class="phone-icon">&#128222;</span> Call Now &#8212; (844) 582-1795</a>
        </div>
        <div class="jump-links" style="margin-top: 16px;">
          <a href="#services" class="jump-link" style="color: rgba(255,255,255,0.8);">Services</a>
          <span class="jump-link-dot" style="color: rgba(255,255,255,0.5);">&#8226;</span>
          <a href="#faqs" class="jump-link" style="color: rgba(255,255,255,0.8);">FAQs</a>
        </div>
      </div>
    </section>

    <!-- Breadcrumb -->
    <div class="breadcrumb-nav">
      <div class="container">
        <nav aria-label="Breadcrumb">
          <ol class="breadcrumb-list">
            <li><a href="../index">Home</a></li>
            <li><a href="../locations">Locations</a></li>
            <li><a href="{s_slug}">{e_state}</a></li>
            <li aria-current="page">{e_city}, {e_st}</li>
          </ol>
        </nav>
      </div>
    </div>

    <!-- Local Context Paragraph -->
    <section class="section" style="padding: 48px 0 0;">
      <div class="container">
        <div class="city-context" style="max-width: 760px; margin: 0 auto 40px;">
          <p>Serving <strong>{pop_str}</strong> residents across <strong>{e_city}, {e_state}</strong>, Cool Call Pro connects homeowners in <strong>{e_neighborhoods}</strong> with independent 24/7 HVAC technicians. Our referral network covers <strong>{e_zips}</strong> and surrounding areas. {cost_sentence} Your contractor should pull mechanical permits through the <strong>{e_permit}</strong>.</p>
        </div>

        <!-- Phase 3 (2026-05-09): the original Climate Context section was
             8 climate-conditional templates — every hot-humid city got the
             same paragraph with city/temp slots filled in. That made
             same-cluster cities highly similar at the shingle level.
             The new climate_by_numbers / housing_stock / operating_cost
             blocks below replace it with NOAA + Census + EIA data-driven
             prose that varies city-by-city, not just climate-by-climate. -->
{climate_numbers_block}{housing_stock_block}{operating_cost_block}{figure_block}{local_context_block}{neighborhoods_block}
        <!-- Services List -->
        <div class="city-services" id="services">
          <h2>HVAC Services in {e_city}, {e_st}</h2>
          <ul>
'''

    for svc in svcs:
        html += f'              <li>{svc}</li>\n'

    html += f'''          </ul>
        </div>

        <!-- Licensing & Trust -->
        <div class="city-services" style="margin-bottom: 40px;">
          <h2>{lic_h2}</h2>
          <p style="font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);">{lic_para}</p>
        </div>

        <!-- Rebates & Incentives -->
        <div class="city-services" style="margin-bottom: 48px;">
          <h2>{reb_h2}</h2>
          <p style="font-size: 1.05rem; line-height: 1.85; color: var(--gray-700);">{reb_para}</p>
        </div>
      </div>
    </section>
{also_serving_html}
    <!-- FAQs -->
    <section class="section" id="faqs">
      <div class="container" style="max-width: 800px;">
        <div class="city-faq">
          <h2>Frequently Asked Questions &#8212; {e_city}, {e_st}</h2>
          <div class="faq-list" style="margin-top: 24px;">
            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>{permit_faq_q} in {e_city}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>Yes, ensure your contractor files a mechanical permit with the <strong>{e_permit}</strong>. Pulling the correct permits protects you as a homeowner and ensures work is inspected to code.</p>
                </div>
              </div>
            </div>

            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>Are there HVAC rebates in {e_city}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>Homeowners may qualify for savings through <strong>{e_utility}</strong>. Check with <strong>{e_rebates}</strong> for current offers. The federal Section 25C credit was terminated for installations after Dec 31, 2025 (OBBBA, P.L. 119-21); check current state and utility programs for 2026.</p>
                </div>
              </div>
            </div>

            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>What ZIP codes do you serve in {e_city}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>Our network covers {e_city} and surrounding areas including <strong>{escape(all_zips)}</strong>. Call <a href="tel:+18445821795" style="color: var(--orange-dark); font-weight: 600;">(844) 582-1795</a> to verify service availability for your specific ZIP code.</p>
                </div>
              </div>
            </div>

            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>{cost_faq_q} in {e_city}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>{faq_cost_html}</p>
                </div>
              </div>
            </div>

            <div class="faq-item">
              <button class="faq-q" aria-expanded="false">
                <span>What HVAC license is required in {e_state}?</span>
                <span class="faq-icon"></span>
              </button>
              <div class="faq-a">
                <div class="faq-a-inner">
                  <p>{license_faq_html} Always verify your contractor's credentials before authorizing work. For {e_city} residents, permits are filed through the <strong>{e_permit}</strong>.</p>
                </div>
              </div>
            </div>
          </div>

        </div>
      </div>
    </section>

    <!-- Nearby Service Areas -->
    <section class="section" style="padding: 32px 0 48px;">
      <div class="container" style="max-width: 800px;">
        <div class="city-services">
          <h2>Nearby HVAC Service Areas</h2>
          <ul>
{nearby_html}          </ul>
        </div>
{sources_block}      </div>
    </section>

  </main>

  <!-- Footer -->
  <footer class="footer">
    <div class="container">
      <div class="footer-grid">
        <div class="footer-brand">
          <a href="../index" class="logo"><span class="logo-icon">&#10052;&#65039;</span><span
              class="logo-text">CoolCall<span class="logo-accent">Pro</span></span></a>
          <p>Connecting homeowners with independent HVAC professionals across the US. Available 24/7 in many areas.</p>
          <div class="footer-contact">
            <a href="tel:+18445821795" class="footer-phone">(844) 582-1795</a>
            <span>24/7 Service Line</span>
          </div>
        </div>
        <div class="footer-col">
          <h3 class="footer-heading">Resources</h3>
          <ul>
            <li><a href="../costs">Cost Guide</a></li>
            <li><a href="../emergency">Emergency Service</a></li>
            <li><a href="../safety">Safety Tips</a></li>
            <li><a href="../articles">All Articles</a></li>
            <li><a href="../locations">Locations</a></li>
          </ul>
        </div>
        <div class="footer-col">
          <h3 class="footer-heading">Company</h3>
          <ul>
            <li><a href="../about">About Us</a></li>
            <li><a href="../contact">Contact</a></li>
            <li><a href="../author-gyanesh">Author</a></li>
            <li><a href="../privacy">Privacy Policy</a></li>
            <li><a href="../advertising-disclosure">Advertising Disclosure</a></li>
          </ul>
        </div>
        <div class="footer-col">
          <h3 class="footer-heading">Service Areas</h3>
          <ul class="service-areas">
            <li><a href="atlanta-ga">Atlanta, GA</a></li>
            <li><a href="chicago-il">Chicago, IL</a></li>
            <li><a href="dallas-tx">Dallas, TX</a></li>
            <li><a href="houston-tx">Houston, TX</a></li>
            <li><a href="phoenix-az">Phoenix, AZ</a></li>
            <li><a href="../locations">All 50 states &#8594;</a></li>
          </ul>
        </div>
      </div>
      <div class="footer-bottom">
        <p><strong style="color: rgba(255,255,255,0.82);">How this site works:</strong> We publish HVAC cost and
          troubleshooting information and may connect callers to independent service providers. We do not perform HVAC
          services ourselves. Pricing, availability, and response times vary by provider and location.</p>
        <p><strong style="color: rgba(255,255,255,0.82);">Editorial standards:</strong> Our content focuses on helpful
          troubleshooting tips and realistic pricing information. We avoid guarantees; actual pricing depends on the
          provider and situation. We clearly disclose compensation for referrals. We update content for clarity and
          accuracy when needed.</p>
        <p><strong style="color: rgba(255,255,255,0.82);">Disclaimer:</strong> Cool Call Pro is a free service to assist homeowners in connecting with local service providers. All contractors/providers are independent and Cool Call Pro does not warrant or guarantee any work performed. It is the responsibility of the homeowner to verify that the hired contractor furnishes the necessary license and insurance required for the work being performed. All persons depicted in a photo or video are actors or models and not contractors listed on Cool Call Pro.</p>
        <p><strong style="color: rgba(255,255,255,0.82);">Service availability:</strong> Same-day and 24/7 emergency services are subject to provider participation, location, technician availability, and demand. Availability is not guaranteed and may vary by market and appointment capacity.</p>
        <p>&copy; <span class="copyright-year">2026</span> Cool Call Pro. All rights reserved. &nbsp;&#183;&nbsp; <a href="../privacy">Privacy
            Policy</a> &nbsp;&#183;&nbsp; <a href="../terms">Terms of Use</a> &nbsp;&#183;&nbsp; <a
            href="../disclaimer">Disclaimer</a> &nbsp;&#183;&nbsp; <a href="../advertising-disclosure">Advertising
            Disclosure</a> &nbsp;&#183;&nbsp; <a href="../image-credits">Image Credits</a></p>
      </div>
    </div>
  </footer>

  <script src="../js/main.min.js" defer></script>

  <!-- Mobile Call Bar -->
  <div class="mobile-call-bar">
    <a href="tel:+18445821795"><svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor" style="width: 1.25em; height: 1.25em; position: relative; top: 0.15em;"><path d="M20.01 15.38c-1.23 0-2.42-.2-3.53-.56-.35-.12-.74-.03-1.01.24l-1.57 1.97c-2.83-1.35-5.48-3.9-6.89-6.83l1.95-1.66c.27-.28.35-.67.24-1.02-.37-1.11-.56-2.3-.56-3.53 0-.54-.45-.99-.99-.99H4.19C3.65 3 3 3.24 3 3.99 3 13.28 10.73 21 20.01 21c.71 0 1.01-.63 1.01-1.18v-3.45c0-.54-.45-.99-.99-.99z"/></svg> Call Now &#8212; (844) 582-1795</a>
  </div>
<script>document.querySelectorAll(".copyright-year").forEach(function(el){{el.textContent=new Date().getFullYear()}});</script>
</body>

</html>'''

    return html


# ============================================================
# MAIN
# ============================================================

def load_states():
    """Load state metadata (license body + URL + name) from states.xlsx, keyed by abbr."""
    wb = openpyxl.load_workbook(STATES_XLSX)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        abbr = data.get('State Abbreviation')
        if abbr:
            out[abbr] = data
    return out


def load_cities():
    """Load all cities from xlsx.

    Returns (cities_by_name, all_cities_list). `cities_by_name` keys by city
    name (first-loaded wins on duplicates like Portland OR vs Portland ME);
    `all_cities_list` contains every row so --force can regenerate both
    same-named cities.
    """
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active

    cities = {}
    all_cities = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pop = row[8]
        if isinstance(pop, str):
            pop = int(pop.replace(',', '').strip())
        elif pop is None:
            pop = 0
        else:
            pop = int(pop)

        city = {
            'city': row[0], 'state': row[1],
            'ac_cost': str(row[2] or ''), 'furnace_cost': str(row[3] or ''),
            'neighborhoods': str(row[4] or ''), 'zips': str(row[5] or ''),
            'permit': str(row[6] or ''), 'utility': str(row[7] or ''),
            'pop': pop,
            'summer_high': str(row[9] or ''),
            'winter_low': str(row[10] or ''),
            'seer': str(row[11] or ''), 'rebates': str(row[12] or ''),
            'license': str(row[13] or ''), 'climate': str(row[14] or ''),
            'published': row[15],
            # Phase 6: researched per-city factual block (column index 17).
            # Optional; empty string means generator renders no local-context
            # section, so un-researched cities keep their Phase 3 template.
            'local_context_html': str(row[17] or '') if len(row) > 17 else '',
            # Phase-3-uniqueness fields (added 2026-05-09 in NOAA + Census + EIA
            # fetcher commits). Columns 18-29. Stored as native types where
            # available (int/float) or None for cities with NCEI sentinels.
            # Drive new content blocks: "Climate by the Numbers", "Housing
            # Stock & HVAC Considerations", "Local Operating Cost Snapshot".
            'noaa_hdd_annual':              row[18] if len(row) > 18 else None,
            'noaa_cdd_annual':              row[19] if len(row) > 19 else None,
            'noaa_days_above_90':           row[20] if len(row) > 20 else None,
            'noaa_days_below_32':           row[21] if len(row) > 21 else None,
            'noaa_snowfall_inches':         row[22] if len(row) > 22 else None,
            'acs_median_home_year':         row[23] if len(row) > 23 else None,
            'acs_heating_fuel_pct_gas':     row[24] if len(row) > 24 else None,
            'acs_heating_fuel_pct_electric':row[25] if len(row) > 25 else None,
            'acs_owner_occupied_pct':       row[26] if len(row) > 26 else None,
            'acs_median_home_value':        row[27] if len(row) > 27 else None,
            'eia_state_elec_rate':          row[28] if len(row) > 28 else None,
            'eia_state_gas_rate':           row[29] if len(row) > 29 else None,
        }
        all_cities.append(city)
        if row[0] not in cities:
            cities[row[0]] = city

    return cities, all_cities


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Generate city pages v3")
    parser.add_argument('--batch', type=int, help="Batch number (1=high priority)")
    parser.add_argument('--city', type=str, help="Generate single city by name")
    parser.add_argument('--dry-run', action='store_true', help="Preview without writing")
    parser.add_argument('--force', action='store_true', help="Regenerate pages even if already published (used for template overhauls)")
    args = parser.parse_args()

    cities, all_cities = load_cities()
    states_meta = load_states()

    # Group by state for nearby logic
    from collections import defaultdict
    by_state = defaultdict(list)
    for c in all_cities:
        if c['city'] not in REMOVED_CITIES and c['city'] not in ABSORBED:
            by_state[c['state']].append(c)

    # Determine which cities to generate. In --force mode we iterate the full
    # all_cities list (covers same-named cities like Portland OR + Portland ME);
    # otherwise we iterate city names against the first-wins dict.
    force_targets = None  # list of city dicts when in --force mode

    if args.city:
        targets = [args.city]
    elif args.batch == 1:
        targets = BATCH_1
    elif args.batch == 2:
        targets = BATCH_2
    elif args.force:
        force_targets = [c for c in all_cities
                         if c['city'] not in REMOVED_CITIES
                         and c['city'] not in ABSORBED]
        targets = []  # unused in force mode
    else:
        targets = [c['city'] for c in all_cities
                   if c['city'] not in PUBLISHED_CITIES
                   and c['city'] not in REMOVED_CITIES
                   and c['city'] not in ABSORBED]

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    generated = 0
    skipped = 0
    errors = []

    # Build iteration list as (city_name, city_dict) pairs
    if force_targets is not None:
        iter_list = [(c['city'], c) for c in force_targets]
    else:
        iter_list = []
        for city_name in targets:
            if city_name not in cities:
                errors.append(f"City '{city_name}' not found in xlsx")
                continue
            iter_list.append((city_name, cities[city_name]))

    for city_name, c in iter_list:
        if city_name in PUBLISHED_CITIES and not args.force and not args.city:
            print(f"  SKIP (already published): {city_name}")
            skipped += 1
            continue

        if city_name in ABSORBED:
            print(f"  SKIP (absorbed into {ABSORBED[city_name]}): {city_name}")
            skipped += 1
            continue

        if city_name in REMOVED_CITIES:
            print(f"  SKIP (removed - too small): {city_name}")
            skipped += 1
            continue

        climate_type = get_climate_type(city_name, c['climate'])

        # Get absorbed cities data for metro anchors
        absorbed_data = []
        if city_name in ANCHOR_ABSORBS:
            for abs_name in ANCHOR_ABSORBS[city_name]:
                if abs_name in cities:
                    absorbed_data.append(cities[abs_name])

        # Get nearby cities
        nearby = get_nearby_cities(c, by_state, cities)

        # Generate HTML
        state_info = states_meta.get(c['state'])
        html = generate_page(c, climate_type, nearby, absorbed_data, cities, state_info=state_info)

        # Write file
        slug = city_slug(city_name, c['state'])
        filepath = os.path.join(OUTPUT_DIR, f"{slug}.html")

        if args.dry_run:
            print(f"  DRY RUN: Would write {filepath} ({len(html)} chars) | climate={climate_type} | absorbed={len(absorbed_data)}")
        else:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html)
            print(f"  GENERATED: {filepath} | climate={climate_type} | absorbed={len(absorbed_data)} cities")

        generated += 1

    print(f"\n{'='*50}")
    print(f"Generated: {generated} | Skipped: {skipped} | Errors: {len(errors)}")
    if errors:
        for e in errors:
            print(f"  ERROR: {e}")


if __name__ == "__main__":
    main()
